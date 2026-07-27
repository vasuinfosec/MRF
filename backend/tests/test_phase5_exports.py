"""
Phase 5 — Export/Download tests (PO PDF/Excel/Tally, MRF PDF/Excel, bulk exports,
validation 422 + missing_mandatory, ?force=1 bypass, audit logging, RBAC).

Runs against the public preview URL (EXPO_PUBLIC_BACKEND_URL) and uses dev-login
to obtain per-role session tokens.
"""
import io
import os
import time
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = (
    os.environ.get("EXPO_BACKEND_URL")
    or os.environ.get("EXPO_PUBLIC_BACKEND_URL")
    or "https://purchase-workflow-10.preview.emergentagent.com"
).rstrip("/")

ROLES = {
    "director":     {"email": "director@vasu.dev",     "name": "Director"},
    "gm":           {"email": "gm@vasu.dev",           "name": "GM"},
    "pm":           {"email": "pm@vasu.dev",           "name": "PM"},
    "purchase":     {"email": "purchase@vasu.dev",     "name": "Purchase"},
    "site_engineer":{"email": "siteeng@vasu.dev",      "name": "Site Eng"},
    "admin":        {"email": "admin@vasu.dev",        "name": "Admin"},
    "store":        {"email": "store@vasu.dev",        "name": "Store"},
}

TOKENS: dict = {}


def _login(role: str) -> str:
    if role in TOKENS:
        return TOKENS[role]
    body = {**ROLES[role], "role": role}
    r = requests.post(f"{BASE_URL}/api/auth/dev-login", json=body, timeout=20)
    assert r.status_code == 200, f"dev-login failed for {role}: {r.status_code} {r.text}"
    tok = r.json()["session_token"]
    TOKENS[role] = tok
    return tok


def _hdr(role: str) -> dict:
    return {"Authorization": f"Bearer {_login(role)}"}


# --------------------- fixtures -------------------------------------------
@pytest.fixture(scope="module")
def any_po_id():
    """Pick a PO that exists in the system."""
    _login("admin")
    # Ensure seed exists (idempotent)
    requests.post(f"{BASE_URL}/api/seed", headers=_hdr("admin"), timeout=30)
    r = requests.get(f"{BASE_URL}/api/po", headers=_hdr("purchase"), timeout=20)
    assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
    pos = r.json()
    assert pos, "No POs in system — cannot run PO export tests"
    return pos[0]["po_id"]


@pytest.fixture(scope="module")
def any_mrf_id():
    r = requests.get(f"{BASE_URL}/api/mrf", headers=_hdr("purchase"), timeout=20)
    assert r.status_code == 200
    mrfs = r.json()
    assert mrfs
    return mrfs[0]["mrf_id"]


# --------------------- Single-PO PDF --------------------------------------
class TestPOPDF:
    def test_po_pdf_returns_pdf_bytes(self, any_po_id):
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/pdf",
                         headers=_hdr("purchase"), timeout=30)
        # Either 200 with %PDF, or 422 validation_error (seed data may miss customer)
        if r.status_code == 422:
            pytest.skip(f"seed PO missing mandatory field: {r.json()}")
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        assert r.content[:4] == b"%PDF", "Response is not a PDF"
        assert len(r.content) > 2000, "PDF too small — likely broken"

    def test_po_pdf_force_bypass_when_validation_fails(self, any_po_id):
        # try with force=1 — should always return PDF
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/pdf?force=1",
                         headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"


# --------------------- Single-PO export dispatcher ------------------------
class TestPOExportDispatcher:
    def test_dispatcher_pdf(self, any_po_id):
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=pdf&force=1",
                         headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_dispatcher_excel_headers_and_content(self, any_po_id):
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=excel&force=1",
                         headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct or "excel" in ct, f"unexpected content-type: {ct}"
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header_row = [c.value for c in ws[1]]
        required_cols = [
            "PO#", "PO Date", "Customer ID", "Customer GSTIN",
            "Vendor Name", "Vendor GSTIN", "MRF Refs", "Customer PO Ref",
            "MAT UID", "VAR UID", "HSN/SAC", "Taxable Value",
            "CGST", "SGST", "IGST", "Line Total",
            "Authorised Signatory", "Approved By (history)",
        ]
        missing = [c for c in required_cols if c not in header_row]
        assert not missing, f"Excel missing columns: {missing}. Actual: {header_row}"

    def test_dispatcher_tally_missing_hsn_returns_422(self, any_po_id):
        # Seed data typically lacks HSN, so this SHOULD 422
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=tally",
                         headers=_hdr("purchase"), timeout=30)
        if r.status_code == 200:
            pytest.skip("This PO has HSN on all items — cannot test 422 path")
        assert r.status_code == 422, f"expected 422, got {r.status_code}: {r.text[:200]}"
        detail = r.json().get("detail") or {}
        assert detail.get("error") == "export_validation_failed"
        missing = detail.get("missing_mandatory") or []
        assert any("HSN" in m for m in missing), f"missing_mandatory has no HSN entry: {missing}"

    def test_dispatcher_tally_force_bypasses_validation(self, any_po_id):
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=tally&force=1",
                         headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.active.max_row >= 1

    def test_forced_tally_writes_audit_with_forced_true(self, any_po_id):
        # trigger a forced export
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=tally&force=1",
                         headers=_hdr("purchase"), timeout=30)
        if r.status_code != 200:
            pytest.skip("could not trigger forced tally export")
        time.sleep(0.5)
        a = requests.get(f"{BASE_URL}/api/audit?entity_id={any_po_id}",
                         headers=_hdr("director"), timeout=20)
        assert a.status_code == 200
        rows = a.json()
        tally_rows = [x for x in rows if x.get("action") == "export_tally"]
        assert tally_rows, "no export_tally audit rows found"
        forced_rows = [x for x in tally_rows if (x.get("details") or {}).get("forced") is True]
        assert forced_rows, f"no forced=true audit rows. Sample: {tally_rows[0] if tally_rows else None}"

    def test_invalid_format(self, any_po_id):
        r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=word",
                         headers=_hdr("purchase"), timeout=20)
        assert r.status_code == 400


# --------------------- MRF PDF + dispatcher -------------------------------
class TestMRFExport:
    def test_mrf_pdf(self, any_mrf_id):
        r = requests.get(f"{BASE_URL}/api/mrf/{any_mrf_id}/pdf?force=1",
                         headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        assert r.content[:4] == b"%PDF"
        assert len(r.content) > 1500

    def test_mrf_export_pdf_dispatcher(self, any_mrf_id):
        r = requests.get(f"{BASE_URL}/api/mrf/{any_mrf_id}/export?format=pdf&force=1",
                         headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    def test_mrf_export_excel_dispatcher(self, any_mrf_id):
        r = requests.get(f"{BASE_URL}/api/mrf/{any_mrf_id}/export?format=excel&force=1",
                         headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        headers = [c.value for c in wb.active[1]]
        for col in ("MRF#", "MAT UID", "VAR UID", "Make", "Model", "Qty Requested"):
            assert col in headers, f"MRF Excel missing '{col}'. Headers: {headers}"

    def test_mrf_export_tally_rejected(self, any_mrf_id):
        r = requests.get(f"{BASE_URL}/api/mrf/{any_mrf_id}/export?format=tally",
                         headers=_hdr("purchase"), timeout=20)
        assert r.status_code == 400


# --------------------- Bulk exports (RBAC) --------------------------------
class TestBulkExports:
    def test_bulk_mrf_purchase(self):
        r = requests.get(f"{BASE_URL}/api/export/mrf", headers=_hdr("purchase"), timeout=30)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        headers = [c.value for c in wb.active[1]]
        for col in ("MRF#", "Customer ID", "MAT UID"):
            assert col in headers

    @pytest.mark.parametrize("role", ["pm", "gm", "director", "admin"])
    def test_bulk_mrf_allowed_roles(self, role):
        r = requests.get(f"{BASE_URL}/api/export/mrf", headers=_hdr(role), timeout=30)
        assert r.status_code == 200, f"{role} got {r.status_code}"

    def test_bulk_mrf_denied_site_engineer(self):
        r = requests.get(f"{BASE_URL}/api/export/mrf",
                         headers=_hdr("site_engineer"), timeout=15)
        assert r.status_code == 403

    def test_bulk_po_denied_site_engineer(self):
        r = requests.get(f"{BASE_URL}/api/export/po",
                         headers=_hdr("site_engineer"), timeout=15)
        assert r.status_code == 403

    @pytest.mark.parametrize("role", ["purchase", "gm", "director", "admin"])
    def test_bulk_po_allowed_roles(self, role):
        r = requests.get(f"{BASE_URL}/api/export/po", headers=_hdr(role), timeout=30)
        assert r.status_code == 200, f"{role} got {r.status_code}"

    def test_bulk_po_pm_role(self):
        # Spec says pm should be allowed for /api/export/po
        r = requests.get(f"{BASE_URL}/api/export/po", headers=_hdr("pm"), timeout=30)
        assert r.status_code == 200, (
            f"Spec says pm should be allowed but got {r.status_code}: {r.text[:200]}"
        )

    def test_bulk_tally_denied_site_engineer(self):
        r = requests.get(f"{BASE_URL}/api/export/tally",
                         headers=_hdr("site_engineer"), timeout=15)
        assert r.status_code == 403

    def test_bulk_tally_denied_pm(self):
        r = requests.get(f"{BASE_URL}/api/export/tally",
                         headers=_hdr("pm"), timeout=15)
        assert r.status_code == 403

    @pytest.mark.parametrize("role", ["purchase", "director", "admin"])
    def test_bulk_tally_allowed_roles(self, role):
        r = requests.get(f"{BASE_URL}/api/export/tally?kind=purchase",
                         headers=_hdr(role), timeout=30)
        assert r.status_code == 200, f"{role} got {r.status_code}"


# --------------------- Audit logging --------------------------------------
class TestAuditLogging:
    def test_successful_export_increments_audit_count(self, any_po_id):
        # baseline audit count for this PO
        base = requests.get(f"{BASE_URL}/api/audit?entity_id={any_po_id}",
                            headers=_hdr("director"), timeout=15)
        assert base.status_code == 200
        before = len([r for r in base.json() if r.get("action") == "export_pdf"])
        # perform export
        exp = requests.get(f"{BASE_URL}/api/po/{any_po_id}/pdf?force=1",
                           headers=_hdr("purchase"), timeout=30)
        assert exp.status_code == 200
        time.sleep(0.5)
        after = requests.get(f"{BASE_URL}/api/audit?entity_id={any_po_id}",
                             headers=_hdr("director"), timeout=15)
        after_c = len([r for r in after.json() if r.get("action") == "export_pdf"])
        assert after_c > before, "audit count did not grow after export"

    def test_export_action_has_correct_shape(self, any_po_id):
        requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=excel&force=1",
                     headers=_hdr("purchase"), timeout=30)
        time.sleep(0.5)
        r = requests.get(f"{BASE_URL}/api/audit?entity_id={any_po_id}",
                         headers=_hdr("director"), timeout=15)
        rows = [x for x in r.json() if x.get("action") == "export_excel"]
        assert rows
        top = rows[0]
        assert top.get("entity") in ("po", "mrf", "tally")
        # details.format may be present
        d = top.get("details") or {}
        assert "format" in d or "record_number" in d, f"details missing keys: {d}"


# --------------------- Validation error content ---------------------------
class TestValidationContent:
    def test_missing_customer_id_and_name(self):
        """Create a PO with no customer_id (via project without customer) → check 422."""
        r = requests.get(f"{BASE_URL}/api/po", headers=_hdr("purchase"), timeout=20)
        assert r.status_code == 200
        for po in r.json():
            if not po.get("customer_id"):
                po_id = po["po_id"]
                ex = requests.get(f"{BASE_URL}/api/po/{po_id}/export?format=pdf",
                                  headers=_hdr("purchase"), timeout=30)
                assert ex.status_code == 422
                miss = ex.json().get("detail", {}).get("missing_mandatory", [])
                assert any("Customer ID" in m for m in miss)
                return
        pytest.skip("No PO with missing customer_id present in seed data")

    def test_hsn_only_mandatory_for_tally(self, any_po_id):
        """PDF/Excel should treat HSN as warning; Tally as mandatory."""
        pdf_r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=pdf",
                             headers=_hdr("purchase"), timeout=30)
        excel_r = requests.get(f"{BASE_URL}/api/po/{any_po_id}/export?format=excel",
                               headers=_hdr("purchase"), timeout=30)
        # If validation fails for pdf/excel, it must NOT be due to HSN
        for r_, fmt in ((pdf_r, "pdf"), (excel_r, "excel")):
            if r_.status_code == 422:
                miss = r_.json().get("detail", {}).get("missing_mandatory", [])
                hsn_missing = [m for m in miss if "HSN" in m]
                assert not hsn_missing, (
                    f"HSN is mandatory in {fmt} export — should be warning only: {miss}"
                )


# --------------------- Regression -----------------------------------------
class TestRegression:
    def test_seed_idempotent(self):
        r1 = requests.post(f"{BASE_URL}/api/seed", headers=_hdr("admin"), timeout=30)
        assert r1.status_code in (200, 201)
        r2 = requests.post(f"{BASE_URL}/api/seed", headers=_hdr("admin"), timeout=30)
        assert r2.status_code in (200, 201)

    def test_grn_pdf_regression(self):
        # find a GRN
        rp = requests.get(f"{BASE_URL}/api/po", headers=_hdr("purchase"), timeout=15)
        for po in rp.json():
            g = requests.get(f"{BASE_URL}/api/po/{po['po_id']}/grns",
                             headers=_hdr("purchase"), timeout=15)
            if g.status_code == 200 and g.json():
                grn_id = g.json()[0]["grn_id"]
                r = requests.get(f"{BASE_URL}/api/grn/{grn_id}/pdf",
                                 headers=_hdr("purchase"), timeout=20)
                assert r.status_code == 200
                assert r.content[:4] == b"%PDF"
                return
        pytest.skip("No GRN present in system")

    def test_materials_default_returns_approved_only(self):
        r = requests.get(f"{BASE_URL}/api/materials",
                         headers=_hdr("purchase"), timeout=15)
        assert r.status_code == 200
        for m in r.json():
            assert m.get("status") in ("approved", None), f"got: {m.get('status')}"
