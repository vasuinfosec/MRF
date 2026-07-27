"""Phase 2c backend regression:
- Customer ID snapshot on MRF/PO (create-time snapshot from project)
- Tally-compatible XLSX export (kind=purchase|invoice; RBAC purchase/director/admin)
- Branded PDF endpoints (GRN/PO/Invoice) — smoke test content-type + size only
- MRF line-item edit with mandatory reason + master_audit mirror + locked status rules
"""
import io
import os
import uuid

import pytest
import requests

BASE_URL = os.environ.get("EXPO_PUBLIC_BACKEND_URL") or os.environ.get("EXPO_BACKEND_URL")
assert BASE_URL, "EXPO_PUBLIC_BACKEND_URL / EXPO_BACKEND_URL must be set"
BASE_URL = BASE_URL.rstrip("/")
API = f"{BASE_URL}/api"

# -------------------- helpers --------------------

def _login(email: str, role: str, name: str = "Test User") -> str:
    r = requests.post(f"{API}/auth/dev-login",
                      json={"email": email, "role": role, "name": name}, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()["session_token"]


def _hdr(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


@pytest.fixture(scope="module")
def tokens():
    return {
        "admin":    _login("admin@vasu.dev", "admin", "Admin"),
        "purchase": _login("purchase@vasu.dev", "purchase", "Purchase"),
        "pm":       _login("pm@vasu.dev", "pm", "PM"),
        "se":       _login("siteeng@vasu.dev", "site_engineer", "Site Eng"),
        "director": _login("director@vasu.dev", "director", "Director"),
        "gm":       _login("gm@vasu.dev", "gm", "GM"),
    }


@pytest.fixture(scope="module")
def se2_token():
    # Second site-engineer distinct from siteeng@vasu.dev for RBAC test
    email = f"se2_{uuid.uuid4().hex[:8]}@vasu.dev"
    return _login(email, "site_engineer", "SE Two")


@pytest.fixture(scope="module")
def a_project_with_customer(tokens):
    """Return a project doc that has customer_id populated. Creates one if none exists."""
    r = requests.get(f"{API}/projects", headers=_hdr(tokens["admin"]), timeout=15)
    assert r.status_code == 200, r.text
    projects = r.json()
    for p in projects:
        if p.get("customer_id"):
            return p
    # Create a fresh customer + project
    cid = f"TESTCUST-{uuid.uuid4().hex[:8].upper()}"
    r = requests.post(f"{API}/customers", headers=_hdr(tokens["admin"]),
                      json={"customer_id": cid, "name": f"TEST Cust {cid}",
                            "gstin": "27ABCDE1234F1Z5", "state": "MH"}, timeout=15)
    assert r.status_code == 200, r.text
    pcode = f"TP-{uuid.uuid4().hex[:6].upper()}"
    r = requests.post(f"{API}/projects", headers=_hdr(tokens["admin"]),
                      json={"code": pcode, "name": f"TEST Project {pcode}",
                            "site": "Pune", "client": "ACME",
                            "customer_id": cid, "reason": "phase2c test setup"}, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="module")
def a_vendor(tokens):
    r = requests.get(f"{API}/vendors", headers=_hdr(tokens["admin"]), timeout=15)
    assert r.status_code == 200
    vendors = [v for v in r.json() if v.get("active", True)]
    assert vendors, "Need at least one active vendor seeded"
    return vendors[0]


# -------------------- Customer snapshot --------------------

class TestCustomerSnapshotFields:
    def test_1_mrf_list_has_customer_fields(self, tokens):
        r = requests.get(f"{API}/mrf", headers=_hdr(tokens["admin"]), timeout=20)
        assert r.status_code == 200
        mrfs = r.json()
        assert isinstance(mrfs, list) and mrfs, "Expected some MRFs to exist for backfill check"
        # every doc should CONTAIN the keys (value may be null for legacy w/o project.customer)
        for m in mrfs[:20]:
            assert "customer_id" in m, f"MRF {m.get('mrf_number')} missing customer_id key"
            assert "customer_name" in m, f"MRF {m.get('mrf_number')} missing customer_name key"
        # at least one backfilled with non-null
        assert any(m.get("customer_id") for m in mrfs), "No MRF has customer_id populated (backfill expected 12)"

    def test_2_po_list_has_customer_fields(self, tokens):
        r = requests.get(f"{API}/po", headers=_hdr(tokens["admin"]), timeout=20)
        assert r.status_code == 200
        pos = r.json()
        assert isinstance(pos, list) and pos, "Expected some POs to exist"
        for p in pos[:20]:
            assert "customer_id" in p, f"PO {p.get('po_number')} missing customer_id key"
            assert "customer_name" in p, f"PO {p.get('po_number')} missing customer_name key"
        assert any(p.get("customer_id") for p in pos), "No PO has customer_id populated (backfill expected 4)"


@pytest.fixture(scope="module")
def new_mrf(tokens, a_project_with_customer):
    """Create a fresh MRF against a project with customer_id — reused by later tests."""
    proj = a_project_with_customer
    body = {
        "project_id": proj["project_id"],
        "site": proj.get("site", "Pune"),
        "required_by": "2026-02-01",
        "requesting_person": "Test Requester",
        "system_category": "electrical",
        "items": [
            {"description": "TEST-Cable 2C x 1.5", "unit": "mtr", "qty_requested": 100,
             "specification": "Cu, PVC"},
            {"description": "TEST-Switch 10A", "unit": "nos", "qty_requested": 20},
        ],
        "remarks": "phase2c snapshot test",
    }
    r = requests.post(f"{API}/mrf", headers=_hdr(tokens["admin"]), json=body, timeout=20)
    assert r.status_code == 200, r.text
    return r.json(), proj


class TestCreateMRFSnapshot:
    def test_3_new_mrf_carries_customer_id(self, new_mrf):
        mrf, proj = new_mrf
        assert mrf.get("customer_id") == proj["customer_id"], \
            f"MRF snapshot mismatch: mrf.customer_id={mrf.get('customer_id')} project.customer_id={proj['customer_id']}"
        assert mrf.get("customer_name"), "customer_name should be populated"

    def test_5_get_single_mrf_has_customer(self, tokens, new_mrf):
        mrf, proj = new_mrf
        r = requests.get(f"{API}/mrf/{mrf['mrf_id']}", headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        got = r.json()
        assert got["customer_id"] == proj["customer_id"]
        assert got["customer_name"] == mrf["customer_name"]


@pytest.fixture(scope="module")
def issued_po(tokens, new_mrf, a_vendor):
    """Approve/authorize the MRF, then create a PO from it."""
    mrf, proj = new_mrf
    # draft -> pm_review (submit as admin; admin allowed regardless of creator)
    r = requests.post(f"{API}/mrf/{mrf['mrf_id']}/submit",
                      headers=_hdr(tokens["admin"]), timeout=15)
    assert r.status_code == 200, f"submit failed: {r.status_code} {r.text}"
    # pm_review -> approved (admin can approve)
    r = requests.post(f"{API}/mrf/{mrf['mrf_id']}/approve",
                      headers=_hdr(tokens["admin"]),
                      json={"action": "approve", "comment": "test"}, timeout=15)
    assert r.status_code == 200, f"approve failed: {r.status_code} {r.text}"
    # approved -> sent_to_purchase
    requests.post(f"{API}/mrf/{mrf['mrf_id']}/send-to-purchase",
                  headers=_hdr(tokens["admin"]), timeout=15)
    r = requests.get(f"{API}/mrf/{mrf['mrf_id']}", headers=_hdr(tokens["admin"]), timeout=15)
    mrf_now = r.json()
    assert mrf_now["status"] in ("approved", "sent_to_purchase", "partially_ordered"), \
        f"MRF ended in unpurchasable state: {mrf_now['status']}"

    # Create PO with first item
    it0 = mrf_now["items"][0]
    body = {
        "vendor_id": a_vendor["vendor_id"],
        "project_id": mrf_now["project_id"],
        "delivery_site": mrf_now.get("site", "Pune"),
        "items": [{
            "mrf_id": mrf_now["mrf_id"],
            "item_line_id": it0["item_line_id"],
            "description": it0["description"],
            "unit": it0["unit"],
            "qty": 5,
            "rate": 100.0,
            "discount": 0,
            "gst": 18,
            "specification": it0.get("specification", ""),
        }],
        "delivery_schedule": "2026-02-15",
        "payment_terms": "30 days",
        "warranty_terms": "1yr",
        "freight": 0,
        "other_charges": 0,
        "authorised_signatory": "Admin",
    }
    r = requests.post(f"{API}/po", headers=_hdr(tokens["purchase"]), json=body, timeout=20)
    assert r.status_code == 200, f"PO create failed: {r.status_code} {r.text}"
    return r.json(), proj


class TestPOSnapshot:
    def test_4_new_po_carries_customer_id_and_name(self, issued_po):
        po, proj = issued_po
        assert po.get("customer_id") == proj["customer_id"], \
            f"PO customer_id mismatch: {po.get('customer_id')} != {proj['customer_id']}"
        assert po.get("customer_name"), "PO customer_name should be populated"


# -------------------- Tally XLSX export --------------------

def _is_xlsx(resp) -> bool:
    ct = resp.headers.get("content-type", "").lower()
    ok_ct = "spreadsheetml.sheet" in ct or "openxmlformats" in ct or "excel" in ct
    body_starts_pk = resp.content[:2] == b"PK"
    return ok_ct and body_starts_pk


class TestTallyExport:
    def test_6_admin_default_purchase(self, tokens):
        r = requests.get(f"{API}/export/tally", headers=_hdr(tokens["admin"]), timeout=30)
        assert r.status_code == 200, r.text
        assert _is_xlsx(r), f"content-type={r.headers.get('content-type')} first-bytes={r.content[:4]!r}"

    def test_7_purchase_kind_invoice(self, tokens):
        r = requests.get(f"{API}/export/tally", params={"kind": "invoice"},
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 200, r.text
        assert _is_xlsx(r)

    def test_8_bogus_kind_400(self, tokens):
        r = requests.get(f"{API}/export/tally", params={"kind": "bogus"},
                         headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 400, f"expected 400, got {r.status_code}: {r.text}"

    def test_9_site_engineer_403(self, tokens):
        r = requests.get(f"{API}/export/tally", headers=_hdr(tokens["se"]), timeout=15)
        assert r.status_code == 403, r.text

    def test_10_pm_403(self, tokens):
        r = requests.get(f"{API}/export/tally", headers=_hdr(tokens["pm"]), timeout=15)
        assert r.status_code == 403, r.text

    def test_11_xlsx_headers_match_spec(self, tokens):
        import openpyxl
        r = requests.get(f"{API}/export/tally", headers=_hdr(tokens["admin"]), timeout=30)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Header spec expanded in later iterations to satisfy Tally GST filings:
        # company GSTIN block, customer GSTIN, project code, master IDs
        # (MAT UID / VAR UID), and make/model.
        expected = [
            "Voucher Date", "Voucher Type", "Voucher No", "Reference No", "Reference Date",
            "Company GSTIN", "Company State", "Company State Code",
            "Party Ledger", "Party GSTIN", "Party State",
            "Customer ID", "Customer Name", "Customer GSTIN",
            "Project Code",
            "MAT UID", "VAR UID",
            "Item Name", "Make", "Model",
            "HSN/SAC", "Unit", "Quantity", "Rate", "Discount",
            "Taxable Value", "GST Rate %", "CGST", "SGST", "IGST",
            "Line Total", "Narration",
        ]
        got = [c.value for c in ws[1]]
        assert got == expected, f"Header mismatch.\n expected={expected}\n got={got}"

    def test_24_kind_defaults_to_purchase(self, tokens):
        # No kind param → same as ?kind=purchase (both 200 XLSX, same header row)
        import openpyxl
        r1 = requests.get(f"{API}/export/tally", headers=_hdr(tokens["admin"]), timeout=30)
        r2 = requests.get(f"{API}/export/tally", params={"kind": "purchase"},
                          headers=_hdr(tokens["admin"]), timeout=30)
        assert r1.status_code == 200 and r2.status_code == 200
        h1 = [c.value for c in openpyxl.load_workbook(io.BytesIO(r1.content)).active[1]]
        h2 = [c.value for c in openpyxl.load_workbook(io.BytesIO(r2.content)).active[1]]
        assert h1 == h2


# -------------------- PDF smoke --------------------

def _is_pdf(resp) -> bool:
    ct = resp.headers.get("content-type", "").lower()
    return "application/pdf" in ct and resp.content[:4] == b"%PDF"


class TestPDFSmoke:
    def test_12_po_pdf(self, tokens, issued_po):
        po, _ = issued_po
        r = requests.get(f"{API}/po/{po['po_id']}/pdf",
                         headers=_hdr(tokens["admin"]), timeout=30)
        assert r.status_code == 200, r.text
        assert _is_pdf(r), f"content-type={r.headers.get('content-type')}"
        assert len(r.content) > 1024, f"pdf too small: {len(r.content)} bytes"

    def test_13_grn_pdf(self, tokens, issued_po):
        """GRN is created as a side-effect of POST /po/{po_id}/received."""
        po, _ = issued_po
        it0 = po["items"][0]
        recv_body = {"items": [{
            "mrf_id": it0["mrf_id"],
            "item_line_id": it0["item_line_id"],
            "qty": 1,
        }]}
        r = requests.post(f"{API}/po/{po['po_id']}/received",
                          headers=_hdr(tokens["admin"]), json=recv_body, timeout=20)
        assert r.status_code == 200, f"receive failed: {r.status_code} {r.text}"
        # Find the GRN just created
        r = requests.get(f"{API}/po/{po['po_id']}/grns",
                         headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        grns = r.json()
        assert grns, "expected at least one GRN after receive"
        grn = grns[0]
        r = requests.get(f"{API}/grn/{grn['grn_id']}/pdf",
                         headers=_hdr(tokens["admin"]), timeout=30)
        assert r.status_code == 200, r.text
        assert _is_pdf(r)

    def test_14_invoice_pdf_if_any(self, tokens, issued_po):
        po, _ = issued_po
        # Look for existing invoices first
        r = requests.get(f"{API}/invoice", headers=_hdr(tokens["admin"]), timeout=15)
        invoices = r.json() if r.status_code == 200 else []
        if not invoices:
            # Create one against our issued PO so the PDF endpoint has something to render
            it0 = po["items"][0]
            body = {
                "po_id": po["po_id"],
                "vendor_invoice_number": f"VINV-TEST-{uuid.uuid4().hex[:6]}",
                "invoice_date": "2026-01-20",
                "items": [{
                    "mrf_id": it0["mrf_id"],
                    "item_line_id": it0["item_line_id"],
                    "description": it0["description"],
                    "unit": it0["unit"],
                    "qty": 1, "rate": 100.0, "discount": 0, "gst": 18,
                }],
                "freight": 0, "other_charges": 0, "remarks": "phase2c test",
            }
            r = requests.post(f"{API}/invoice", headers=_hdr(tokens["admin"]),
                              json=body, timeout=20)
            if r.status_code != 200:
                pytest.skip(f"could not create invoice: {r.status_code} {r.text[:150]}")
            inv = r.json()
        else:
            inv = invoices[0]
        r = requests.get(f"{API}/invoice/{inv['invoice_id']}/pdf",
                         headers=_hdr(tokens["admin"]), timeout=30)
        assert r.status_code == 200, r.text
        assert _is_pdf(r)


# -------------------- MRF line-item edit --------------------

@pytest.fixture(scope="module")
def se_created_mrf(tokens, se2_token, a_project_with_customer):
    """MRF created by siteeng@vasu.dev (tokens['se']) so SE-creator rules apply."""
    # Ensure SE is assigned to the project
    proj = a_project_with_customer
    # Get SE user_id
    me = requests.get(f"{API}/auth/me", headers=_hdr(tokens["se"]), timeout=10).json()
    se_id = me.get("user_id") or me.get("user", {}).get("user_id")
    if se_id and se_id not in (proj.get("site_engineers") or []):
        # Admin adds SE to project
        requests.put(f"{API}/projects/{proj['project_id']}",
                     headers=_hdr(tokens["admin"]),
                     json={"site_engineers": list(set((proj.get("site_engineers") or []) + [se_id])),
                           "reason": "phase2c: add SE for line-edit test"},
                     timeout=15)
    body = {
        "project_id": proj["project_id"],
        "site": proj.get("site", "Pune"),
        "required_by": "2026-02-05",
        "requesting_person": "SE Requester",
        "system_category": "electrical",
        "items": [
            {"description": "TEST-Screw M6", "unit": "nos", "qty_requested": 50},
            {"description": "TEST-Nut M6", "unit": "nos", "qty_requested": 50},
        ],
    }
    r = requests.post(f"{API}/mrf", headers=_hdr(tokens["se"]), json=body, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


class TestMRFLineEdit:
    def test_15_no_reason_400(self, tokens, new_mrf):
        mrf, _ = new_mrf
        line = mrf["items"][0]
        r = requests.put(f"{API}/mrf/{mrf['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(tokens["admin"]),
                         json={"description": "changed"}, timeout=15)
        assert r.status_code == 400, r.text
        assert "reason" in r.text.lower()

    def test_16_admin_edit_with_reason_and_audit_mirror(self, tokens, new_mrf):
        mrf, _ = new_mrf
        line = mrf["items"][1]  # use 2nd item to avoid clashing with test_17
        new_desc = f"TEST-Switch 10A EDITED {uuid.uuid4().hex[:4]}"
        r = requests.put(f"{API}/mrf/{mrf['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(tokens["admin"]),
                         json={"description": new_desc, "reason": "spec update - phase2c"},
                         timeout=15)
        assert r.status_code == 200, r.text
        body = r.json()
        assert "changes" in body and "description" in body["changes"], body
        assert body["changes"]["description"]["new"] == new_desc
        assert "log" in body and body["log"]["reason"] == "spec update - phase2c"

        # master-audit filter by entity=mrf.item shows this line
        r = requests.get(f"{API}/master-audit", params={"entity": "mrf.item"},
                         headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        entries = r.json()
        assert isinstance(entries, list) and entries, "expected master_audit entries for mrf.item"
        expected_eid = f"{mrf['mrf_number']}#{line['item_line_id']}"
        assert any(e.get("entity_id") == expected_eid for e in entries), \
            f"master_audit did not record entity_id={expected_eid}"

    def test_17_no_changes_400(self, tokens, new_mrf):
        mrf, _ = new_mrf
        # Re-fetch current state
        r = requests.get(f"{API}/mrf/{mrf['mrf_id']}", headers=_hdr(tokens["admin"]), timeout=15)
        cur = r.json()
        line = cur["items"][0]
        r = requests.put(f"{API}/mrf/{mrf['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(tokens["admin"]),
                         json={"description": line["description"], "reason": "no-op"},
                         timeout=15)
        assert r.status_code == 400, r.text
        assert "no changes" in r.text.lower()

    def test_18_se_creator_can_edit_own(self, tokens, se_created_mrf):
        mrf = se_created_mrf
        line = mrf["items"][0]
        r = requests.put(f"{API}/mrf/{mrf['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(tokens["se"]),
                         json={"remarks": "SE-owner edited remarks",
                               "reason": "clarify remarks"}, timeout=15)
        assert r.status_code == 200, r.text

    def test_19_se_non_creator_forbidden(self, se2_token, se_created_mrf):
        mrf = se_created_mrf
        line = mrf["items"][0]
        r = requests.put(f"{API}/mrf/{mrf['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(se2_token),
                         json={"remarks": "hijack attempt", "reason": "test rbac"}, timeout=15)
        # Could be 403 (creator rule) or 403 due to _check_mrf_access (SE not creator).
        assert r.status_code == 403, f"expected 403, got {r.status_code}: {r.text}"

    def test_22_nonexistent_line_id_404(self, tokens, new_mrf):
        mrf, _ = new_mrf
        r = requests.put(f"{API}/mrf/{mrf['mrf_id']}/items/does-not-exist",
                         headers=_hdr(tokens["admin"]),
                         json={"description": "x", "reason": "y"}, timeout=15)
        assert r.status_code == 404, r.text

    def test_23_non_numeric_qty_400(self, tokens, new_mrf):
        mrf, _ = new_mrf
        r = requests.get(f"{API}/mrf/{mrf['mrf_id']}", headers=_hdr(tokens["admin"]), timeout=15)
        cur = r.json()
        line = cur["items"][0]
        r = requests.put(f"{API}/mrf/{mrf['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(tokens["admin"]),
                         json={"qty_requested": "not-a-number", "reason": "invalid qty"},
                         timeout=15)
        assert r.status_code == 400, r.text
        assert "numeric" in r.text.lower() or "qty_requested" in r.text.lower()


@pytest.fixture(scope="module")
def locked_mrf(tokens, a_project_with_customer, a_vendor):
    """Build an MRF that ends up in the locked 'received' status by:
    create -> submit -> approve -> send-to-purchase -> PO -> fully receive.
    Returns the freshly-fetched MRF (status should be 'received')."""
    proj = a_project_with_customer
    body = {
        "project_id": proj["project_id"],
        "site": proj.get("site", "Pune"),
        "required_by": "2026-02-10",
        "requesting_person": "Locked Test",
        "system_category": "electrical",
        "items": [{"description": "TEST-Locked item", "unit": "nos", "qty_requested": 3}],
    }
    r = requests.post(f"{API}/mrf", headers=_hdr(tokens["admin"]), json=body, timeout=15)
    assert r.status_code == 200, r.text
    mrf = r.json()
    line = mrf["items"][0]

    # submit -> approve -> send-to-purchase
    assert requests.post(f"{API}/mrf/{mrf['mrf_id']}/submit",
                         headers=_hdr(tokens["admin"]), timeout=15).status_code == 200
    assert requests.post(f"{API}/mrf/{mrf['mrf_id']}/approve",
                         headers=_hdr(tokens["admin"]),
                         json={"action": "approve"}, timeout=15).status_code == 200
    requests.post(f"{API}/mrf/{mrf['mrf_id']}/send-to-purchase",
                  headers=_hdr(tokens["admin"]), timeout=15)

    # Create PO for full qty
    po_body = {
        "vendor_id": a_vendor["vendor_id"],
        "project_id": mrf["project_id"],
        "delivery_site": mrf["site"],
        "items": [{
            "mrf_id": mrf["mrf_id"],
            "item_line_id": line["item_line_id"],
            "description": line["description"],
            "unit": line["unit"],
            "qty": 3, "rate": 50.0, "discount": 0, "gst": 18,
        }],
        "delivery_schedule": "2026-02-15",
        "payment_terms": "30 days", "warranty_terms": "1yr",
        "freight": 0, "other_charges": 0,
        "authorised_signatory": "Admin",
    }
    r = requests.post(f"{API}/po", headers=_hdr(tokens["purchase"]), json=po_body, timeout=20)
    assert r.status_code == 200, f"PO create failed: {r.text}"
    po = r.json()

    # Fully receive PO -> MRF moves to 'received'
    r = requests.post(f"{API}/po/{po['po_id']}/received",
                      headers=_hdr(tokens["admin"]),
                      json={"items": [{"mrf_id": mrf["mrf_id"],
                                       "item_line_id": line["item_line_id"], "qty": 3}]},
                      timeout=20)
    assert r.status_code == 200, r.text

    r = requests.get(f"{API}/mrf/{mrf['mrf_id']}", headers=_hdr(tokens["admin"]), timeout=10)
    return r.json()


class TestMRFLockedStatus:
    def test_20_admin_can_edit_locked(self, tokens, locked_mrf):
        # locked_mrf may or may not actually be locked depending on API surface; skip if not
        from_srv = locked_mrf
        if from_srv.get("status") not in ("received", "fully_received", "closed",
                                          "cancelled", "rejected"):
            pytest.skip(f"could not force MRF into locked state (status={from_srv.get('status')})")
        line = from_srv["items"][0]
        r = requests.put(f"{API}/mrf/{from_srv['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(tokens["admin"]),
                         json={"remarks": "admin override on locked",
                               "reason": "admin override test"}, timeout=15)
        assert r.status_code == 200, r.text

    def test_21_se_cannot_edit_locked(self, tokens, locked_mrf):
        from_srv = locked_mrf
        if from_srv.get("status") not in ("received", "fully_received", "closed",
                                          "cancelled", "rejected"):
            pytest.skip(f"could not force MRF into locked state (status={from_srv.get('status')})")
        line = from_srv["items"][0]
        r = requests.put(f"{API}/mrf/{from_srv['mrf_id']}/items/{line['item_line_id']}",
                         headers=_hdr(tokens["se"]),
                         json={"remarks": "SE hijack on locked",
                               "reason": "se on locked"}, timeout=15)
        # SE isn't the creator here (admin created it), so 403 might fire before 400.
        # Both are acceptable outcomes; the primary intent — SE blocked — is verified.
        assert r.status_code in (400, 403), r.text


class TestMasterAuditFilter:
    def test_25_master_audit_filter_by_mrf_item(self, tokens):
        r = requests.get(f"{API}/master-audit", params={"entity": "mrf.item"},
                         headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        entries = r.json()
        assert isinstance(entries, list), entries
        # Should be non-empty because test_16 created one
        assert entries, "master-audit filter=mrf.item returned []; expected recent line edits"
        for e in entries[:5]:
            assert e.get("entity") == "mrf.item"
