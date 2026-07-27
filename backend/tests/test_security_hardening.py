"""
Iteration 7 — Security hardening tests.
Validates:
  SEC-001: dev-login gated behind ENABLE_DEV_LOGIN; /api/seed requires admin once users exist.
  SEC-002: role/project scoping on GET /api/mrf/{id}, /api/po/{id}, /api/invoice*, /api/audit,
           /api/billing/items, /api/users (email redaction), /api/export/*, /api/po/{id}/grns, /api/grns.
  SEC-003: /api/po/{id}/received restricted to purchase/billing/admin.
  SEC-004: Excel formula-injection sanitization + 5MB upload cap + .xlsx/.xlsm allowlist on imports.
"""
import io
import os
import uuid
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("EXPO_PUBLIC_BACKEND_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"

SEEDED = {
    "site_engineer": "site@vasu.dev",
    "project_manager": "pm@vasu.dev",
    "purchase": "purchase@vasu.dev",
    "billing": "billing@vasu.dev",
    "admin": "admin@vasu.dev",
}


# ---------- helpers ----------
def _login(email, role, name=None):
    r = requests.post(f"{API}/auth/dev-login",
                      json={"email": email, "role": role, "name": name or email})
    assert r.status_code == 200, f"dev-login failed for {email}/{role}: {r.status_code} {r.text}"
    return r.json()["session_token"], r.json()["user"]


def _hdr(tok):
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def tokens():
    # Ensure seed data exists first (idempotent if admin, or bootstrap on empty db)
    admin_tok, _ = _login(SEEDED["admin"], "admin")
    r = requests.post(f"{API}/seed", headers=_hdr(admin_tok))
    assert r.status_code == 200, f"seed failed: {r.status_code} {r.text}"

    toks = {}
    users = {}
    for role, email in SEEDED.items():
        t, u = _login(email, role)
        toks[role] = t
        users[role] = u
    # Fresh unrelated site engineer for negative auth tests
    stranger_email = f"stranger_{uuid.uuid4().hex[:6]}@vasu.dev"
    t, u = _login(stranger_email, "site_engineer", name="Stranger SE")
    toks["stranger"] = t
    users["stranger"] = u
    return {"tokens": toks, "users": users, "admin_tok": admin_tok}


@pytest.fixture(scope="module")
def sample_ids(tokens):
    """Create an MRF (as seeded site engineer) + PO (as purchase) for reuse."""
    ttoks = tokens["tokens"]
    # find project VIS-101
    r = requests.get(f"{API}/projects", headers=_hdr(ttoks["admin"]))
    assert r.status_code == 200, r.text
    projects = r.json()
    proj = next((p for p in projects if p.get("code") == "VIS-101"), None) or projects[0]
    proj_id = proj["project_id"]

    # Create MRF as site engineer (seeded is assigned to VIS-101)
    mrf_payload = {
        "project_id": proj_id,
        "site": "Bangalore",
        "required_by": "2026-02-15",
        "requesting_person": "TEST_Tester",
        "system_category": "Fire Safety",
        "items": [{"description": "TEST_Smoke Detector", "unit": "Nos",
                   "qty_requested": 5.0, "brand_preference": "Honeywell"}],
        "remarks": "TEST security"
    }
    r = requests.post(f"{API}/mrf", json=mrf_payload, headers=_hdr(ttoks["site_engineer"]))
    assert r.status_code == 200, f"create mrf: {r.text}"
    mrf = r.json()
    mrf_id = mrf["mrf_id"]

    # Submit + approve so PO can be created
    r = requests.post(f"{API}/mrf/{mrf_id}/submit", headers=_hdr(ttoks["site_engineer"]))
    assert r.status_code == 200, r.text
    r = requests.post(f"{API}/mrf/{mrf_id}/approve",
                      json={"action": "approve", "comment": "ok"},
                      headers=_hdr(ttoks["project_manager"]))
    assert r.status_code == 200, r.text

    # Fetch approved mrf to get item_line_id
    r = requests.get(f"{API}/mrf/{mrf_id}", headers=_hdr(ttoks["admin"]))
    assert r.status_code == 200, r.text
    mrf_full = r.json()
    line = mrf_full["items"][0]

    # vendor
    r = requests.get(f"{API}/vendors", headers=_hdr(ttoks["purchase"]))
    assert r.status_code == 200, r.text
    vendors = r.json()
    vendor_id = vendors[0]["vendor_id"]

    po_payload = {
        "vendor_id": vendor_id,
        "project_id": proj_id,
        "delivery_site": "Bangalore Whitefield",
        "items": [{
            "mrf_id": mrf_id, "item_line_id": line["item_line_id"],
            "description": line["description"], "unit": line["unit"],
            "qty": line.get("qty_approved") or line["qty_requested"],
            "rate": 500, "gst": 18,
        }],
        "delivery_schedule": "1 week", "payment_terms": "Net 30",
        "warranty_terms": "1yr", "freight": 0, "other_charges": 0,
    }
    r = requests.post(f"{API}/po", json=po_payload, headers=_hdr(ttoks["purchase"]))
    assert r.status_code == 200, f"create po: {r.text}"
    po = r.json()
    return {"mrf_id": mrf_id, "po_id": po["po_id"], "project_id": proj_id, "vendor_id": vendor_id}


# ============ SEC-001: dev-login gate + seed admin gate ============
class TestSec001DevLoginAndSeed:
    def test_dev_login_enabled(self):
        # ENABLE_DEV_LOGIN=1 per backend/.env
        r = requests.post(f"{API}/auth/dev-login",
                          json={"email": "admin@vasu.dev", "role": "admin"})
        assert r.status_code == 200
        assert "session_token" in r.json()

    def test_seed_without_auth_401(self, tokens):
        # users already exist → must require admin auth
        r = requests.post(f"{API}/seed")
        assert r.status_code == 401, f"expected 401, got {r.status_code}: {r.text}"

    def test_seed_as_site_engineer_403(self, tokens):
        r = requests.post(f"{API}/seed",
                         headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 403

    def test_seed_as_admin_200(self, tokens):
        r = requests.post(f"{API}/seed", headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200
        assert r.json().get("ok") is True


# ============ SEC-002: /api/audit admin-only ============
class TestSec002Audit:
    @pytest.mark.parametrize("role", ["site_engineer", "project_manager", "purchase", "billing"])
    def test_non_admin_forbidden(self, tokens, role):
        r = requests.get(f"{API}/audit", headers=_hdr(tokens["tokens"][role]))
        assert r.status_code == 403, f"{role} got {r.status_code}"

    def test_admin_ok(self, tokens):
        r = requests.get(f"{API}/audit", headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200
        assert isinstance(r.json(), list)


# ============ SEC-002: /api/users email redaction for non-admin ============
class TestSec002UsersRedaction:
    def test_site_engineer_gets_empty_emails(self, tokens):
        r = requests.get(f"{API}/users", headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 200
        users = r.json()
        assert len(users) > 0
        for u in users:
            assert u.get("email", "") == "", f"non-admin saw email: {u}"

    def test_admin_gets_real_emails(self, tokens):
        r = requests.get(f"{API}/users", headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200
        users = r.json()
        emails = [u["email"] for u in users]
        assert "admin@vasu.dev" in emails
        assert any("@" in e for e in emails)


# ============ SEC-003: /api/po/{id}/received role gate ============
class TestSec003POReceived:
    @pytest.mark.parametrize("role", ["site_engineer", "project_manager"])
    def test_non_authorized_403(self, tokens, sample_ids, role):
        r = requests.post(f"{API}/po/{sample_ids['po_id']}/received",
                          json={}, headers=_hdr(tokens["tokens"][role]))
        assert r.status_code == 403, f"{role} got {r.status_code}: {r.text}"

    def test_purchase_ok(self, tokens, sample_ids):
        # partial receipt (0 items) — endpoint accepts empty body, receives full remainder
        # Use per_line dict targeting no matching keys to avoid actually receiving here
        r = requests.post(f"{API}/po/{sample_ids['po_id']}/received",
                          json={"items": [{"mrf_id": "none", "item_line_id": "none", "qty": 0}]},
                          headers=_hdr(tokens["tokens"]["purchase"]))
        assert r.status_code == 200, r.text

    def test_billing_ok(self, tokens, sample_ids):
        r = requests.post(f"{API}/po/{sample_ids['po_id']}/received",
                          json={"items": [{"mrf_id": "none", "item_line_id": "none", "qty": 0}]},
                          headers=_hdr(tokens["tokens"]["billing"]))
        assert r.status_code == 200

    def test_admin_ok(self, tokens, sample_ids):
        r = requests.post(f"{API}/po/{sample_ids['po_id']}/received",
                          json={"items": [{"mrf_id": "none", "item_line_id": "none", "qty": 0}]},
                          headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200


# ============ SEC-002: PO/MRF access scoping ============
class TestSec002POMRFScoping:
    def test_get_po_stranger_403(self, tokens, sample_ids):
        r = requests.get(f"{API}/po/{sample_ids['po_id']}",
                         headers=_hdr(tokens["tokens"]["stranger"]))
        assert r.status_code == 403

    def test_get_po_admin_200(self, tokens, sample_ids):
        r = requests.get(f"{API}/po/{sample_ids['po_id']}",
                         headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200
        assert r.json()["po_id"] == sample_ids["po_id"]

    def test_get_po_purchase_200(self, tokens, sample_ids):
        r = requests.get(f"{API}/po/{sample_ids['po_id']}",
                         headers=_hdr(tokens["tokens"]["purchase"]))
        assert r.status_code == 200

    def test_get_mrf_stranger_403(self, tokens, sample_ids):
        r = requests.get(f"{API}/mrf/{sample_ids['mrf_id']}",
                         headers=_hdr(tokens["tokens"]["stranger"]))
        assert r.status_code == 403

    def test_get_mrf_creator_200(self, tokens, sample_ids):
        r = requests.get(f"{API}/mrf/{sample_ids['mrf_id']}",
                         headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 200

    def test_get_mrf_admin_200(self, tokens, sample_ids):
        r = requests.get(f"{API}/mrf/{sample_ids['mrf_id']}",
                         headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200


# ============ SEC-002: Invoice + billing/items scoping ============
class TestSec002InvoiceBillingScoping:
    def test_list_invoice_site_engineer_403(self, tokens):
        r = requests.get(f"{API}/invoice", headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 403

    def test_list_invoice_purchase_200(self, tokens):
        r = requests.get(f"{API}/invoice", headers=_hdr(tokens["tokens"]["purchase"]))
        assert r.status_code == 200

    def test_po_invoices_site_engineer_403(self, tokens, sample_ids):
        r = requests.get(f"{API}/po/{sample_ids['po_id']}/invoices",
                         headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 403

    def test_po_invoices_billing_200(self, tokens, sample_ids):
        r = requests.get(f"{API}/po/{sample_ids['po_id']}/invoices",
                         headers=_hdr(tokens["tokens"]["billing"]))
        assert r.status_code == 200

    def test_invoice_by_id_site_engineer_403(self, tokens):
        # Even a random id should get 403 before 404 for non-authorized role
        r = requests.get(f"{API}/invoice/nonexistent",
                         headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 403

    def test_billing_items_site_engineer_403(self, tokens):
        r = requests.get(f"{API}/billing/items",
                         headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 403

    def test_billing_items_billing_200(self, tokens):
        r = requests.get(f"{API}/billing/items",
                         headers=_hdr(tokens["tokens"]["billing"]))
        assert r.status_code == 200


# ============ SEC-002: Export scoping + SEC-004: formula-injection ============
class TestSec002Sec004Exports:
    def test_export_mrf_site_engineer_403(self, tokens):
        tok = tokens["tokens"]["site_engineer"]
        r = requests.get(f"{API}/export/mrf?token={tok}")
        assert r.status_code == 403

    @pytest.mark.parametrize("role", ["project_manager", "purchase", "billing", "admin"])
    def test_export_mrf_authorized_roles(self, tokens, role):
        tok = tokens["tokens"][role]
        r = requests.get(f"{API}/export/mrf?token={tok}")
        assert r.status_code == 200, f"{role} got {r.status_code}"
        assert r.headers.get("content-type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_export_po_site_engineer_403(self, tokens):
        tok = tokens["tokens"]["site_engineer"]
        r = requests.get(f"{API}/export/po?token={tok}")
        assert r.status_code == 403

    def test_export_po_pm_403(self, tokens):
        tok = tokens["tokens"]["project_manager"]
        r = requests.get(f"{API}/export/po?token={tok}")
        assert r.status_code == 403

    @pytest.mark.parametrize("role", ["purchase", "billing", "admin"])
    def test_export_po_authorized(self, tokens, role):
        tok = tokens["tokens"][role]
        r = requests.get(f"{API}/export/po?token={tok}")
        assert r.status_code == 200

    def test_formula_injection_defense(self, tokens):
        """Create an MRF whose description begins with '=' then export and confirm apostrophe prefix."""
        ttoks = tokens["tokens"]
        r = requests.get(f"{API}/projects", headers=_hdr(ttoks["admin"]))
        proj = next((p for p in r.json() if p.get("code") == "VIS-101"), r.json()[0])
        payload = {
            "project_id": proj["project_id"], "site": "Bangalore",
            "required_by": "2026-02-15", "requesting_person": "TEST_Inj",
            "system_category": "Fire Safety",
            "items": [{"description": "=DANGER_FORMULA()", "unit": "Nos", "qty_requested": 1.0}],
            "remarks": "TEST_injection"
        }
        r = requests.post(f"{API}/mrf", json=payload, headers=_hdr(ttoks["site_engineer"]))
        assert r.status_code == 200, r.text
        # Export as admin
        tok = ttoks["admin"]
        r = requests.get(f"{API}/export/mrf?token={tok}")
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
        ws = wb.active
        # Item column is index 8 (1-based col 8: 'Item')
        found_injected = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            # description column index 7 (0-based)
            if row[7] and isinstance(row[7], str) and "DANGER_FORMULA" in row[7]:
                found_injected = True
                assert row[7].startswith("'="), f"formula not neutralized: {row[7]!r}"
                break
        assert found_injected, "Injected MRF item not present in export"


# ============ SEC-004: Import upload cap + extension allowlist ============
class TestSec004Imports:
    def _make_xlsx_bytes(self, size_bytes=None, headers_row=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers_row or ["name", "address", "gstin", "contact", "email"])
        ws.append(["TEST_VENDOR_ROW", "addr", "gst", "9990001111", "x@y.com"])
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        if size_bytes and len(content) < size_bytes:
            # pad up
            content = content + b"\0" * (size_bytes - len(content))
        return content

    def test_vendors_import_bad_extension_400(self, tokens):
        content = self._make_xlsx_bytes()
        files = {"file": ("bad.txt", content, "text/plain")}
        r = requests.post(f"{API}/import/vendors", files=files,
                          headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 400, r.text

    def test_vendors_import_too_large_413(self, tokens):
        # 6MB of arbitrary bytes with .xlsx extension
        big = b"A" * (6 * 1024 * 1024)
        files = {"file": ("big.xlsx", big,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/import/vendors", files=files,
                          headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 413, f"expected 413, got {r.status_code}: {r.text[:200]}"

    def test_mrf_import_bad_extension_400(self, tokens):
        content = b"nothing"
        files = {"file": ("bad.csv", content, "text/csv")}
        r = requests.post(f"{API}/import/mrf", files=files,
                          headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 400

    def test_mrf_import_too_large_413(self, tokens):
        big = b"A" * (6 * 1024 * 1024)
        files = {"file": ("big.xlsx", big,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/import/mrf", files=files,
                          headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 413


# ============ SEC-002: GRN endpoints ============
class TestSec002GRN:
    def test_po_grns_stranger_403(self, tokens, sample_ids):
        r = requests.get(f"{API}/po/{sample_ids['po_id']}/grns",
                         headers=_hdr(tokens["tokens"]["stranger"]))
        assert r.status_code == 403

    @pytest.mark.parametrize("role", ["admin", "purchase", "billing"])
    def test_po_grns_authorized(self, tokens, sample_ids, role):
        r = requests.get(f"{API}/po/{sample_ids['po_id']}/grns",
                         headers=_hdr(tokens["tokens"][role]))
        assert r.status_code == 200, f"{role} got {r.status_code}"
        assert isinstance(r.json(), list)

    def test_grns_list_site_engineer_403(self, tokens):
        r = requests.get(f"{API}/grns", headers=_hdr(tokens["tokens"]["site_engineer"]))
        assert r.status_code == 403

    @pytest.mark.parametrize("role", ["admin", "purchase", "billing"])
    def test_grns_list_authorized(self, tokens, role):
        r = requests.get(f"{API}/grns", headers=_hdr(tokens["tokens"][role]))
        assert r.status_code == 200


# ============ Regression: core flows still work ============
class TestRegressionCoreFlows:
    def test_mrf_list_admin(self, tokens):
        r = requests.get(f"{API}/mrf", headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_po_list_admin(self, tokens):
        r = requests.get(f"{API}/po", headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200

    def test_po_pdf_admin(self, tokens, sample_ids):
        tok = tokens["tokens"]["admin"]
        r = requests.get(f"{API}/po/{sample_ids['po_id']}/pdf?token={tok}")
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/pdf")

    def test_invoice_create_and_pdf(self, tokens, sample_ids):
        ttoks = tokens["tokens"]
        # get PO details
        r = requests.get(f"{API}/po/{sample_ids['po_id']}", headers=_hdr(ttoks["admin"]))
        assert r.status_code == 200
        po = r.json()
        line = po["items"][0]
        inv_payload = {
            "po_id": po["po_id"],
            "vendor_id": po["vendor_id"],
            "project_id": po["project_id"],
            "invoice_ref": f"TEST_INV_{uuid.uuid4().hex[:6]}",
            "invoice_date": "2026-01-15",
            "items": [{
                "mrf_id": line["mrf_id"], "item_line_id": line["item_line_id"],
                "description": line["description"], "unit": line["unit"],
                "qty": 1, "rate": line["rate"], "gst": line.get("gst", 18)
            }],
            "notes": "TEST inv"
        }
        r = requests.post(f"{API}/invoice", json=inv_payload, headers=_hdr(ttoks["billing"]))
        assert r.status_code == 200, r.text
        inv = r.json()
        # PDF
        r = requests.get(f"{API}/invoice/{inv['invoice_id']}/pdf?token={ttoks['admin']}")
        assert r.status_code == 200
        assert r.headers.get("content-type", "").startswith("application/pdf")

    def test_dashboard_admin(self, tokens):
        r = requests.get(f"{API}/dashboard", headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code in (200, 404)  # if dashboard exists

    def test_import_vendors_valid_xlsx(self, tokens):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "address", "gstin", "contact", "email"])
        ws.append([f"TEST_VNDR_{uuid.uuid4().hex[:6]}", "addr", "gst123", "9998887777", "v@x.com"])
        buf = io.BytesIO()
        wb.save(buf)
        files = {"file": ("vendors.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/import/vendors", files=files,
                          headers=_hdr(tokens["tokens"]["admin"]))
        assert r.status_code == 200
        j = r.json()
        assert j["inserted"] >= 1
