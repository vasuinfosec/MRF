"""Iteration 19 — Backend regression for GRN export dispatcher, Delivery Challan
(DC) as a first-class document, and Comparative Statement (CS) import-only feature.

Also verifies audit_logs row is written with details.old_value / details.new_value
for every create/edit/cancel/upload/download op.

Requires ENABLE_DEV_LOGIN=1 (dev-login endpoint gated by env).
BASE_URL resolves from EXPO_PUBLIC_BACKEND_URL.
"""
from __future__ import annotations

import base64
import io
import os
import uuid
from typing import Any, Dict, List, Optional

import pytest
import requests

# openpyxl for xlsx validation
try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


BASE_URL = (os.environ.get("EXPO_PUBLIC_BACKEND_URL")
            or os.environ.get("EXPO_BACKEND_URL"))
assert BASE_URL, "EXPO_PUBLIC_BACKEND_URL / EXPO_BACKEND_URL must be set"
BASE_URL = BASE_URL.rstrip("/")
API = f"{BASE_URL}/api"


# ---------------- helpers ----------------
def _login(email: str, role: str, name: str = "Iter19") -> str:
    r = requests.post(f"{API}/auth/dev-login",
                      json={"email": email, "role": role, "name": name},
                      timeout=15)
    assert r.status_code == 200, r.text
    return r.json()["session_token"]


def _hdr(t: str) -> Dict[str, str]:
    return {"Authorization": f"Bearer {t}", "Content-Type": "application/json"}


def _me(t: str) -> Dict[str, Any]:
    r = requests.get(f"{API}/auth/me", headers=_hdr(t), timeout=10)
    assert r.status_code == 200, r.text
    j = r.json()
    return j if "user_id" in j else j.get("user", j)


def _latest_audit(t: str, entity_id: str, limit: int = 5) -> List[Dict[str, Any]]:
    r = requests.get(f"{API}/audit",
                     params={"entity_id": entity_id, "limit": limit},
                     headers=_hdr(t), timeout=15)
    assert r.status_code == 200, r.text
    j = r.json()
    # Some /api/audit endpoints wrap in {items:[...]}; adapt:
    return j.get("items", j) if isinstance(j, dict) else j


# ---------------- fixtures ----------------
@pytest.fixture(scope="module")
def tokens() -> Dict[str, str]:
    return {
        "admin":    _login("admin@vasu.dev",    "admin",         "Iter19 Admin"),
        "director": _login("director@vasu.dev", "director",      "Iter19 Director"),
        "gm":       _login("gm@vasu.dev",       "gm",            "Iter19 GM"),
        "pm":       _login("pm@vasu.dev",       "pm",            "Iter19 PM"),
        "purchase": _login("purchase@vasu.dev", "purchase",      "Iter19 Purchase"),
        "se":       _login("siteeng@vasu.dev",  "site_engineer", "Iter19 SE"),
        "store":    _login("store@vasu.dev",    "store",         "Iter19 Store"),
    }


@pytest.fixture(scope="module")
def uid(tokens) -> Dict[str, str]:
    return {k: _me(t)["user_id"] for k, t in tokens.items()}


@pytest.fixture(scope="module")
def masters(tokens) -> Dict[str, Any]:
    r = requests.get(f"{API}/masters", headers=_hdr(tokens["admin"]), timeout=10)
    assert r.status_code == 200, r.text
    d = r.json()
    assert d.get("material"), "need >=1 material master"
    return d


@pytest.fixture(scope="module")
def project_a(tokens, uid) -> Dict[str, Any]:
    """Project with pm=pm@vasu.dev and se=siteeng@vasu.dev in the team."""
    cid = f"TESTCUS19-{uuid.uuid4().hex[:6].upper()}"
    r = requests.post(f"{API}/customers", headers=_hdr(tokens["admin"]),
                      json={"customer_id": cid,
                            "name": f"TEST Cust {cid}",
                            "gstin": "27ABCDE1234F1Z5", "state": "MH"},
                      timeout=15)
    assert r.status_code == 200, r.text
    cust = r.json()
    pcode = f"TP-19-{uuid.uuid4().hex[:5].upper()}"
    r = requests.post(f"{API}/projects", headers=_hdr(tokens["admin"]),
                      json={"code": pcode, "name": f"TEST Proj {pcode}",
                            "site": "Pune", "client": "ClientA",
                            "customer_id": cust["customer_id"],
                            "site_engineers": [uid["se"]],
                            "project_managers": [uid["pm"]],
                            "reason": "iter19 setup"},
                      timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="module")
def project_b_no_se(tokens, uid) -> Dict[str, Any]:
    """Project WITHOUT the SE in the site_engineers list — used to test SE denial."""
    pcode = f"TP-19B-{uuid.uuid4().hex[:5].upper()}"
    r = requests.post(f"{API}/projects", headers=_hdr(tokens["admin"]),
                      json={"code": pcode, "name": f"TEST ProjB {pcode}",
                            "site": "Delhi", "client": "ClientB",
                            "project_managers": [uid["pm"]],  # note: SE NOT here
                            "reason": "iter19 setup no-se"},
                      timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="module")
def vendor(tokens) -> Dict[str, Any]:
    vid = f"TESTVEND19-{uuid.uuid4().hex[:6].upper()}"
    r = requests.post(f"{API}/vendors", headers=_hdr(tokens["admin"]),
                      json={"vendor_id": vid, "name": f"TEST Vendor {vid}",
                            "gstin": "27AABCT1234A1Z5", "state": "MH",
                            "active": True},
                      timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


def _mk_mrf_item(masters, qty: float = 3.0) -> Dict[str, Any]:
    mat = masters["material"][0]
    return {"description": "TEST-Iter19 item",
            "unit": "nos", "qty_requested": qty,
            "material_id": mat["item_id"],
            "priority": "normal"}


@pytest.fixture(scope="module")
def grn_ready(tokens, project_a, masters, vendor) -> Dict[str, Any]:
    """SE MRF → PM approve/send → Purchase PO → Store receives → GRN created."""
    # 1. MRF create + submit + approve + send
    body = {"project_id": project_a["project_id"],
            "site": project_a.get("site", "Pune"),
            "required_by": "2026-03-15",
            "requesting_person": "Iter19",
            "system_category": "electrical",
            "items": [_mk_mrf_item(masters)],
            "remarks": "iter19"}
    r = requests.post(f"{API}/mrf", headers=_hdr(tokens["se"]), json=body, timeout=15)
    assert r.status_code == 200, f"mrf create: {r.status_code} {r.text}"
    mrf = r.json()
    mid = mrf["mrf_id"]
    r = requests.post(f"{API}/mrf/{mid}/submit", headers=_hdr(tokens["se"]), timeout=15)
    assert r.status_code == 200, f"submit: {r.status_code} {r.text}"
    r = requests.post(f"{API}/mrf/{mid}/approve",
                      headers=_hdr(tokens["pm"]),
                      json={"action": "approve", "comment": "ok"}, timeout=15)
    assert r.status_code == 200, f"approve: {r.status_code} {r.text}"
    r = requests.post(f"{API}/mrf/{mid}/send-to-purchase",
                      headers=_hdr(tokens["pm"]), timeout=15)
    assert r.status_code == 200, f"send-to-purchase: {r.status_code} {r.text}"
    r = requests.get(f"{API}/mrf/{mid}", headers=_hdr(tokens["admin"]), timeout=15)
    assert r.status_code == 200, r.text
    approved_mrf = r.json()
    line = approved_mrf["items"][0]

    # 2. PO with small value (auto-issued)
    body = {"vendor_id": vendor["vendor_id"],
            "project_id": project_a["project_id"],
            "delivery_site": project_a.get("site", "Pune"),
            "items": [{"mrf_id": mid, "item_line_id": line["item_line_id"],
                       "description": line["description"], "unit": line["unit"],
                       "qty": line.get("qty_approved") or line["qty_requested"],
                       "rate": 100.0, "discount": 0, "gst": 18, "hsn_code": "8536"}],
            "delivery_schedule": "asap", "payment_terms": "net 30"}
    r = requests.post(f"{API}/po", headers=_hdr(tokens["purchase"]), json=body, timeout=15)
    assert r.status_code == 200, r.text
    po = r.json()

    # 3. Purchase receives (has global access; store would need to be in project SE list)
    q_total = float(line.get("qty_approved") or line["qty_requested"])
    r = requests.post(f"{API}/po/{po['po_id']}/received",
                      headers=_hdr(tokens["purchase"]),
                      json={"items": [{"mrf_id": mid,
                                        "item_line_id": line["item_line_id"],
                                        "qty": q_total}]},
                      timeout=15)
    assert r.status_code == 200, f"po receive: {r.status_code} {r.text}"
    j = r.json()
    assert j.get("grn_id"), j
    return {"mrf_id": mid, "po_id": po["po_id"], "grn_id": j["grn_id"],
            "grn_number": j["grn_number"], "project_id": project_a["project_id"]}


# =====================================================================
# 1. GRN /api/grn/{id}/export dispatcher
# =====================================================================
class TestGRNExport:
    def test_pdf_returns_pdf_bytes(self, tokens, grn_ready):
        r = requests.get(f"{API}/grn/{grn_ready['grn_id']}/export",
                         params={"format": "pdf"},
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 200, r.text
        assert r.content[:4] == b"%PDF", "expected %PDF magic header"

    def test_excel_returns_valid_xlsx(self, tokens, grn_ready):
        r = requests.get(f"{API}/grn/{grn_ready['grn_id']}/export",
                         params={"format": "excel"},
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 200, r.text
        # xlsx = zip = PK\x03\x04 signature
        assert r.content[:2] == b"PK", "expected xlsx PK signature"
        if openpyxl is not None:
            wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            assert "MAT UID" in headers and "VAR UID" in headers, headers

    def test_export_audit_logged(self, tokens, grn_ready):
        # trigger + inspect fresh audit for THIS entity_id
        requests.get(f"{API}/grn/{grn_ready['grn_id']}/export",
                     params={"format": "excel"},
                     headers=_hdr(tokens["purchase"]), timeout=30)
        rows = _latest_audit(tokens["admin"], grn_ready["grn_id"], limit=5)
        assert any(r["action"].startswith("export_") for r in rows), rows

    def test_bulk_export_grn_roles(self, tokens):
        allowed = ["purchase", "director", "admin", "store", "pm", "gm"]
        for role in allowed:
            r = requests.get(f"{API}/export/grn",
                             headers=_hdr(tokens[role]), timeout=30)
            assert r.status_code == 200, f"{role}: {r.status_code} {r.text[:120]}"
        # site_engineer forbidden
        r = requests.get(f"{API}/export/grn",
                         headers=_hdr(tokens["se"]), timeout=30)
        assert r.status_code == 403, r.text

    def test_zero_qty_validation_and_force(self, tokens, project_a, grn_ready):
        """Inject a temp GRN with qty=0 line via pymongo, then verify:
            - normal export → 422 with missing_mandatory
            - ?force=1 → 200 + audit-log details.forced=true
        Falls back to skip if pymongo/local mongo not reachable."""
        try:
            from pymongo import MongoClient
            from datetime import datetime, timezone
        except Exception:
            pytest.skip("pymongo not available")
        mongo_url = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
        db_name = os.environ.get("DB_NAME", "test_database")
        try:
            cli = MongoClient(mongo_url, serverSelectionTimeoutMS=2000)
            cli.server_info()
        except Exception as e:
            pytest.skip(f"Local mongo not reachable: {e}")

        db = cli[db_name]
        grn_id = f"grn_TEST19_{uuid.uuid4().hex[:8]}"
        db.grns.insert_one({
            "grn_id": grn_id,
            "grn_number": f"GRN/TEST/{uuid.uuid4().hex[:4].upper()}",
            "po_id": grn_ready["po_id"],
            "po_number": "TEST-PO",
            "mrf_refs": [grn_ready["mrf_id"]],
            "vendor_id": "TESTVEND19",
            "project_id": project_a["project_id"],
            "date": datetime.now(timezone.utc),
            "items": [{"mrf_id": grn_ready["mrf_id"],
                       "item_line_id": "tl_1",
                       "description": "zero-qty test",
                       "unit": "nos", "qty": 0}],
            "received_by": "u", "received_by_name": "Iter19",
        })
        try:
            r = requests.get(f"{API}/grn/{grn_id}/export",
                             params={"format": "excel"},
                             headers=_hdr(tokens["admin"]), timeout=30)
            assert r.status_code == 422, f"expected 422, got {r.status_code}: {r.text[:200]}"
            j = r.json()
            missing = (j.get("detail") or {}).get("missing_mandatory") or j.get("missing_mandatory") or []
            assert any("qty" in m for m in missing), j
            # force=1 should bypass
            r = requests.get(f"{API}/grn/{grn_id}/export",
                             params={"format": "excel", "force": 1},
                             headers=_hdr(tokens["admin"]), timeout=30)
            assert r.status_code == 200, r.text
            rows = _latest_audit(tokens["admin"], grn_id, limit=5)
            forced_hit = any((r.get("details") or {}).get("forced") is True for r in rows)
            assert forced_hit, rows
        finally:
            db.grns.delete_one({"grn_id": grn_id})
            db.audit_logs.delete_many({"entity_id": grn_id})


# =====================================================================
# 2. Delivery Challan
# =====================================================================
@pytest.fixture
def dc_body(project_a) -> Dict[str, Any]:
    return {"dc_type": "outbound",
            "project_id": project_a["project_id"],
            "from_location": "Pune Warehouse",
            "to_location": "Delhi Site",
            "dispatch_date": "2026-02-05",
            "vehicle_no": "MH12 AB 1234",
            "driver_name": "Ravi K",
            "items": [{"description": "Cable 4c 2.5sqmm",
                       "unit": "mtr", "qty": 100.0,
                       "material_uid": "MAT-CABLE",
                       "variant_uid": "VAR-2.5"}],
            "remarks": "iter19 dc test"}


class TestDCCreate:
    def test_create_dc_success(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        assert r.status_code == 200, r.text
        j = r.json()
        assert j.get("dc_id") and j.get("dc_number"), j
        assert j["dc_number"].startswith("DC/"), j["dc_number"]
        parts = j["dc_number"].split("/")
        assert len(parts) == 3 and parts[1].isdigit() and parts[2].isdigit(), j
        # Audit
        rows = _latest_audit(tokens["admin"], j["dc_id"], limit=5)
        create_row = next((x for x in rows if x["action"] == "create"), None)
        assert create_row, rows
        d = create_row["details"]
        assert d.get("old_value") is None
        assert d["new_value"]["item_count"] == 1

    def test_dc_number_increments(self, tokens, dc_body):
        r1 = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                           json=dc_body, timeout=15)
        assert r1.status_code == 200, r1.text
        r2 = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                           json=dc_body, timeout=15)
        assert r2.status_code == 200, r2.text
        n1 = int(r1.json()["dc_number"].split("/")[-1])
        n2 = int(r2.json()["dc_number"].split("/")[-1])
        assert n2 == n1 + 1, (r1.json()["dc_number"], r2.json()["dc_number"])

    def test_missing_items_400(self, tokens, dc_body):
        b = dict(dc_body); b["items"] = []
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=b, timeout=15)
        assert r.status_code == 400, r.text

    def test_missing_project_404(self, tokens, dc_body):
        b = dict(dc_body); b["project_id"] = "bogus-project"
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=b, timeout=15)
        assert r.status_code == 404, r.text

    def test_se_denied_when_not_in_project_se_list(self, tokens, project_b_no_se, dc_body):
        b = dict(dc_body); b["project_id"] = project_b_no_se["project_id"]
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["se"]),
                          json=b, timeout=15)
        assert r.status_code == 403, r.text


class TestDCListScoping:
    def test_pm_sees_only_own_project(self, tokens, uid, project_a, dc_body):
        # create a DC in project_a (pm is in project_a's PMs)
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        assert r.status_code == 200
        r = requests.get(f"{API}/dc", headers=_hdr(tokens["pm"]), timeout=15)
        assert r.status_code == 200, r.text
        rows = r.json()
        # PM must see the newly created DC in project_a
        assert any(row["project_id"] == project_a["project_id"] for row in rows), rows
        # Every returned DC belongs to a project where pm is a project_manager
        pids = {row["project_id"] for row in rows}
        if pids:
            prjs = requests.get(f"{API}/projects",
                                 headers=_hdr(tokens["admin"]), timeout=15).json()
            pm_prjs = {p["project_id"] for p in prjs
                        if uid["pm"] in (p.get("project_managers") or [])}
            assert pids.issubset(pm_prjs), (pids - pm_prjs)

    def test_se_scoped_to_projects_they_belong_to(self, tokens, uid, project_a):
        r = requests.get(f"{API}/dc", headers=_hdr(tokens["se"]), timeout=15)
        assert r.status_code == 200, r.text
        rows = r.json()
        pids = {row["project_id"] for row in rows}
        if pids:
            prjs = requests.get(f"{API}/projects",
                                 headers=_hdr(tokens["admin"]), timeout=15).json()
            se_prjs = {p["project_id"] for p in prjs
                        if uid["se"] in (p.get("site_engineers") or [])}
            assert pids.issubset(se_prjs), (pids - se_prjs)

    def test_status_filter(self, tokens):
        r = requests.get(f"{API}/dc", params={"status": "issued"},
                         headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        assert all(row["status"] == "issued" for row in r.json())

    def test_dc_type_filter(self, tokens):
        r = requests.get(f"{API}/dc", params={"dc_type": "outbound"},
                         headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        assert all(row["dc_type"] == "outbound" for row in r.json())


class TestDCUpdate:
    def test_put_allowed_fields_and_audit(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        assert r.status_code == 200
        dc_id = r.json()["dc_id"]
        r = requests.put(f"{API}/dc/{dc_id}",
                         headers=_hdr(tokens["purchase"]),
                         json={"vehicle_no": "MH12 XY 9999",
                               "driver_name": "Suresh",
                               "status": "in_transit",
                               "remarks": "updated"},
                         timeout=15)
        assert r.status_code == 200, r.text
        j = r.json()
        assert j["vehicle_no"] == "MH12 XY 9999"
        rows = _latest_audit(tokens["admin"], dc_id, limit=8)
        edit_row = next((x for x in rows if x["action"] == "edit"), None)
        assert edit_row, rows
        d = edit_row["details"]
        assert d["old_value"]["vehicle_no"] == "MH12 AB 1234"
        assert d["new_value"]["vehicle_no"] == "MH12 XY 9999"

    def test_put_only_unknown_fields_400(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        dc_id = r.json()["dc_id"]
        r = requests.put(f"{API}/dc/{dc_id}",
                         headers=_hdr(tokens["purchase"]),
                         json={"__unknown_field__": "x"}, timeout=15)
        assert r.status_code == 400, r.text
        assert "Nothing to update" in r.text


class TestDCExport:
    def test_pdf_ok(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        dc_id = r.json()["dc_id"]
        r = requests.get(f"{API}/dc/{dc_id}/pdf",
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 200, r.text
        assert r.content[:4] == b"%PDF"

    def test_pdf_missing_mandatory_422(self, tokens, dc_body):
        b = dict(dc_body); b["from_location"] = ""; b["to_location"] = ""
        b["dispatch_date"] = ""
        # Post via purchase and re-fetch — from/to blank may still create,
        # let's send via PUT to force-blank scenario.
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        dc_id = r.json()["dc_id"]
        requests.put(f"{API}/dc/{dc_id}",
                     headers=_hdr(tokens["purchase"]),
                     json={"from_location": "", "to_location": "",
                           "dispatch_date": ""}, timeout=15)
        r = requests.get(f"{API}/dc/{dc_id}/pdf",
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 422, r.text
        j = r.json()
        missing = (j.get("detail") or {}).get("missing_mandatory") or j.get("missing_mandatory") or []
        joined = " ".join(missing).lower()
        assert "from" in joined and "to" in joined and "dispatch" in joined, missing

        # force=1 bypass
        r = requests.get(f"{API}/dc/{dc_id}/pdf",
                         params={"force": 1},
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 200, r.text

    def test_export_excel_headers(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        dc_id = r.json()["dc_id"]
        r = requests.get(f"{API}/dc/{dc_id}/export",
                         params={"format": "excel"},
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 200, r.text
        if openpyxl is not None:
            wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            assert len(headers) > 20, f"expected >20 headers, got {len(headers)}"
            assert "MAT UID" in headers and "VAR UID" in headers, headers

    def test_export_pdf_dispatch(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        dc_id = r.json()["dc_id"]
        r = requests.get(f"{API}/dc/{dc_id}/export",
                         params={"format": "pdf"},
                         headers=_hdr(tokens["purchase"]), timeout=30)
        assert r.status_code == 200, r.text
        assert r.content[:4] == b"%PDF"


class TestDCDelete:
    def test_non_admin_forbidden(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        dc_id = r.json()["dc_id"]
        for role in ("purchase", "pm", "gm", "store", "se"):
            r = requests.delete(f"{API}/dc/{dc_id}",
                                headers=_hdr(tokens[role]), timeout=15)
            assert r.status_code == 403, f"{role}: {r.status_code} {r.text}"

    def test_admin_can_cancel(self, tokens, dc_body):
        r = requests.post(f"{API}/dc", headers=_hdr(tokens["purchase"]),
                          json=dc_body, timeout=15)
        dc_id = r.json()["dc_id"]
        r = requests.delete(f"{API}/dc/{dc_id}?reason=iter19-test",
                            headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        # Verify deleted + status=cancelled by direct GET
        r = requests.get(f"{API}/dc/{dc_id}",
                         headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200
        d = r.json()
        assert d.get("deleted") is True
        assert d.get("status") == "cancelled"
        # Audit
        rows = _latest_audit(tokens["admin"], dc_id, limit=5)
        cancel_row = next((x for x in rows if x["action"] == "cancel"), None)
        assert cancel_row, rows
        assert cancel_row["details"]["new_value"]["status"] == "cancelled"


# =====================================================================
# 3. Comparative Statement
# =====================================================================
def _sample_b64(size_bytes: int = 200) -> str:
    return base64.b64encode(os.urandom(size_bytes)).decode("ascii")


@pytest.fixture
def mrf_owned_by_se(tokens, project_a, masters) -> Dict[str, Any]:
    body = {"project_id": project_a["project_id"],
            "site": project_a.get("site", "Pune"),
            "required_by": "2026-04-01",
            "requesting_person": "Iter19-CS",
            "system_category": "electrical",
            "items": [_mk_mrf_item(masters)],
            "remarks": "iter19-cs"}
    r = requests.post(f"{API}/mrf", headers=_hdr(tokens["se"]), json=body, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


class TestCSUpload:
    def test_missing_file_400(self, tokens, mrf_owned_by_se):
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS T", "file_name": "cs.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": ""},
                          timeout=15)
        assert r.status_code == 400, r.text
        assert "file_base64" in r.text.lower()

    def test_missing_link_400(self, tokens):
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"title": "CS T", "file_name": "cs.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": _sample_b64(100)},
                          timeout=15)
        assert r.status_code == 400, r.text
        assert "mrf_id" in r.text or "po_id" in r.text

    def test_oversize_413(self, tokens, mrf_owned_by_se):
        # >15 MB raw ⇒ ~ 20 MB base64
        big = _sample_b64(16 * 1024 * 1024)
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS BIG", "file_name": "big.bin",
                                "mime_type": "application/octet-stream",
                                "file_base64": big},
                          timeout=60)
        assert r.status_code == 413, r.status_code

    def test_success_no_base64_in_response(self, tokens, mrf_owned_by_se):
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS Good", "file_name": "cs.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": _sample_b64(400)},
                          timeout=15)
        assert r.status_code == 200, r.text
        j = r.json()
        assert "file_base64" not in j, j.keys()
        assert j["cs_number"].startswith("CS/"), j["cs_number"]
        parts = j["cs_number"].split("/")
        assert len(parts) == 3 and parts[1].isdigit() and parts[2].isdigit()
        # Audit
        rows = _latest_audit(tokens["admin"], j["cs_id"], limit=5)
        up = next((x for x in rows if x["action"] == "upload"), None)
        assert up, rows
        assert up["details"]["old_value"] is None
        assert up["details"]["new_value"]["file_name"] == "cs.xlsx"


class TestCSListDownload:
    def test_list_never_returns_file_base64(self, tokens, mrf_owned_by_se):
        # Ensure at least one exists
        requests.post(f"{API}/comparative-statement",
                      headers=_hdr(tokens["purchase"]),
                      json={"mrf_id": mrf_owned_by_se["mrf_id"],
                            "title": "CS L", "file_name": "cs.xlsx",
                            "mime_type": "application/vnd.ms-excel",
                            "file_base64": _sample_b64(300)},
                      timeout=15)
        r = requests.get(f"{API}/comparative-statement",
                         headers=_hdr(tokens["purchase"]), timeout=15)
        assert r.status_code == 200, r.text
        for row in r.json():
            assert "file_base64" not in row, row

    def test_se_sees_only_own_mrf_cs(self, tokens, mrf_owned_by_se):
        # Purchase uploads a CS attached to SE's MRF
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS SE",
                                "file_name": "cs.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": _sample_b64(200)},
                          timeout=15)
        assert r.status_code == 200
        r = requests.get(f"{API}/comparative-statement",
                         headers=_hdr(tokens["se"]), timeout=15)
        assert r.status_code == 200, r.text
        rows = r.json()
        # Every row must correspond to an MRF whose created_by == se's user_id
        for row in rows:
            assert row.get("mrf_id"), row
            # if list is not empty, mrf must be visible to SE
            r2 = requests.get(f"{API}/mrf/{row['mrf_id']}",
                              headers=_hdr(tokens["se"]), timeout=15)
            assert r2.status_code == 200, r2.text

    def test_download_ok_and_audit(self, tokens, mrf_owned_by_se):
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS DL",
                                "file_name": "csdl.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": _sample_b64(500)},
                          timeout=15)
        cs_id = r.json()["cs_id"]
        r = requests.get(f"{API}/comparative-statement/{cs_id}/download",
                         headers=_hdr(tokens["purchase"]), timeout=15)
        assert r.status_code == 200, r.text
        assert r.headers.get("content-type", "").startswith("application/"), r.headers
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd and "csdl.xlsx" in cd, cd
        rows = _latest_audit(tokens["admin"], cs_id, limit=8)
        assert any(r["action"] == "export_download" for r in rows), rows


class TestCSUpdateDelete:
    def test_put_whitelist(self, tokens, mrf_owned_by_se):
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS Upd",
                                "file_name": "csu.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": _sample_b64(300),
                                "l1_vendor": "V1", "l1_amount": 100.0,
                                "selected_vendor": "V1"},
                          timeout=15)
        cs_id = r.json()["cs_id"]
        r = requests.put(f"{API}/comparative-statement/{cs_id}",
                         headers=_hdr(tokens["purchase"]),
                         json={"title": "CS Upd v2",
                               "l1_vendor": "V2",
                               "__unknown_field__": "no",
                               "reason": "correction"},
                         timeout=15)
        assert r.status_code == 200, r.text
        j = r.json()
        assert j["title"] == "CS Upd v2"
        assert j["l1_vendor"] == "V2"
        assert "__unknown_field__" not in j
        rows = _latest_audit(tokens["admin"], cs_id, limit=8)
        edit = next((x for x in rows if x["action"] == "edit"), None)
        assert edit, rows
        assert edit["details"]["old_value"]["title"] == "CS Upd"
        assert edit["details"]["new_value"]["title"] == "CS Upd v2"

    def test_put_only_unknown_400(self, tokens, mrf_owned_by_se):
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS Blank",
                                "file_name": "csb.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": _sample_b64(300)},
                          timeout=15)
        cs_id = r.json()["cs_id"]
        r = requests.put(f"{API}/comparative-statement/{cs_id}",
                         headers=_hdr(tokens["purchase"]),
                         json={"__nope__": 1}, timeout=15)
        assert r.status_code == 400, r.text

    def test_delete_admin_only(self, tokens, mrf_owned_by_se):
        r = requests.post(f"{API}/comparative-statement",
                          headers=_hdr(tokens["purchase"]),
                          json={"mrf_id": mrf_owned_by_se["mrf_id"],
                                "title": "CS Del",
                                "file_name": "csd.xlsx",
                                "mime_type": "application/vnd.ms-excel",
                                "file_base64": _sample_b64(200)},
                          timeout=15)
        cs_id = r.json()["cs_id"]
        for role in ("purchase", "pm", "gm", "store", "se"):
            r = requests.delete(f"{API}/comparative-statement/{cs_id}",
                                headers=_hdr(tokens[role]), timeout=15)
            assert r.status_code == 403, f"{role}: {r.status_code}"
        r = requests.delete(f"{API}/comparative-statement/{cs_id}",
                            headers=_hdr(tokens["admin"]), timeout=15)
        assert r.status_code == 200, r.text
        rows = _latest_audit(tokens["admin"], cs_id, limit=5)
        assert any(x["action"] == "delete" for x in rows), rows


# =====================================================================
# 4. Regression: PO/MRF export + audit RBAC
# =====================================================================
class TestRegression:
    def test_po_export_still_works(self, tokens, grn_ready):
        for fmt in ("pdf", "excel", "tally"):
            r = requests.get(f"{API}/po/{grn_ready['po_id']}/export",
                             params={"format": fmt, "force": 1},
                             headers=_hdr(tokens["purchase"]), timeout=30)
            assert r.status_code == 200, f"{fmt}: {r.status_code} {r.text[:200]}"

    def test_mrf_export_still_works(self, tokens, grn_ready):
        r = requests.get(f"{API}/mrf/{grn_ready['mrf_id']}/export",
                         params={"format": "pdf", "force": 1},
                         headers=_hdr(tokens["pm"]), timeout=30)
        assert r.status_code == 200, r.text
        assert r.content[:4] == b"%PDF"

    def test_audit_role_visibility(self, tokens):
        for role in ("gm", "purchase", "pm", "site_engineer", "store"):
            key = "se" if role == "site_engineer" else role
            r = requests.get(f"{API}/audit", headers=_hdr(tokens[key]), timeout=15)
            assert r.status_code == 200, f"{role}: {r.status_code} {r.text[:120]}"

    def test_audit_facets_roles(self, tokens):
        for role in ("admin", "director", "gm", "purchase", "pm"):
            r = requests.get(f"{API}/audit/facets",
                             headers=_hdr(tokens[role]), timeout=15)
            assert r.status_code == 200, f"{role}: {r.status_code}"
