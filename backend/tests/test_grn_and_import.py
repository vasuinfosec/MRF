"""
Backend tests for new features:
  * GRN (Goods Received Note) creation on PO receipt
  * Per-line `receipts` history array
  * GRN listing (per PO, and global) + GRN PDF download
  * Excel bulk import for vendors and MRFs (template download + upload)
"""
import io
import os
import pytest
import requests
import openpyxl

BASE_URL = "https://purchase-workflow-10.preview.emergentagent.com"


def _login(role, email, name):
    r = requests.post(f"{BASE_URL}/api/auth/dev-login",
                      json={"email": email, "role": role, "name": name}, timeout=20)
    assert r.status_code == 200, r.text
    return r.json()["session_token"]


def _h(t):
    return {"Authorization": f"Bearer {t}"}


@pytest.fixture(scope="module")
def tokens():
    admin_tok = _login("admin", "admin@vasu.dev", "Director")
    r = requests.post(f"{BASE_URL}/api/seed", headers=_h(admin_tok), timeout=30)
    assert r.status_code == 200
    return {
        "site": _login("site_engineer", "site@vasu.dev", "Ravi"),
        "pm": _login("project_manager", "pm@vasu.dev", "Priya"),
        "purchase": _login("purchase", "purchase@vasu.dev", "Kumar"),
        "billing": _login("billing", "billing@vasu.dev", "Anita"),
        "admin": admin_tok,
    }


@pytest.fixture(scope="module")
def po_context(tokens):
    """Create a fresh MRF → approve → send-to-purchase → PO so tests below have a live PO."""
    projects = requests.get(f"{BASE_URL}/api/projects", headers=_h(tokens["site"])).json()
    vendors = requests.get(f"{BASE_URL}/api/vendors", headers=_h(tokens["purchase"])).json()
    proj = projects[0]
    body = {
        "project_id": proj["project_id"], "site": proj["site"],
        "required_by": "2026-04-01", "requesting_person": "TEST_GRN_Ravi",
        "system_category": "Fire Alarm",
        "items": [
            {"description": "TEST_GRN_Detector", "unit": "Nos", "qty_requested": 10, "billable": True},
            {"description": "TEST_GRN_Sounder", "unit": "Nos", "qty_requested": 6, "billable": True},
        ],
    }
    r = requests.post(f"{BASE_URL}/api/mrf", headers=_h(tokens["site"]), json=body)
    assert r.status_code == 200, r.text
    mrf = r.json()
    mrf_id = mrf["mrf_id"]
    requests.post(f"{BASE_URL}/api/mrf/{mrf_id}/submit", headers=_h(tokens["site"])).raise_for_status()

    item_actions = [{"item_line_id": it["item_line_id"], "action": "approve",
                     "qty_approved": it["qty_requested"]} for it in mrf["items"]]
    requests.post(f"{BASE_URL}/api/mrf/{mrf_id}/approve", headers=_h(tokens["pm"]),
                  json={"action": "approve", "comment": "OK", "item_actions": item_actions}).raise_for_status()
    requests.post(f"{BASE_URL}/api/mrf/{mrf_id}/send-to-purchase",
                  headers=_h(tokens["pm"])).raise_for_status()

    # Refresh MRF to get updated item_line_ids/approved state
    mrf = requests.get(f"{BASE_URL}/api/mrf/{mrf_id}", headers=_h(tokens["purchase"])).json()
    approved = [it for it in mrf["items"] if it["status"] == "approved"]

    po_body = {
        "vendor_id": vendors[0]["vendor_id"],
        "project_id": proj["project_id"],
        "delivery_site": "TEST GRN site",
        "items": [{
            "mrf_id": mrf_id, "item_line_id": approved[0]["item_line_id"],
            "description": approved[0]["description"], "unit": "Nos",
            "qty": 10, "rate": 100, "gst": 18, "discount": 0
        }, {
            "mrf_id": mrf_id, "item_line_id": approved[1]["item_line_id"],
            "description": approved[1]["description"], "unit": "Nos",
            "qty": 6, "rate": 200, "gst": 18, "discount": 0
        }],
        "freight": 0, "other_charges": 0,
    }
    r = requests.post(f"{BASE_URL}/api/po", headers=_h(tokens["purchase"]), json=po_body)
    assert r.status_code == 200, r.text
    po = r.json()
    return {"mrf_id": mrf_id, "po_id": po["po_id"], "po_number": po["po_number"],
            "items": approved}


# ---------- GRN + Receipts ----------
class TestGRNCreation:
    def test_partial_receive_returns_grn_ids(self, tokens, po_context):
        """First receipt: partial qty on both lines → creates a GRN and returns grn_id + grn_number."""
        items = po_context["items"]
        body = {
            "items": [
                {"mrf_id": po_context["mrf_id"], "item_line_id": items[0]["item_line_id"], "qty": 4},
                {"mrf_id": po_context["mrf_id"], "item_line_id": items[1]["item_line_id"], "qty": 2},
            ],
            "remarks": "TEST_partial_1"
        }
        r = requests.post(f"{BASE_URL}/api/po/{po_context['po_id']}/received",
                          headers=_h(tokens["purchase"]), json=body)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["ok"] is True
        assert d["status"] == "partially_received"
        assert d.get("grn_id"), "grn_id missing from response"
        assert d.get("grn_number", "").startswith("GRN/"), f"grn_number bad: {d.get('grn_number')}"
        # Track first GRN
        po_context["grn_id_1"] = d["grn_id"]
        po_context["grn_number_1"] = d["grn_number"]

    def test_receipts_array_populated_on_po_items(self, tokens, po_context):
        r = requests.get(f"{BASE_URL}/api/po/{po_context['po_id']}", headers=_h(tokens["purchase"]))
        assert r.status_code == 200
        po = r.json()
        for pit in po["items"]:
            assert isinstance(pit.get("receipts"), list) and len(pit["receipts"]) >= 1
            recv = pit["receipts"][0]
            assert set(["date", "qty", "user_id", "user_name"]).issubset(recv.keys())
            assert recv["qty"] > 0
            assert recv["user_name"]  # non-empty

    def test_second_partial_receipt_creates_second_grn(self, tokens, po_context):
        items = po_context["items"]
        body = {
            "items": [
                {"mrf_id": po_context["mrf_id"], "item_line_id": items[0]["item_line_id"], "qty": 6},  # completes line1
                {"mrf_id": po_context["mrf_id"], "item_line_id": items[1]["item_line_id"], "qty": 4},  # completes line2
            ],
            "remarks": "TEST_final_receipt"
        }
        r = requests.post(f"{BASE_URL}/api/po/{po_context['po_id']}/received",
                          headers=_h(tokens["purchase"]), json=body)
        assert r.status_code == 200
        d = r.json()
        assert d["status"] == "received", f"expected fully received, got {d['status']}"
        assert d.get("grn_id")
        assert d.get("grn_number", "").startswith("GRN/")
        po_context["grn_id_2"] = d["grn_id"]

    def test_over_receive_is_clamped_and_no_grn(self, tokens, po_context):
        """After full receipt, sending another receive with zero-outstanding should not create a GRN."""
        items = po_context["items"]
        body = {"items": [{"mrf_id": po_context["mrf_id"],
                           "item_line_id": items[0]["item_line_id"], "qty": 100}]}
        r = requests.post(f"{BASE_URL}/api/po/{po_context['po_id']}/received",
                          headers=_h(tokens["purchase"]), json=body)
        assert r.status_code == 200
        d = r.json()
        # Nothing outstanding → no GRN should be created
        assert d.get("grn_id") is None
        assert d.get("grn_number") is None


class TestGRNList:
    def test_list_grns_for_po(self, tokens, po_context):
        r = requests.get(f"{BASE_URL}/api/po/{po_context['po_id']}/grns",
                         headers=_h(tokens["purchase"]))
        assert r.status_code == 200
        grns = r.json()
        assert isinstance(grns, list) and len(grns) >= 2
        # Should be reverse chronological
        dates = [g["date"] for g in grns]
        assert dates == sorted(dates, reverse=True)
        for g in grns:
            assert g["po_id"] == po_context["po_id"]
            assert g["po_number"] == po_context["po_number"]
            assert isinstance(g["items"], list) and len(g["items"]) >= 1

    def test_list_all_grns(self, tokens):
        r = requests.get(f"{BASE_URL}/api/grns", headers=_h(tokens["admin"]))
        assert r.status_code == 200
        assert isinstance(r.json(), list) and len(r.json()) >= 1

    def test_grn_list_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/grns")
        assert r.status_code == 401


class TestGRNPdf:
    def test_grn_pdf_download_with_token(self, tokens, po_context):
        gid = po_context.get("grn_id_1")
        assert gid, "grn_id_1 fixture missing"
        r = requests.get(f"{BASE_URL}/api/grn/{gid}/pdf?token={tokens['purchase']}")
        assert r.status_code == 200, r.text[:200]
        assert r.headers["content-type"].startswith("application/pdf")
        assert r.content[:4] == b"%PDF"

    def test_grn_pdf_missing_returns_404(self, tokens):
        r = requests.get(f"{BASE_URL}/api/grn/does-not-exist/pdf?token={tokens['purchase']}")
        assert r.status_code == 404

    def test_grn_pdf_without_auth(self):
        r = requests.get(f"{BASE_URL}/api/grn/anything/pdf")
        assert r.status_code == 401


# ---------- Excel Import Templates ----------
class TestImportTemplates:
    def test_vendor_template_download(self, tokens):
        r = requests.get(f"{BASE_URL}/api/import/vendors/template?token={tokens['admin']}")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert r.content[:2] == b"PK"
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        hdr = [c.value for c in wb.active[1]]
        assert "name" in hdr and "gstin" in hdr

    def test_mrf_template_download(self, tokens):
        r = requests.get(f"{BASE_URL}/api/import/mrf/template?token={tokens['admin']}")
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        hdr = [c.value for c in wb.active[1]]
        for k in ["project_code", "site", "required_by", "requesting_person",
                  "system_category", "description", "qty_requested"]:
            assert k in hdr, f"missing header {k}"

    def test_templates_require_auth(self):
        assert requests.get(f"{BASE_URL}/api/import/vendors/template").status_code == 401
        assert requests.get(f"{BASE_URL}/api/import/mrf/template").status_code == 401


# ---------- Excel Import - Vendors ----------
class TestVendorImport:
    def _make_xlsx(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "address", "gstin", "contact", "email"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_vendors_inserts_new_skips_dupes(self, tokens):
        # Use unique name each run
        unique = f"TEST_Vendor_{os.urandom(4).hex()}"
        buf = self._make_xlsx([
            [unique, "Bengaluru", "29AAAAA0000A1Z5", "9999999999", "t@t.com"],
            [unique, "Bengaluru", "29AAAAA0000A1Z5", "9999999999", "t@t.com"],  # duplicate row in same file - name lookup, but second is after first insert
            ["", "no name row", "", "", ""],  # skipped: no name
        ])
        r = requests.post(f"{BASE_URL}/api/import/vendors",
                          headers=_h(tokens["admin"]),
                          files={"file": ("vendors.xlsx", buf.getvalue(),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["inserted"] == 1, f"expected 1 insert, got {d}"
        assert d["skipped"] >= 2, f"expected ≥2 skipped, got {d}"
        assert isinstance(d["errors"], list)
        # Verify the vendor exists via GET
        vendors = requests.get(f"{BASE_URL}/api/vendors", headers=_h(tokens["purchase"])).json()
        assert any(v["name"] == unique for v in vendors)

    def test_import_vendors_rejects_non_admin_non_purchase(self, tokens):
        buf = self._make_xlsx([["TEST_x", "a", "b", "c", "d"]])
        r = requests.post(f"{BASE_URL}/api/import/vendors",
                          headers=_h(tokens["site"]),
                          files={"file": ("v.xlsx", buf.getvalue(),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.status_code == 403

    def test_import_vendors_bad_file_returns_400(self, tokens):
        r = requests.post(f"{BASE_URL}/api/import/vendors",
                          headers=_h(tokens["admin"]),
                          files={"file": ("bad.xlsx", b"not-an-xlsx",
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.status_code == 400


# ---------- Excel Import - MRF ----------
class TestMRFImport:
    def _make_mrf_xlsx(self, project_code, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["project_code", "site", "required_by", "requesting_person", "system_category",
                   "description", "specification", "part_number", "unit", "qty_requested",
                   "purpose", "drawing_ref", "billable", "boq_ref", "remarks"])
        for r in rows:
            ws.append([project_code] + r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_mrf_groups_rows(self, tokens):
        projects = requests.get(f"{BASE_URL}/api/projects", headers=_h(tokens["admin"])).json()
        proj_code = projects[0]["code"]

        buf = self._make_mrf_xlsx(proj_code, [
            # Two rows same group → one MRF with 2 items
            ["Site-A", "2026-06-01", "TEST_Imp_Rk", "Fire Alarm", "TEST_IMP_Det1", "", "", "Nos", 5,
             "", "", "Y", "", ""],
            ["Site-A", "2026-06-01", "TEST_Imp_Rk", "Fire Alarm", "TEST_IMP_Det2", "", "", "Nos", 3,
             "", "", "Y", "", ""],
            # Different site → separate MRF
            ["Site-B", "2026-06-05", "TEST_Imp_Vk", "CCTV", "TEST_IMP_Cam", "", "", "Nos", 8,
             "", "", "Y", "", ""],
        ])
        r = requests.post(f"{BASE_URL}/api/import/mrf",
                          headers=_h(tokens["admin"]),
                          files={"file": ("mrf.xlsx", buf.getvalue(),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["count"] == 2, f"expected 2 MRFs created, got {d}"
        assert len(d["created"]) == 2
        assert all(mn.startswith("MRF/") for mn in d["created"])
        # Verify at least one is retrievable and has 2 items
        mrfs = requests.get(f"{BASE_URL}/api/mrf", headers=_h(tokens["admin"])).json()
        matched = [m for m in mrfs if m["mrf_number"] in d["created"]]
        assert len(matched) == 2
        by_num = {m["mrf_number"]: m for m in matched}
        two_item_mrf = next(m for m in matched if len(m["items"]) == 2)
        assert any(it["description"].startswith("TEST_IMP_Det") for it in two_item_mrf["items"])

    def test_import_mrf_reports_bad_project(self, tokens):
        buf = self._make_mrf_xlsx("NOPE-999", [
            ["Site-X", "2026-07-01", "TEST_Imp_X", "Fire Alarm", "TEST_IMP_BAD", "", "", "Nos", 1,
             "", "", "Y", "", ""],
        ])
        r = requests.post(f"{BASE_URL}/api/import/mrf",
                          headers=_h(tokens["admin"]),
                          files={"file": ("mrf.xlsx", buf.getvalue(),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert r.status_code == 200
        d = r.json()
        assert d["count"] == 0
        assert any("unknown project_code" in e for e in d["errors"])
