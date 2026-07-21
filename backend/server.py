"""
Vasu Infosec — Material Requisition & Purchase Order System
Backend: FastAPI + MongoDB + Emergent Google Auth
"""
import os
import io
import re
import uuid
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal, Dict, Any, Tuple

import httpx
from fastapi import FastAPI, APIRouter, HTTPException, Header, Request, Response, UploadFile, File, Body
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo.errors import DuplicateKeyError
from pydantic import BaseModel, Field, EmailStr, field_validator

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="Vasu Infosec MRF & PO System")
api = APIRouter(prefix="/api")

# ---------------------- Constants ----------------------
# Vasu operational chain: Site Engineer -> PM -> Purchase -> GM -> Director.
# Admin is a system-management role and NOT part of the approval chain by default.
ROLES = ["site_engineer", "pm", "purchase", "gm", "director", "admin", "store"]
LEGACY_ROLE_MAP = {
    # Historic aliases -> canonical roles. site_engineer is now a first-class role
    # and MUST NOT be mapped to pm any more.
    "project_manager": "pm",
    "billing": "purchase",
}
def _canon_role(r: str) -> str:
    return LEGACY_ROLE_MAP.get(r, r) if r in LEGACY_ROLE_MAP else r


def _ensure_roles_shape(user: dict) -> dict:
    """Task 3A.1 — DB-driven multi-role backfill.

    Ensures every user dict has both:
      * `role`  : legacy scalar (source of truth when ACCESS_SECURITY_V2=0)
      * `roles` : canonical list (source of truth when ACCESS_SECURITY_V2=1)

    Rules (non-mutating on DB, only mutates in-memory dict):
      1. If `roles` is missing / empty and `role` is set → roles = [role].
      2. If `roles` is set and `role` missing → role = roles[0].
      3. Legacy role names inside `roles` are canonicalised.
      4. Duplicates removed while preserving order.
      5. If both empty (fresh pending user) → role="", roles=[].
    """
    if not isinstance(user, dict):
        return user
    role = user.get("role") or ""
    roles = user.get("roles") or []
    if isinstance(roles, str):
        roles = [roles]
    # Canonicalise
    roles = [_canon_role(r) for r in roles if r]
    # Dedupe while preserving order
    seen = set()
    dedup = []
    for r in roles:
        if r and r not in seen:
            seen.add(r)
            dedup.append(r)
    roles = dedup
    if not roles and role:
        roles = [_canon_role(role)]
    if not role and roles:
        role = roles[0]
    user["role"] = role or ""
    user["roles"] = roles
    return user

MRF_APPROVERS = {"pm", "gm", "director", "admin"}
MRF_EDITORS = {"site_engineer", "pm", "gm", "purchase", "director", "admin"}
MRF_CREATORS = {"site_engineer", "pm", "director", "admin"}
# Upload size limit shared across bulk-import endpoints
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5MB
GRN_ROLES = {"purchase", "admin"}  # Only Purchase creates GRN; Admin (Accounts) can also.
GRN_APPROVER_ROLES = {"admin", "director"}  # Accounts / Director approve GRN
DC_ROLES = {"purchase", "admin"}
DC_APPROVER_ROLES = {"admin", "director"}
CS_ROLES_UPLOAD_DEFERRED_ROLES = {"purchase", "admin"}
CS_APPROVER_ROLES = {"pm", "gm", "director"}

# --- Purchase approval thresholds (INR). Editable via /api/settings/thresholds ---
DEFAULT_THRESHOLD_GM = 50000.0        # PO value above this needs GM approval
DEFAULT_THRESHOLD_DIRECTOR = 500000.0  # PO value above this needs Director approval

async def get_thresholds() -> Dict[str, float]:
    doc = await db.settings.find_one({"_id": "thresholds"}, {"_id": 0})
    if not doc:
        return {"gm": DEFAULT_THRESHOLD_GM, "director": DEFAULT_THRESHOLD_DIRECTOR}
    return {
        "gm": float(doc.get("gm", DEFAULT_THRESHOLD_GM)),
        "director": float(doc.get("director", DEFAULT_THRESHOLD_DIRECTOR)),
    }

def _po_initial_status(total: float, gm_t: float, dir_t: float) -> str:
    if total > dir_t:
        return "pending_director_approval"
    if total > gm_t:
        return "pending_gm_approval"
    return "issued"
SYSTEMS = ["Fire Alarm", "Fire Fighting", "Gas Suppression", "Water Mist",
           "CCTV", "Access Control", "Structured Cabling", "Electrical", "Other"]
MRF_STATUS = ["draft", "under_review", "authorised", "rejected", "returned",
              "purchase_pending", "quotation_received", "po_pending",
              "po_issued", "partially_received", "fully_received",
              "closed", "cancelled",
              # legacy aliases retained for backwards compat with existing data:
              "submitted", "pm_review", "approved", "sent_to_purchase",
              "partially_ordered", "fully_ordered", "received"]

# Canonical status alias map (legacy -> new). New statuses map to themselves.
MRF_STATUS_ALIAS = {
    "submitted": "under_review",
    "pm_review": "under_review",
    "approved": "authorised",
    "sent_to_purchase": "purchase_pending",
    "partially_ordered": "po_issued",
    "fully_ordered": "po_issued",
    "received": "fully_received",
}
def canonical_mrf_status(s: str) -> str:
    return MRF_STATUS_ALIAS.get(s or "", s or "")
BILLING_STATUS = ["not_billed", "partially_billed", "fully_billed", "non_billable"]

# ---------------------- Models ----------------------
def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def gid(prefix: str = "id") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"

class UserOut(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "site_engineer"          # LEGACY (mirrors roles[0]) — kept for BC
    roles: List[str] = []                # NEW (Task 3A.1) — DB-driven source of truth
    is_active: bool = True

class RoleUpdate(BaseModel):
    user_id: str
    role: str

class RolesUpdate(BaseModel):
    """Multi-role update — Task 3A.1.
    `roles` is a canonical set (order not significant, deduped)."""
    user_id: str
    roles: List[str]

class SessionRequest(BaseModel):
    session_id: str

class Site(BaseModel):
    site_id: str = Field(default_factory=lambda: gid("st"))
    name: str
    location: Optional[str] = ""
    site_engineers: List[str] = []
    active: bool = True

class Project(BaseModel):
    project_id: str = Field(default_factory=lambda: gid("prj"))
    code: str
    name: str
    site: str
    client: Optional[str] = None                 # legacy free-text client name (kept for backwards compat)
    customer_id: Optional[str] = None            # FK to Customer.customer_id (permanent, alphanumeric)
    active: bool = True
    site_engineers: List[str] = []      # user_ids at project level
    project_managers: List[str] = []    # user_ids
    sites: List[Site] = []              # sub-sites under this project
    system_categories: List[str] = []   # e.g. ["Fire Alarm", "CCTV"]

class Vendor(BaseModel):
    vendor_id: str = Field(default_factory=lambda: gid("vnd"))
    name: str
    address: Optional[str] = ""
    gstin: Optional[str] = ""
    contact: Optional[str] = ""
    email: Optional[str] = ""
    active: bool = True

class MasterItem(BaseModel):
    item_id: str = Field(default_factory=lambda: gid("itm"))
    name: str
    category: str  # unit/system/brand/material/model/gst/department
    parent_id: Optional[str] = None   # e.g. Model.parent_id -> Brand.item_id
    value: Optional[str] = None       # generic extra (e.g. GST rate "18", HSN code, dept code)
    active: bool = True

# ---------------------- Customer & Customer PO ----------------------
CUSTOMER_ID_REGEX = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_-]{1,31}$")

class CustomerPO(BaseModel):
    """A customer-side PO issued to Vasu — anchors project billing/collection."""
    cpo_id: str = Field(default_factory=lambda: gid("cpo"))
    customer_id: Optional[str] = ""      # populated from URL path; client may omit
    po_number: str
    po_date: Optional[str] = None        # ISO date "YYYY-MM-DD"
    value: float = 0.0
    validity_till: Optional[str] = None  # ISO date
    attachment_name: Optional[str] = None
    attachment_b64: Optional[str] = None  # base64 (small files only)
    remarks: Optional[str] = ""
    active: bool = True

class Customer(BaseModel):
    customer_id: str                     # PERMANENT, admin-typed alphanumeric (e.g. "VASU-CUST-001")
    name: str
    gstin: Optional[str] = ""
    pan: Optional[str] = ""
    billing_address: Optional[str] = ""
    shipping_address: Optional[str] = ""
    contact_person: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    remarks: Optional[str] = ""
    active: bool = True
    created_at: datetime = Field(default_factory=now_utc)

class MRFItemIn(BaseModel):
    description: str
    specification: Optional[str] = ""
    part_number: Optional[str] = ""
    unit: str
    qty_requested: float
    qty_approved: Optional[float] = None
    purpose: Optional[str] = ""
    drawing_ref: Optional[str] = ""
    billable: bool = True
    boq_ref: Optional[str] = ""
    remarks: Optional[str] = ""
    # ---- Phase 3: master-driven selection (dropdown-only for site engineers) ----
    material_id: Optional[str] = None      # FK -> masters.item_id (category=material)
    make: Optional[str] = ""               # from masters.brand
    make_id: Optional[str] = None
    model: Optional[str] = ""              # from masters.model
    model_id: Optional[str] = None
    priority: Optional[str] = "normal"     # low | normal | high | urgent

    @field_validator("priority")
    @classmethod
    def _validate_priority(cls, v):
        if v is None or v == "":
            return "normal"
        s = str(v).lower().strip()
        if s not in ("low", "normal", "high", "urgent"):
            raise ValueError("priority must be one of low, normal, high, urgent")
        return s

class MRFItem(MRFItemIn):
    item_line_id: str = Field(default_factory=lambda: gid("mli"))
    status: str = "pending"  # pending/approved/rejected
    rejection_reason: Optional[str] = ""
    billing_status: str = "not_billed"
    qty_received: float = 0
    qty_issued: float = 0
    qty_billed: float = 0
    qty_ordered: float = 0
    client_bill_ref: Optional[str] = ""
    ra_bill_no: Optional[str] = ""
    billing_date: Optional[str] = ""
    billing_remarks: Optional[str] = ""

class MRFCreate(BaseModel):
    project_id: str
    site: str
    required_by: str
    requesting_person: str
    system_category: str
    items: List[MRFItemIn]
    attachments: List[str] = []  # base64 or descriptions
    remarks: Optional[str] = ""

class MRF(BaseModel):
    mrf_id: str = Field(default_factory=lambda: gid("mrf"))
    mrf_number: str
    date: datetime = Field(default_factory=now_utc)
    project_id: str
    site: str
    required_by: str
    requesting_person: str
    system_category: str
    customer_id: Optional[str] = None          # snapshot from project at create time
    customer_name: Optional[str] = None        # snapshot for display / exports
    items: List[MRFItem]
    attachments: List[str] = []
    remarks: str = ""
    status: str = "draft"
    created_by: str
    created_at: datetime = Field(default_factory=now_utc)
    updated_at: datetime = Field(default_factory=now_utc)
    pm_comments: str = ""
    deleted: bool = False

class POItemIn(BaseModel):
    mrf_id: str
    item_line_id: str
    description: str
    specification: Optional[str] = ""
    unit: str
    qty: float
    rate: float = 0
    discount: float = 0
    gst: float = 18
    remarks: Optional[str] = ""

class POCreate(BaseModel):
    vendor_id: str
    project_id: str
    delivery_site: str
    items: List[POItemIn]
    delivery_schedule: Optional[str] = ""
    payment_terms: Optional[str] = ""
    warranty_terms: Optional[str] = ""
    freight: float = 0
    other_charges: float = 0
    authorised_signatory: Optional[str] = ""
    vendor_quotation: Optional[str] = ""

class PO(POCreate):
    po_id: str = Field(default_factory=lambda: gid("po"))
    po_number: str
    date: datetime = Field(default_factory=now_utc)
    mrf_refs: List[str] = []
    status: str = "issued"  # issued/received/closed
    total: float = 0
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=now_utc)
    deleted: bool = False

class BillingUpdate(BaseModel):
    mrf_id: str
    item_line_id: str
    qty_received: Optional[float] = None
    qty_issued: Optional[float] = None
    qty_billed: Optional[float] = None
    client_bill_ref: Optional[str] = None
    ra_bill_no: Optional[str] = None
    billing_date: Optional[str] = None
    billing_remarks: Optional[str] = None
    billing_status: Optional[str] = None
    override_reason: Optional[str] = None

class ApprovalAction(BaseModel):
    action: str  # approve/reject/return
    comment: Optional[str] = ""
    item_actions: Optional[List[Dict[str, Any]]] = None  # [{item_line_id, action, qty_approved, reason}]

# ---------------------- Delivery Challan ----------------------
class DCItemIn(BaseModel):
    """One line on a Delivery Challan."""
    description: str
    unit: str = "nos"
    qty: float
    material_uid: Optional[str] = ""
    variant_uid: Optional[str] = ""
    make: Optional[str] = ""
    model: Optional[str] = ""
    remarks: Optional[str] = ""
    # optional back-links so we can trace DC back to source records
    mrf_id: Optional[str] = ""
    po_id: Optional[str] = ""
    grn_id: Optional[str] = ""
    item_line_id: Optional[str] = ""

class DCCreate(BaseModel):
    dc_type: str = "outbound"  # inbound | outbound
    po_id: Optional[str] = ""
    grn_id: Optional[str] = ""
    mrf_refs: List[str] = []
    project_id: str
    customer_id: Optional[str] = ""
    vendor_id: Optional[str] = ""
    from_location: str
    to_location: str
    dispatch_date: str
    vehicle_no: Optional[str] = ""
    driver_name: Optional[str] = ""
    driver_contact: Optional[str] = ""
    transporter: Optional[str] = ""
    e_way_bill_no: Optional[str] = ""
    e_way_bill_date: Optional[str] = ""
    items: List[DCItemIn]
    remarks: Optional[str] = ""
    authorised_signatory: Optional[str] = ""
    vendor_dc_ref: Optional[str] = ""  # for inbound DC — vendor's DC number

class DC(DCCreate):
    dc_id: str = Field(default_factory=lambda: gid("dc"))
    dc_number: str
    date: datetime = Field(default_factory=now_utc)
    status: str = "pending_approval"  # pending_approval | approved | rejected | cancelled | in_transit | delivered
    approval_history: List[Dict[str, Any]] = []
    customer_name: Optional[str] = ""
    vendor_name: Optional[str] = ""
    created_by: str
    created_at: datetime = Field(default_factory=now_utc)
    updated_at: Optional[datetime] = None
    deleted: bool = False

# ---------------------- Comparative Statement (import-only) ----------------------
class ComparativeVendorQuote(BaseModel):
    vendor_id: Optional[str] = ""
    vendor_name: str
    quote_number: Optional[str] = ""
    quote_date: Optional[str] = ""
    amount: float = 0.0
    gst_incl: bool = False
    validity: Optional[str] = ""
    remarks: Optional[str] = ""

class ComparativeStatementCreate(BaseModel):
    mrf_id: Optional[str] = ""
    po_id: Optional[str] = ""
    project_id: Optional[str] = ""
    customer_id: Optional[str] = ""
    title: str
    remarks: Optional[str] = ""
    prepared_by: Optional[str] = ""  # who created it outside the app
    prepared_on: Optional[str] = ""
    vendors: List[ComparativeVendorQuote] = []
    l1_vendor: Optional[str] = ""
    l1_amount: Optional[float] = 0
    selected_vendor: Optional[str] = ""
    selection_reason: Optional[str] = ""
    file_name: str
    mime_type: str
    file_base64: str  # base64-encoded attachment

class ComparativeStatement(ComparativeStatementCreate):
    cs_id: str = Field(default_factory=lambda: gid("cs"))
    cs_number: str
    uploaded_by: str
    uploaded_by_name: Optional[str] = ""
    uploaded_at: datetime = Field(default_factory=now_utc)
    file_size: int = 0
    status: str = "pending_approval"  # pending_approval | approved | rejected | superseded | cancelled
    approval_history: List[Dict[str, Any]] = []
    deleted: bool = False

# ---------------------- Auth ----------------------
async def get_current_user(authorization: Optional[str] = Header(None)) -> UserOut:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1]
    # ------------------------------------------------------------------
    # SEC-003: Short-lived, single-use download tokens (`dt_*`) issued via
    # POST /api/auth/download-token. They live in a separate collection so
    # they cannot be replayed after use, and expire in 5 minutes.
    # ------------------------------------------------------------------
    if token.startswith("dt_"):
        dt = await db.download_tokens.find_one({"token": token}, {"_id": 0})
        if not dt:
            raise HTTPException(status_code=401, detail="Invalid download token")
        exp = dt.get("expires_at")
        if exp and exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp and exp < now_utc():
            raise HTTPException(status_code=401, detail="Download token expired")
        if dt.get("used"):
            raise HTTPException(status_code=401, detail="Download token already used")
        # Mark used atomically to prevent replay (compareAndSet).
        r = await db.download_tokens.update_one(
            {"token": token, "used": {"$ne": True}},
            {"$set": {"used": True, "used_at": now_utc()}},
        )
        if r.matched_count == 0:
            raise HTTPException(status_code=401, detail="Download token already used")
        user = await db.users.find_one({"user_id": dt["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        if user.get("role") in LEGACY_ROLE_MAP:
            canon = LEGACY_ROLE_MAP[user["role"]]
            await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"role": canon}})
            user["role"] = canon
        user = _ensure_roles_shape(user)
        return UserOut(**user)

    sess = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not sess:
        raise HTTPException(status_code=401, detail="Invalid session")
    exp = sess["expires_at"]
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if exp < now_utc():
        raise HTTPException(status_code=401, detail="Session expired")
    user = await db.users.find_one({"user_id": sess["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    # Auto-migrate legacy roles on read
    if user.get("role") in LEGACY_ROLE_MAP:
        canon = LEGACY_ROLE_MAP[user["role"]]
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"role": canon}})
        user["role"] = canon
    user = _ensure_roles_shape(user)
    return UserOut(**user)

def require_roles(*allowed):
    async def dep(user: UserOut = None, authorization: Optional[str] = Header(None)):
        u = await get_current_user(authorization)
        if u.role not in allowed and u.role != "admin":
            raise HTTPException(status_code=403, detail=f"Role {u.role} not allowed")
        return u
    return dep

# ---------------------- Audit ----------------------
async def audit(entity: str, entity_id: str, action: str, user: UserOut, details: dict = None):
    await db.audit_logs.insert_one({
        "audit_id": gid("aud"),
        "entity": entity,
        "entity_id": entity_id,
        "action": action,
        "user_id": user.user_id,
        "user_name": user.name,
        "user_role": user.role,
        "details": details or {},
        "timestamp": now_utc(),
    })

def _strip_oids(o):
    """Recursively strip `_id` keys and convert ObjectId/datetime to JSON-safe primitives."""
    from bson import ObjectId
    if isinstance(o, ObjectId):
        return str(o)
    if isinstance(o, dict):
        return {k: _strip_oids(v) for k, v in o.items() if k != "_id"}
    if isinstance(o, list):
        return [_strip_oids(v) for v in o]
    if isinstance(o, datetime):
        return o.isoformat()
    return o

async def master_audit(entity: str, entity_id: str, action: str, user: UserOut,
                       reason: str, old_value: Any = None, new_value: Any = None):
    """Master-data change log — captures WHO, WHEN, WHY, and BEFORE/AFTER values."""
    old_value = _strip_oids(old_value)
    new_value = _strip_oids(new_value)
    ts = now_utc()
    await db.master_audit_logs.insert_one({
        "audit_id": gid("maud"),
        "entity": entity,
        "entity_id": entity_id,
        "action": action,
        "user_id": user.user_id,
        "user_name": user.name,
        "user_role": user.role,
        "reason": reason or "",
        "old_value": old_value,
        "new_value": new_value,
        "timestamp": ts,
    })
    # Mirror into the main audit log too so /api/audit sees it
    await db.audit_logs.insert_one({
        "audit_id": gid("aud"),
        "entity": entity,
        "entity_id": entity_id,
        "action": action,
        "user_id": user.user_id,
        "user_name": user.name,
        "user_role": user.role,
        "details": {"reason": reason, "old": old_value, "new": new_value},
        "timestamp": ts,
    })

async def notify(user_ids: List[str], title: str, body_text: str, link: str = ""):
    for uid in user_ids:
        await db.notifications.insert_one({
            "notification_id": gid("ntf"),
            "user_id": uid,
            "title": title,
            "body": body_text,
            "link": link,
            "read": False,
            "created_at": now_utc(),
        })

# ---------------------- Master Data ----------------------


# ---------------------- Customers ----------------------
def _validate_customer_id(cid: str) -> str:
    cid = (cid or "").strip()
    if not cid:
        raise HTTPException(400, "Customer ID is required")
    if not CUSTOMER_ID_REGEX.match(cid):
        raise HTTPException(400, "Customer ID must be alphanumeric (letters, digits, - or _, max 32 chars, must start with a letter or digit)")
    return cid


# ---------------------- Phase 4: Material Master (MAT-####) + Variants (VAR-####) ----------------------
MATERIAL_STATUSES = {"pending_pm_review", "approved", "rejected", "pending_gm_approval", "needs_correction"}
MAT_WORD_LIMIT = 100

async def _next_seq(name: str) -> int:
    """Atomic monotonic counter for MAT/VAR sequences."""
    doc = await db.counters.find_one_and_update(
        {"_id": name}, {"$inc": {"n": 1}}, upsert=True, return_document=True
    )
    return int(doc["n"]) if doc else 1

def _mat_uid(n: int) -> str: return f"MAT-{n:04d}"
def _var_uid(n: int) -> str: return f"VAR-{n:04d}"
def _norm(s: str) -> str: return re.sub(r"\s+", " ", (s or "").strip()).lower()
def _word_count(s: str) -> int: return len([w for w in re.split(r"\s+", (s or "").strip()) if w])

class MaterialIn(BaseModel):
    description: str
    category: str = ""      # system_category (FAS/FFT/GSS/CCTV etc.)
    unit: str = ""
    gst_rate: Optional[float] = None
    item_code: Optional[str] = ""   # user-editable short recall code
    remarks: Optional[str] = ""

async def _ensure_variant(material_uid: str, make: str, model: str,
                          material_id: Optional[str] = None) -> Dict[str, Any]:
    """Get or create a variant (idempotent). Prevents duplicate (mat, make, model)."""
    make = (make or "").strip()
    model = (model or "").strip()
    q = {"material_uid": material_uid, "make_norm": _norm(make), "model_norm": _norm(model)}
    existing = await db.variants.find_one(q, {"_id": 0})
    if existing:
        return existing
    n = await _next_seq("variants")
    doc = {
        "variant_uid": _var_uid(n),
        "material_uid": material_uid,
        "material_id": material_id,
        "make": make, "make_norm": _norm(make),
        "model": model, "model_norm": _norm(model),
        "created_at": now_utc(),
    }
    try:
        await db.variants.insert_one(doc)
    except DuplicateKeyError:
        existing = await db.variants.find_one(q, {"_id": 0})
        return existing or doc
    doc.pop("_id", None)
    return doc



# ---------------------- MRF ----------------------
async def next_mrf_number() -> str:
    year = datetime.now().year
    r = await db.counters.find_one_and_update(
        {"_id": f"mrf_{year}"},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True
    )
    seq = r["seq"] if r else 1
    return f"MRF/{year}/{seq:04d}"

async def next_po_number() -> str:
    year = datetime.now().year
    r = await db.counters.find_one_and_update(
        {"_id": f"po_{year}"},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True
    )
    seq = r["seq"] if r else 1
    return f"PO/{year}/{seq:04d}"

async def next_grn_number() -> str:
    year = datetime.now().year
    r = await db.counters.find_one_and_update(
        {"_id": f"grn_{year}"},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True
    )
    seq = r["seq"] if r else 1
    return f"GRN/{year}/{seq:04d}"

async def next_invoice_number() -> str:
    year = datetime.now().year
    r = await db.counters.find_one_and_update(
        {"_id": f"inv_{year}"},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True
    )
    seq = r["seq"] if r else 1
    return f"INV/{year}/{seq:04d}"

async def next_dc_number() -> str:
    year = datetime.now().year
    r = await db.counters.find_one_and_update(
        {"_id": f"dc_{year}"},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True,
    )
    seq = r["seq"] if r else 1
    return f"DC/{year}/{seq:04d}"

async def next_cs_number() -> str:
    year = datetime.now().year
    r = await db.counters.find_one_and_update(
        {"_id": f"cs_{year}"},
        {"$inc": {"seq": 1}},
        upsert=True, return_document=True,
    )
    seq = r["seq"] if r else 1
    return f"CS/{year}/{seq:04d}"


async def _check_mrf_access(u: UserOut, mrf: dict):
    """Raise 403 if user cannot see this MRF."""
    if u.role in ("admin", "purchase", "gm", "director"):
        return
    if u.role == "site_engineer":
        if mrf.get("created_by") == u.user_id:
            return
        raise HTTPException(403, "Not allowed")
    if u.role == "pm":
        p = await db.projects.find_one({"project_id": mrf["project_id"]}, {"_id": 0})
        if p and u.user_id in (p.get("project_managers") or []):
            return
        raise HTTPException(403, "Not allowed")
    if u.role == "store":
        p = await db.projects.find_one({"project_id": mrf["project_id"]}, {"_id": 0})
        if p and (u.user_id in (p.get("site_engineers") or []) or
                  any(u.user_id in (s.get("site_engineers") or [])
                      for s in (p.get("sites") or []) if s.get("active", True))):
            return
        raise HTTPException(403, "Not allowed")
    raise HTTPException(403, "Not allowed")

async def _check_po_access(u: UserOut, po: dict):
    if u.role in ("admin", "purchase", "gm", "director"):
        return
    if u.role == "pm":
        p = await db.projects.find_one({"project_id": po.get("project_id")}, {"_id": 0})
        if p and u.user_id in (p.get("project_managers") or []):
            return
    if u.role == "site_engineer":
        # Site engineer can see POs originating from MRFs they created
        for mid in po.get("mrf_refs") or []:
            m = await db.mrfs.find_one({"mrf_id": mid}, {"_id": 0, "created_by": 1})
            if m and m.get("created_by") == u.user_id:
                return
    if u.role == "store":
        p = await db.projects.find_one({"project_id": po.get("project_id")}, {"_id": 0})
        if p and (u.user_id in (p.get("site_engineers") or []) or
                  any(u.user_id in (s.get("site_engineers") or [])
                      for s in (p.get("sites") or []) if s.get("active", True))):
            return
    raise HTTPException(403, "Not allowed")

# ---------------------- Purchase Orders ----------------------

# ---------------------- Billing ----------------------


# ---------------------- PDF ----------------------
# ---------------------- Branded PDF helpers ----------------------
VASU_PRIMARY = colors.HexColor('#002FA7')  # Vasu brand blue
VASU_ACCENT = colors.HexColor('#F7941D')   # Vasu orange
VASU_TAGLINE = "Enterprise Fire Safety · CCTV · Access Control · Structured Cabling"
VASU_ADDR = "Vasu Infosec Pvt Ltd  ·  Pune · Delhi  ·  vasuinfosec.com"
# Detailed company block (env-overridable). These appear on all outbound documents.
VASU_LEGAL_NAME = os.getenv("VASU_LEGAL_NAME", "Vasu Infosec Pvt Ltd")
VASU_GSTIN = os.getenv("VASU_GSTIN", "27AAFCV0000A1Z0")
VASU_PAN = os.getenv("VASU_PAN", "AAFCV0000A")
VASU_STATE = os.getenv("VASU_STATE", "Maharashtra")
VASU_STATE_CODE = os.getenv("VASU_STATE_CODE", "27")
VASU_CIN = os.getenv("VASU_CIN", "")
VASU_ADDR_FULL = os.getenv(
    "VASU_ADDR_FULL",
    "Corporate Office: Pune, Maharashtra · Branch: Delhi · www.vasuinfosec.com",
)
VASU_CONTACT = os.getenv("VASU_CONTACT", "accounts@vasuinfosec.com  ·  +91-20-0000-0000")

# ---------------------- Export helpers (validation / audit / enrichment) ----------------------
async def _log_export(user: UserOut, entity: str, entity_id: str, fmt: str,
                      record_number: str = "", extras: Optional[dict] = None):
    """Audit every download/export so ops can prove regulatory compliance."""
    details = {"format": fmt, "record_number": record_number}
    if extras:
        details.update(extras)
    await audit(entity, entity_id, f"export_{fmt}", user, details)

async def _resolve_material_info(item: dict) -> dict:
    """Enrich a PO/MRF line with MAT/VAR UIDs, HSN, category, make/model.
    Returns a dict with keys: material_uid, variant_uid, hsn_code, make, model, category, gst_pct.
    Non-fatal — returns empty strings when unresolved.
    """
    out = {
        "material_uid": item.get("material_uid") or "",
        "variant_uid": item.get("variant_uid") or "",
        "hsn_code": item.get("hsn_code") or "",
        "make": item.get("make") or "",
        "model": item.get("model") or "",
        "category": item.get("category") or "",
        "gst_pct": float(item.get("gst") or 0),
    }
    mat = None
    if out["material_uid"]:
        mat = await db.materials.find_one({"material_uid": out["material_uid"]}, {"_id": 0})
    if not mat:
        desc = (item.get("description") or "").strip()
        if desc:
            mat = await db.materials.find_one({"description_norm": _norm(desc)}, {"_id": 0})
    if mat:
        out["material_uid"] = out["material_uid"] or mat.get("material_uid", "")
        out["hsn_code"] = out["hsn_code"] or (mat.get("hsn_code") or "")
        out["category"] = out["category"] or (mat.get("category") or "")
        if not out["gst_pct"] and mat.get("gst_rate") is not None:
            out["gst_pct"] = float(mat.get("gst_rate") or 0)
    # Resolve variant by (material_uid, make, model)
    if out["material_uid"] and (out["make"] or out["model"]) and not out["variant_uid"]:
        v = await db.variants.find_one(
            {"material_uid": out["material_uid"],
             "make_norm": _norm(out["make"]),
             "model_norm": _norm(out["model"])},
            {"_id": 0, "variant_uid": 1},
        )
        if v:
            out["variant_uid"] = v.get("variant_uid", "")
    return out


def _split_gst_amt(taxable: float, gst_pct: float, intra_state: bool) -> tuple:
    """Return (cgst, sgst, igst)."""
    gst_amt = round(taxable * gst_pct / 100, 2)
    if intra_state:
        half = round(gst_amt / 2, 2)
        return (half, gst_amt - half, 0.0)
    return (0.0, 0.0, gst_amt)


def _is_intra_state(party_state_or_addr: str) -> bool:
    s = (party_state_or_addr or "").upper()
    return ("MAHARASHTRA" in s) or (" MH" in f" {s}") or s.endswith(" MH") or s == "MH"


def _validate_po_for_export(po: dict, vendor: Optional[dict], customer: Optional[dict],
                            project: Optional[dict], fmt: str) -> tuple:
    """Return (missing_mandatory, warnings). Mandatory = blocks export unless ?force=1."""
    missing: List[str] = []
    warnings: List[str] = []

    # Vendor mandatory
    if not (vendor and vendor.get("name")):
        missing.append("Vendor name")
    if vendor and not (vendor.get("gstin") or "").strip():
        warnings.append("Vendor GSTIN")
    if vendor and not (vendor.get("address") or "").strip():
        warnings.append("Vendor address")

    # Customer mandatory (permanent Customer ID + name)
    if not (customer and customer.get("customer_id")):
        missing.append("Customer ID (permanent)")
    if not (customer and customer.get("name")):
        missing.append("Customer name")
    if customer and not (customer.get("gstin") or "").strip():
        warnings.append("Customer GSTIN")

    # Project mandatory
    if not project:
        missing.append("Project")
    else:
        if not project.get("code"):
            warnings.append("Project code")

    # PO metadata
    if not po.get("po_number"):
        missing.append("PO number")
    if not po.get("date"):
        missing.append("PO date")

    # Line items
    items = po.get("items") or []
    if not items:
        missing.append("At least one line item")
    for i, it in enumerate(items, 1):
        if not (it.get("description") or "").strip():
            missing.append(f"Item {i} description")
        if float(it.get("qty") or 0) <= 0:
            missing.append(f"Item {i} qty > 0")
        if float(it.get("rate") or 0) < 0:
            missing.append(f"Item {i} rate ≥ 0")
        if not (it.get("unit") or "").strip():
            warnings.append(f"Item {i} unit")
        # HSN mandatory ONLY for Tally (GST filing). Warning for PDF/Excel.
        if not (it.get("hsn_code") or "").strip():
            if fmt == "tally":
                missing.append(f"Item {i} HSN/SAC (required for Tally)")
            else:
                warnings.append(f"Item {i} HSN/SAC")

    # Vasu GSTIN — mandatory for tally, warning otherwise
    if not VASU_GSTIN:
        (missing if fmt == "tally" else warnings).append("Vasu GSTIN (env: VASU_GSTIN)")

    return missing, warnings


def _validate_mrf_for_export(mrf: dict, project: Optional[dict], customer: Optional[dict],
                             fmt: str) -> tuple:
    missing: List[str] = []
    warnings: List[str] = []
    if not (customer and customer.get("customer_id")):
        # MRFs may pre-date customer master — warn only.
        warnings.append("Customer ID (permanent)")
    if not project:
        missing.append("Project")
    if not mrf.get("mrf_number"):
        missing.append("MRF number")
    items = mrf.get("items") or []
    if not items:
        missing.append("At least one line item")
    for i, it in enumerate(items, 1):
        if not (it.get("description") or "").strip():
            missing.append(f"Item {i} description")
        if float(it.get("qty_requested") or 0) <= 0:
            missing.append(f"Item {i} qty_requested > 0")
    return missing, warnings


def _validation_error(missing: List[str], warnings: List[str], fmt: str) -> HTTPException:
    return HTTPException(
        status_code=422,
        detail={
            "error": "export_validation_failed",
            "format": fmt,
            "missing_mandatory": missing,
            "warnings": warnings,
            "hint": "Fix mandatory fields, or retry with ?force=1 to bypass (audit-logged).",
        },
    )


def _vasu_footer(canvas, doc):
    """Draws Vasu footer + page number on every page."""
    canvas.saveState()
    w, _ = A4
    canvas.setStrokeColor(VASU_PRIMARY)
    canvas.setLineWidth(0.6)
    canvas.line(15*mm, 12*mm, w - 15*mm, 12*mm)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.grey)
    canvas.drawString(15*mm, 7*mm, VASU_ADDR)
    canvas.drawRightString(w - 15*mm, 7*mm, f"Page {doc.page}")
    canvas.restoreState()

def _pdf_story(doc_type: str, doc_number: str, doc_date, project: dict,
               customer: Optional[dict], vendor: Optional[dict] = None,
               extra_ref: str = "") -> List:
    """Build a common branded header (VASU banner + Customer + Vendor + Project block)."""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('t', parent=styles['Title'], fontSize=20,
                                 textColor=VASU_PRIMARY, alignment=0, spaceAfter=2)
    tag_style = ParagraphStyle('tg', parent=styles['Normal'], fontSize=8,
                               textColor=colors.HexColor('#555'), spaceAfter=8)
    doc_title_style = ParagraphStyle('dt', parent=styles['Heading1'], fontSize=14,
                                     alignment=1, textColor=colors.HexColor('#111'),
                                     backColor=colors.HexColor('#EEF2FF'),
                                     borderPadding=6, spaceAfter=8)
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=9, leading=12)

    story: List = []
    # Header row: brand banner + Vasu identity
    story.append(Paragraph("VASU INFOSEC", title_style))
    story.append(Paragraph(VASU_TAGLINE, tag_style))

    # Vasu identity strip (legal name + GSTIN + PAN + state)
    identity_bits = [f"<b>{VASU_LEGAL_NAME}</b>"]
    if VASU_GSTIN: identity_bits.append(f"GSTIN: {VASU_GSTIN}")
    if VASU_PAN: identity_bits.append(f"PAN: {VASU_PAN}")
    if VASU_STATE: identity_bits.append(f"State: {VASU_STATE} ({VASU_STATE_CODE})")
    story.append(Paragraph("  ·  ".join(identity_bits),
                           ParagraphStyle('id', parent=styles['Normal'], fontSize=8,
                                          textColor=colors.HexColor('#333'), spaceAfter=2)))
    story.append(Paragraph(VASU_ADDR_FULL,
                           ParagraphStyle('id2', parent=styles['Normal'], fontSize=8,
                                          textColor=colors.HexColor('#555'), spaceAfter=6)))

    # Document title band
    story.append(Paragraph(doc_type, doc_title_style))

    # Meta table: doc no + date + refs
    date_str = doc_date.strftime('%d-%b-%Y') if hasattr(doc_date, 'strftime') else str(doc_date)[:10]
    meta_rows = [
        [Paragraph(f"<b>{doc_type.split()[-1]} No.</b>", small), Paragraph(doc_number, small),
         Paragraph("<b>Date</b>", small), Paragraph(date_str, small)],
    ]
    if extra_ref:
        meta_rows.append([Paragraph("<b>References</b>", small),
                          Paragraph(extra_ref, small), "", ""])
    meta_tbl = Table(meta_rows, colWidths=[70, 200, 40, 100])
    meta_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor('#D5D8DE')),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor('#F2F5FF')),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor('#F2F5FF')),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 6))

    # Customer + Project + Vendor combined info block
    cust_lines = []
    if customer:
        cust_lines.append(f"<b>Customer ID:</b> {customer.get('customer_id','')}")
        cust_lines.append(f"<b>Customer:</b> {customer.get('name','')}")
        if customer.get('gstin'): cust_lines.append(f"GSTIN: {customer['gstin']}")
        if customer.get('billing_address'): cust_lines.append(customer['billing_address'])
        if customer.get('contact_person'): cust_lines.append(f"Contact: {customer['contact_person']} · {customer.get('phone','')}")
    else:
        cust_lines.append(f"<b>Client:</b> {project.get('client','—')}")

    proj_lines = [
        f"<b>Project:</b> {project.get('name','')} ({project.get('code','')})",
        f"<b>Site:</b> {project.get('site','')}",
    ]
    if project.get('system_categories'):
        proj_lines.append(f"Systems: {', '.join(project['system_categories'])}")

    party_rows = [[Paragraph("BILL TO / CUSTOMER", small),
                   Paragraph("PROJECT DETAILS", small)],
                  [Paragraph("<br/>".join(cust_lines), small),
                   Paragraph("<br/>".join(proj_lines), small)]]
    party = Table(party_rows, colWidths=[210, 210])
    party.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), VASU_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor('#D5D8DE')),
        ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]))
    story.append(party)
    story.append(Spacer(1, 8))

    if vendor:
        vend_txt = (f"<b>Vendor:</b> {vendor.get('name','')} "
                    f"({vendor.get('gstin','')})<br/>"
                    f"{vendor.get('address','')} · Contact: {vendor.get('contact','')}")
        story.append(Paragraph(vend_txt, small))
        story.append(Spacer(1, 6))

    return story

def _safe_cell(v: Any) -> Any:
    """Neutralize spreadsheet formula-injection payloads."""
    if isinstance(v, str) and v and v[0] in ("=", "+", "-", "@"):
        return "'" + v
    return v

def _excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

async def _po_excel_workbook(pos_list: List[dict], single: bool = False):
    """Build a rich per-line PO Excel workbook. Reused by bulk & single exports."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Line Items"
    headers = [
        "PO#", "PO Date", "Status", "Customer ID", "Customer Name", "Customer GSTIN",
        "Project Code", "Project Name", "Site",
        "Vendor Name", "Vendor GSTIN", "Vendor Address",
        "MRF Refs", "Customer PO Ref", "Vendor Quote",
        "Line #", "MAT UID", "VAR UID", "Description", "Make", "Model",
        "HSN/SAC", "Unit", "Qty", "Rate", "Discount",
        "Taxable Value", "GST %", "CGST", "SGST", "IGST", "Line Total",
        "Freight (PO)", "Other Charges (PO)", "PO Grand Total",
        "Delivery Schedule", "Payment Terms", "Warranty",
        "Authorised Signatory", "Approved By (history)",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Preload lookups
    vend_map = {v["vendor_id"]: v for v in await db.vendors.find({}, {"_id": 0}).to_list(2000)}
    proj_map = {p["project_id"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(500)}
    cust_map = {}
    for c in await db.customers.find({}, {"_id": 0}).to_list(500):
        cust_map[c["customer_id"]] = c
    # Preload MRF numbers
    mrf_ids = set()
    for p in pos_list:
        for mid in p.get("mrf_refs") or []:
            mrf_ids.add(mid)
    mrf_num_map = {}
    if mrf_ids:
        for m in await db.mrfs.find({"mrf_id": {"$in": list(mrf_ids)}},
                                     {"_id": 0, "mrf_id": 1, "mrf_number": 1}).to_list(2000):
            mrf_num_map[m["mrf_id"]] = m.get("mrf_number", "")

    for p in pos_list:
        vend = vend_map.get(p.get("vendor_id")) or {}
        proj = proj_map.get(p.get("project_id")) or {}
        cust = cust_map.get(p.get("customer_id")) or {}
        intra = _is_intra_state(vend.get("address") or VASU_STATE)
        pdate = p["date"].strftime("%Y-%m-%d") if isinstance(p.get("date"), datetime) else str(p.get("date", ""))[:10]
        mrf_refs = ", ".join(sorted({mrf_num_map.get(x, x) for x in (p.get("mrf_refs") or [])}))
        cpo_ref = ""
        if cust.get("customer_id"):
            cpos = cust.get("pos") or []
            active = sorted([x for x in cpos if x.get("active", True)],
                            key=lambda x: x.get("po_date") or "", reverse=True)
            if active:
                cpo_ref = f"{active[0].get('po_number','')} dt {active[0].get('po_date','')}"
        signers = "; ".join([
            f"{h.get('user_name','')}({(h.get('user_role') or '').upper()})"
            for h in (p.get("approval_history") or [])
            if (h.get("action") or "").lower() == "approve"
        ]) or ""
        for idx, it in enumerate(p.get("items") or [], 1):
            info = await _resolve_material_info(it)
            qty = float(it.get("qty") or 0)
            rate = float(it.get("rate") or 0)
            disc = float(it.get("discount") or 0)
            gst_pct = float(it.get("gst") or 0) or info.get("gst_pct") or 0
            taxable = round(qty * rate - disc, 2)
            cgst, sgst, igst = _split_gst_amt(taxable, gst_pct, intra)
            line_total = round(taxable + cgst + sgst + igst, 2)
            ws.append([
                _safe_cell(p.get("po_number", "")), _safe_cell(pdate),
                _safe_cell(p.get("status", "")),
                _safe_cell(cust.get("customer_id") or p.get("customer_id") or ""),
                _safe_cell(cust.get("name") or p.get("customer_name") or ""),
                _safe_cell(cust.get("gstin") or ""),
                _safe_cell(proj.get("code") or ""), _safe_cell(proj.get("name") or ""),
                _safe_cell(p.get("delivery_site") or proj.get("site") or ""),
                _safe_cell(vend.get("name") or ""),
                _safe_cell(vend.get("gstin") or ""),
                _safe_cell(vend.get("address") or ""),
                _safe_cell(mrf_refs), _safe_cell(cpo_ref),
                _safe_cell(p.get("vendor_quotation") or ""),
                idx,
                _safe_cell(info.get("material_uid") or ""),
                _safe_cell(info.get("variant_uid") or ""),
                _safe_cell(it.get("description") or ""),
                _safe_cell(info.get("make") or ""), _safe_cell(info.get("model") or ""),
                _safe_cell(info.get("hsn_code") or ""),
                _safe_cell(it.get("unit") or ""),
                qty, rate, disc,
                taxable, gst_pct, cgst, sgst, igst, line_total,
                float(p.get("freight") or 0),
                float(p.get("other_charges") or 0),
                float(p.get("total") or 0),
                _safe_cell(p.get("delivery_schedule") or ""),
                _safe_cell(p.get("payment_terms") or ""),
                _safe_cell(p.get("warranty_terms") or ""),
                _safe_cell(p.get("authorised_signatory") or ""),
                _safe_cell(signers),
            ])

    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(len(h) + 2, 30))
    return wb


async def _po_tally_workbook(pos_list: List[dict]):
    """Build a Tally-compatible Purchase-voucher workbook for one or many POs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Voucher"
    headers = [
        "Voucher Date", "Voucher Type", "Voucher No", "Reference No", "Reference Date",
        "Company GSTIN", "Company State", "Company State Code",
        "Party Ledger", "Party GSTIN", "Party State",
        "Customer ID", "Customer Name", "Customer GSTIN",
        "Project Code",
        "MAT UID", "VAR UID", "Item Name", "Make", "Model",
        "HSN/SAC", "Unit", "Quantity", "Rate", "Discount",
        "Taxable Value", "GST Rate %", "CGST", "SGST", "IGST",
        "Line Total", "Narration",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    vend_map = {v["vendor_id"]: v for v in await db.vendors.find({}, {"_id": 0}).to_list(2000)}
    proj_map = {p["project_id"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(500)}
    cust_map = {c["customer_id"]: c for c in await db.customers.find({}, {"_id": 0}).to_list(500)}

    for p in pos_list:
        vend = vend_map.get(p.get("vendor_id")) or {}
        proj = proj_map.get(p.get("project_id")) or {}
        cust = cust_map.get(p.get("customer_id")) or {}
        intra = _is_intra_state(vend.get("address") or VASU_STATE)
        vdate = p["date"].strftime("%d-%m-%Y") if isinstance(p.get("date"), datetime) else str(p.get("date"))[:10]
        for it in p.get("items", []):
            info = await _resolve_material_info(it)
            qty = float(it.get("qty") or 0)
            rate = float(it.get("rate") or 0)
            disc = float(it.get("discount") or 0)
            gst_pct = float(it.get("gst") or 0) or info.get("gst_pct") or 0
            taxable = round(qty * rate - disc, 2)
            cgst, sgst, igst = _split_gst_amt(taxable, gst_pct, intra)
            line_total = round(taxable + cgst + sgst + igst, 2)
            narration = f"PO {p.get('po_number','')} · Project {proj.get('code','')} · Customer {cust.get('customer_id','')}"
            ws.append([
                _safe_cell(vdate), _safe_cell("Purchase"), _safe_cell(p.get("po_number", "")),
                "", "",
                _safe_cell(VASU_GSTIN), _safe_cell(VASU_STATE), _safe_cell(VASU_STATE_CODE),
                _safe_cell(vend.get("name", "")),
                _safe_cell(vend.get("gstin", "")),
                _safe_cell((vend.get("address", "") or "")[:40]),
                _safe_cell(cust.get("customer_id", "")),
                _safe_cell(cust.get("name", "")),
                _safe_cell(cust.get("gstin", "")),
                _safe_cell(proj.get("code", "")),
                _safe_cell(info.get("material_uid") or ""),
                _safe_cell(info.get("variant_uid") or ""),
                _safe_cell(it.get("description", "")),
                _safe_cell(info.get("make") or ""), _safe_cell(info.get("model") or ""),
                _safe_cell(info.get("hsn_code") or ""),
                _safe_cell(it.get("unit", "")),
                qty, rate, disc,
                taxable, gst_pct, cgst, sgst, igst,
                line_total, _safe_cell(narration),
            ])

    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(len(h) + 2, 28))
    return wb


async def _mrf_pdf_bytes(mrf: dict, project: dict, customer: Optional[dict]) -> bytes:
    """Branded MRF PDF (Material Requisition Form)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    mdate = mrf.get("date")
    extra = f"Requested by: {mrf.get('requesting_person','')}  ·  System: {mrf.get('system_category','')}"
    if mrf.get("required_by"):
        extra += f"  ·  Required by: {mrf['required_by']}"
    story = _pdf_story("MATERIAL REQUISITION FORM", mrf.get("mrf_number", ""),
                       mdate, project, customer, extra_ref=extra)
    styles = getSampleStyleSheet()
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, leading=10)
    smallb = ParagraphStyle('sb', parent=styles['Normal'], fontSize=8, leading=10,
                            fontName="Helvetica-Bold")

    header = ["#", "MAT / VAR", "Description", "Make / Model", "Unit",
              "Qty Req.", "Qty Appr.", "Priority", "Status", "Purpose"]
    data = [[Paragraph(h, smallb) for h in header]]
    for idx, it in enumerate(mrf.get("items") or [], 1):
        info = await _resolve_material_info(it)
        mv = ((info.get("material_uid") or "-") + "\n" + (info.get("variant_uid") or "-"))
        mkmd = ((info.get("make") or it.get("make") or "-") + "\n" +
                (info.get("model") or it.get("model") or "-"))
        data.append([
            Paragraph(str(idx), small),
            Paragraph(mv, small),
            Paragraph((it.get("description") or "")[:70], small),
            Paragraph(mkmd, small),
            Paragraph(it.get("unit") or "", small),
            Paragraph(f"{float(it.get('qty_requested') or 0):g}", small),
            Paragraph(f"{float(it.get('qty_approved') or 0):g}", small),
            Paragraph((it.get("priority") or "normal").upper(), small),
            Paragraph((it.get("status") or "pending").upper(), small),
            Paragraph((it.get("purpose") or "-")[:40], small),
        ])
    tbl = Table(data, colWidths=[16, 52, 130, 60, 26, 38, 38, 36, 44, 76], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), VASU_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor('#B7BEC9')),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F8FC')]),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 8))
    if mrf.get("remarks"):
        story.append(Paragraph(f"<b>Remarks:</b> {mrf['remarks']}", small))
    story.append(Paragraph(f"<b>Status:</b> {(mrf.get('status') or 'draft').upper()}", small))
    story.append(Spacer(1, 6))
    # Approval history
    hist = mrf.get("approval_history") or []
    if hist:
        story.append(Paragraph("<b>Approval Trail</b>", smallb))
        for h in hist:
            ts = (h.get("timestamp") or "")[:16].replace("T", " ")
            story.append(Paragraph(
                f"• {h.get('user_name','')} ({(h.get('user_role') or '').upper()}) — "
                f"{(h.get('action') or '').upper()} at {ts}"
                + (f" — “{h['comment']}”" if h.get('comment') else ""), small))
    doc.build(story, onFirstPage=_vasu_footer, onLaterPages=_vasu_footer)
    buf.seek(0)
    return buf.read()





# ---------------------- Unified single-record export dispatchers ----------------------
_ALLOWED_PO_FORMATS = ("pdf", "excel", "tally")
_ALLOWED_MRF_FORMATS = ("pdf", "excel")






# ---------------------- GRN ----------------------


# ---------------------- Delivery Challan ----------------------
def _dc_pdf_story(dc: dict, project: dict, customer: Optional[dict],
                  vendor: Optional[dict]) -> list:
    styles = getSampleStyleSheet()
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, leading=10)
    smallb = ParagraphStyle('sb', parent=styles['Normal'], fontSize=8, leading=10,
                            fontName="Helvetica-Bold")
    ttl = "DELIVERY CHALLAN"
    if (dc.get("dc_type") or "").lower() == "inbound":
        ttl = "DELIVERY CHALLAN (Inbound)"
    story = _pdf_story(ttl, dc.get("dc_number", ""), dc.get("date"),
                       project, customer, vendor,
                       extra_ref=f"Type: {(dc.get('dc_type') or '').upper()}"
                                 + (f"  ·  PO {dc.get('po_id','')}" if dc.get("po_id") else ""))
    # Logistics block
    logistics = [
        f"<b>From:</b> {dc.get('from_location','—')}",
        f"<b>To:</b> {dc.get('to_location','—')}",
        f"<b>Dispatch Date:</b> {dc.get('dispatch_date','—')}",
        f"<b>Vehicle:</b> {dc.get('vehicle_no','—')}",
        f"<b>Driver:</b> {dc.get('driver_name','—')} ({dc.get('driver_contact','')})",
        f"<b>Transporter:</b> {dc.get('transporter','—')}",
    ]
    if dc.get("e_way_bill_no"):
        logistics.append(f"<b>E-Way Bill:</b> {dc['e_way_bill_no']}"
                          + (f" dt {dc['e_way_bill_date']}" if dc.get("e_way_bill_date") else ""))
    if dc.get("vendor_dc_ref"):
        logistics.append(f"<b>Vendor DC Ref:</b> {dc['vendor_dc_ref']}")
    for l in logistics:
        story.append(Paragraph(l, small))
    story.append(Spacer(1, 6))

    header = ["#", "MAT / VAR", "Description", "Make / Model", "Unit", "Qty", "Remarks"]
    data = [[Paragraph(h, smallb) for h in header]]
    for idx, it in enumerate(dc.get("items") or [], 1):
        mv = ((it.get("material_uid") or "-") + "\n" + (it.get("variant_uid") or "-"))
        mm2 = ((it.get("make") or "-") + "\n" + (it.get("model") or "-"))
        data.append([
            Paragraph(str(idx), small),
            Paragraph(mv, small),
            Paragraph((it.get("description") or "")[:80], small),
            Paragraph(mm2, small),
            Paragraph(it.get("unit") or "", small),
            Paragraph(f"{float(it.get('qty') or 0):g}", small),
            Paragraph((it.get("remarks") or "")[:60], small),
        ])
    tbl = Table(data, colWidths=[16, 60, 170, 74, 30, 44, 110], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), VASU_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor('#B7BEC9')),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F8FC')]),
    ]))
    story.append(tbl)
    if dc.get("remarks"):
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<b>Remarks:</b> {dc['remarks']}", small))
    # Declaration
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "<i>This is a Delivery Challan, not a tax invoice. Material dispatched for "
        "project execution / job-work / return. Goods received in good condition unless "
        "recorded otherwise.</i>", small))
    story += [
        Spacer(1, 24),
        Paragraph(f"<b>Authorised Signatory</b> — {dc.get('authorised_signatory') or 'Vasu Infosec Pvt Ltd'}", small),
        Spacer(1, 20),
        Paragraph("Prepared By: __________________  &nbsp;&nbsp;&nbsp;  Driver Signature: __________________  &nbsp;&nbsp;&nbsp;  Receiver Signature: __________________", small),
    ]
    return story


async def _dc_excel_workbook(dcs: List[dict]) -> "openpyxl.Workbook":
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery Challan"
    headers = [
        "DC#", "Date", "Type", "Status", "PO#", "GRN#", "Customer ID", "Customer Name",
        "Vendor Name", "Project Code", "From", "To", "Dispatch Date",
        "Vehicle", "Driver", "Driver Contact", "Transporter", "E-Way Bill", "E-Way Date",
        "Vendor DC Ref",
        "Line#", "MAT UID", "VAR UID", "Description", "Make", "Model", "Unit", "Qty",
        "Item Remarks", "DC Remarks", "Authorised Signatory",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    proj_map = {p["project_id"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(500)}
    for dc in dcs:
        proj = proj_map.get(dc.get("project_id")) or {}
        ddate = dc["date"].strftime("%Y-%m-%d") if isinstance(dc.get("date"), datetime) else str(dc.get("date", ""))[:10]
        po = None
        if dc.get("po_id"):
            po = await db.pos.find_one({"po_id": dc["po_id"]}, {"_id": 0, "po_number": 1})
        grn = None
        if dc.get("grn_id"):
            grn = await db.grns.find_one({"grn_id": dc["grn_id"]}, {"_id": 0, "grn_number": 1})
        for idx, it in enumerate(dc.get("items") or [], 1):
            ws.append([
                _safe_cell(dc.get("dc_number", "")), _safe_cell(ddate),
                _safe_cell(dc.get("dc_type", "")),
                _safe_cell(dc.get("status", "")),
                _safe_cell((po or {}).get("po_number", "")),
                _safe_cell((grn or {}).get("grn_number", "")),
                _safe_cell(dc.get("customer_id", "")),
                _safe_cell(dc.get("customer_name", "")),
                _safe_cell(dc.get("vendor_name", "")),
                _safe_cell(proj.get("code", "")),
                _safe_cell(dc.get("from_location", "")),
                _safe_cell(dc.get("to_location", "")),
                _safe_cell(dc.get("dispatch_date", "")),
                _safe_cell(dc.get("vehicle_no", "")),
                _safe_cell(dc.get("driver_name", "")),
                _safe_cell(dc.get("driver_contact", "")),
                _safe_cell(dc.get("transporter", "")),
                _safe_cell(dc.get("e_way_bill_no", "")),
                _safe_cell(dc.get("e_way_bill_date", "")),
                _safe_cell(dc.get("vendor_dc_ref", "")),
                idx,
                _safe_cell(it.get("material_uid", "")),
                _safe_cell(it.get("variant_uid", "")),
                _safe_cell(it.get("description", "")),
                _safe_cell(it.get("make", "")),
                _safe_cell(it.get("model", "")),
                _safe_cell(it.get("unit", "")),
                float(it.get("qty") or 0),
                _safe_cell(it.get("remarks", "")),
                _safe_cell(dc.get("remarks", "")),
                _safe_cell(dc.get("authorised_signatory", "")),
            ])
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(len(h) + 2, 28))
    return wb


def _validate_dc_for_export(dc: dict, project: Optional[dict], fmt: str) -> tuple:
    missing: List[str] = []
    warnings: List[str] = []
    if not dc.get("dc_number"):
        missing.append("DC number")
    if not project:
        missing.append("Project")
    if not (dc.get("from_location") or "").strip():
        missing.append("From location")
    if not (dc.get("to_location") or "").strip():
        missing.append("To location")
    if not (dc.get("dispatch_date") or "").strip():
        missing.append("Dispatch date")
    if not (dc.get("items") or []):
        missing.append("At least one line item")
    if not (dc.get("vehicle_no") or "").strip():
        warnings.append("Vehicle number")
    if not (dc.get("driver_name") or "").strip():
        warnings.append("Driver name")
    for i, it in enumerate(dc.get("items") or [], 1):
        if not (it.get("description") or "").strip():
            missing.append(f"Item {i} description")
        if float(it.get("qty") or 0) <= 0:
            missing.append(f"Item {i} qty > 0")
    return missing, warnings



# ---------------------- Comparative Statement (import-only) ----------------------


# ---------------------- LLM Co-pilot ----------------------
# Extracted to routers/llm.py in Phase 9C. Import is deferred to the bottom
# of this file (below), after every symbol it depends on is defined.



# ---------------------- Invoice ----------------------
class InvoiceItemIn(BaseModel):
    mrf_id: Optional[str] = ""
    item_line_id: Optional[str] = ""
    description: str
    unit: str = "Nos"
    qty: float
    rate: float
    discount: float = 0
    gst: float = 18

class InvoiceCreate(BaseModel):
    po_id: str
    vendor_invoice_number: Optional[str] = ""
    invoice_date: Optional[str] = ""
    items: List[InvoiceItemIn]
    freight: float = 0
    other_charges: float = 0
    remarks: Optional[str] = ""



# ---------------------- Seed ----------------------
@api.post("/seed")
async def seed_data(authorization: Optional[str] = Header(None)):
    """Idempotent seed of demo data. Requires admin, or dev-mode when no users exist yet."""
    user_count = await db.users.count_documents({})
    if user_count > 0:
        # Bootstrap complete — only admin may reseed
        u = await get_current_user(authorization)
        if u.role != "admin":
            raise HTTPException(403, "Only admin can seed once users exist")
    # (Otherwise: no users yet, allow unauthenticated bootstrap.)
    # Users — legacy demo (dev/QA) + real Vasu Infosec roster
    demo_users = [
        # Legacy demo users (kept for dev-login & smoke tests)
        {"email": "director@vasu.dev", "name": "Vishal (Director demo)", "role": "director"},
        {"email": "pm@vasu.dev", "name": "Priya (PM demo)", "role": "pm"},
        {"email": "gm@vasu.dev", "name": "Girish (GM demo)", "role": "gm"},
        {"email": "purchase@vasu.dev", "name": "Kumar (Purchase demo)", "role": "purchase"},
        {"email": "admin@vasu.dev", "name": "Master Admin", "role": "admin"},
        {"email": "siteeng@vasu.dev", "name": "Sanjay (Site Engineer demo)", "role": "site_engineer"},
        # Real Vasu Infosec roster (Google-sign-in enabled). Roles pre-assigned.
        {"email": "vivek@vasuinfosec.com", "name": "Vivek (Director)", "role": "director"},
        {"email": "balkrishna@vasuinfosec.com", "name": "Balkrishna (GM)", "role": "gm"},
        {"email": "saket.iyer@vasuinfosec.com", "name": "Saket Iyer (PM · Pune)", "role": "pm"},
        {"email": "himanshu@vasuinfosec.com", "name": "Himanshu (PM · Delhi)", "role": "pm"},
        {"email": "wasim@vasuinfosec.com", "name": "Wasim (Purchase · Pune)", "role": "purchase"},
        {"email": "sanjeev@vasuinfosec.com", "name": "Sanjeev (Purchase · Delhi)", "role": "purchase"},
        {"email": "pundlik@vasuinfosec.com", "name": "Pundlik (Admin)", "role": "admin"},
        {"email": "chand@vasuinfosec.com", "name": "Chand (Site Engineer)", "role": "site_engineer"},
        {"email": "saurabh@vasuinfosec.com", "name": "Saurabh (Site Engineer)", "role": "site_engineer"},
        {"email": "abhishek.yadav@vasuinfosec.com", "name": "Abhishek Yadav (Site Engineer)", "role": "site_engineer"},
        {"email": "shivsaran@vasuinfosec.com", "name": "Shiv Saran (Site Engineer)", "role": "site_engineer"},
        {"email": "abhishek@vasuinfosec.com", "name": "Abhishek (Stores)", "role": "store"},
    ]
    for u in demo_users:
        existing = await db.users.find_one({"email": u["email"]}, {"_id": 0})
        if not existing:
            await db.users.insert_one({
                "user_id": f"user_{uuid.uuid4().hex[:12]}",
                "email": u["email"], "name": u["name"], "picture": "",
                "role": u["role"], "is_active": True, "created_at": now_utc(),
            })
        else:
            # Only backfill display name; do NOT overwrite an already-set role
            # (admins may have adjusted roles in the UI).
            await db.users.update_one(
                {"email": u["email"]},
                {"$set": {"name": u["name"]},
                 "$setOnInsert": {"role": u["role"]}}
            )
            # If existing role is legacy or missing, canonicalize
            cur = existing.get("role")
            if not cur or cur in LEGACY_ROLE_MAP:
                await db.users.update_one({"email": u["email"]},
                                          {"$set": {"role": u["role"]}})

    # Projects
    projects = [
        {"code": "VIS-101", "name": "ABC IT Park Fire Safety", "site": "Bangalore Whitefield", "client": "ABC Ltd"},
        {"code": "VIS-102", "name": "XYZ Data Center CCTV", "site": "Hyderabad Gachibowli", "client": "XYZ Corp"},
        {"code": "VIS-103", "name": "PQR Mall Access Control", "site": "Mumbai Andheri", "client": "PQR Group"},
    ]
    for p in projects:
        existing = await db.projects.find_one({"code": p["code"]}, {"_id": 0})
        if not existing:
            await db.projects.insert_one({"project_id": gid("prj"), **p, "active": True,
                                          "site_engineers": [], "project_managers": []})

    # Seed team assignments (idempotent)
    # - site_engineer demo user -> site_engineers on all seed projects
    # - pm demo + real PMs (saket, himanshu) -> project_managers on all seed projects
    se_user = await db.users.find_one({"email": "siteeng@vasu.dev"}, {"_id": 0, "user_id": 1})
    pm_user = await db.users.find_one({"email": "pm@vasu.dev"}, {"_id": 0, "user_id": 1})
    saket = await db.users.find_one({"email": "saket.iyer@vasuinfosec.com"}, {"_id": 0, "user_id": 1})
    himanshu = await db.users.find_one({"email": "himanshu@vasuinfosec.com"}, {"_id": 0, "user_id": 1})
    se_ids = [x["user_id"] for x in (se_user, pm_user) if x]  # legacy pm@ kept as SE for backwards compat
    pm_ids = [x["user_id"] for x in (pm_user, saket, himanshu) if x]
    if se_ids or pm_ids:
        add_ops: Dict[str, Any] = {}
        if se_ids:
            add_ops["site_engineers"] = {"$each": se_ids}
        if pm_ids:
            add_ops["project_managers"] = {"$each": pm_ids}
        await db.projects.update_many(
            {"code": {"$in": [p["code"] for p in projects]}},
            {"$addToSet": add_ops},
        )

    # Vendors
    vendors = [
        {"name": "Honeywell India Pvt Ltd", "gstin": "29AAACH1234A1Z5", "contact": "9880012345", "address": "Bangalore"},
        {"name": "Siemens Ltd", "gstin": "27AAACS4567B1Z6", "contact": "9820098765", "address": "Mumbai"},
        {"name": "Bosch Security Systems", "gstin": "33AAACB7890C1Z7", "contact": "9840001122", "address": "Chennai"},
        {"name": "Ravel Electronics", "gstin": "33AABCR2345D1Z8", "contact": "9840066001", "address": "Chennai"},
        {"name": "Anixter India Cabling", "gstin": "06AAACA9988E1Z9", "contact": "9910044556", "address": "Gurgaon"},
    ]
    for v in vendors:
        existing = await db.vendors.find_one({"name": v["name"]}, {"_id": 0})
        if not existing:
            await db.vendors.insert_one({"vendor_id": gid("vnd"), **v, "email": "", "active": True})

    # Masters (units, brands)
    masters = [
        {"name": "Nos", "category": "unit"}, {"name": "Meter", "category": "unit"},
        {"name": "Kg", "category": "unit"}, {"name": "Set", "category": "unit"},
        {"name": "Roll", "category": "unit"},
        {"name": "Honeywell", "category": "brand"}, {"name": "Siemens", "category": "brand"},
        {"name": "Bosch", "category": "brand"}, {"name": "Hikvision", "category": "brand"},
        {"name": "CP Plus", "category": "brand"},
    ]
    for m in masters:
        existing = await db.masters.find_one({"name": m["name"], "category": m["category"]}, {"_id": 0})
        if not existing:
            await db.masters.insert_one({"item_id": gid("itm"), **m, "active": True})

    # Materials
    materials = [
        "Smoke Detector Photoelectric", "Heat Detector Fixed 57°C", "Manual Call Point",
        "Sounder Strobe Wall Mount", "Fire Extinguisher CO2 4.5kg",
        "IP CCTV Camera 4MP Bullet", "Access Control Reader RFID",
        "Cat6 UTP Cable 305m Box", "Fire Alarm Control Panel 4-Zone",
        "FM200 Gas Cylinder 40L",
    ]
    for m in materials:
        existing = await db.masters.find_one({"name": m, "category": "material"}, {"_id": 0})
        if not existing:
            await db.masters.insert_one({"item_id": gid("itm"), "name": m, "category": "material", "active": True})

    # GST rates (standard Indian slabs)
    gst_rates = [("Nil", "0"), ("5%", "5"), ("12%", "12"), ("18%", "18"), ("28%", "28")]
    for name, val in gst_rates:
        existing = await db.masters.find_one({"name": name, "category": "gst"}, {"_id": 0})
        if not existing:
            await db.masters.insert_one({"item_id": gid("itm"), "name": name, "category": "gst",
                                         "value": val, "active": True})

    # Departments
    departments = ["Fire Safety", "IT & Networking", "Operations", "Projects", "Purchase", "Accounts"]
    for d in departments:
        existing = await db.masters.find_one({"name": d, "category": "department"}, {"_id": 0})
        if not existing:
            await db.masters.insert_one({"item_id": gid("itm"), "name": d, "category": "department",
                                         "active": True})

    # Phase 4: Migrate existing masters.category="material" records into the new
    # `materials` collection with MAT-#### UIDs, in creation-date order. Idempotent.
    legacy_mats = await db.masters.find(
        {"category": "material", "active": True}, {"_id": 0}
    ).sort("_id", 1).to_list(2000)
    for lm in legacy_mats:
        norm = _norm(lm.get("name", ""))
        if not norm:
            continue
        existing = await db.materials.find_one({"description_norm": norm}, {"_id": 0})
        if existing:
            continue
        n = await _next_seq("materials")
        wc = _word_count(lm.get("name", ""))
        doc = {
            "material_uid": _mat_uid(n),
            "description": lm["name"],
            "description_norm": norm,
            "category": "", "unit": "", "gst_rate": None,
            "item_code": "", "remarks": "Migrated from legacy master",
            "status": "approved" if wc <= MAT_WORD_LIMIT else "needs_correction",
            "created_by": "system", "created_by_name": "System Migration",
            "created_at": now_utc(), "updated_at": now_utc(),
            "approved_by": "system" if wc <= MAT_WORD_LIMIT else None,
            "approved_at": now_utc() if wc <= MAT_WORD_LIMIT else None,
            "legacy_item_id": lm.get("item_id"),
            "source": "migration",
        }
        await db.materials.insert_one(doc)

    # Migrate legacy project.client strings -> Customer records (permanent Customer IDs).
    # Any project without a customer_id but with a non-empty client string gets a matching Customer.
    legacy_projects = await db.projects.find(
        {"$and": [{"client": {"$nin": [None, ""]}},
                  {"$or": [{"customer_id": None}, {"customer_id": {"$exists": False}}]}]},
        {"_id": 0, "project_id": 1, "client": 1}
    ).to_list(500)
    for lp in legacy_projects:
        client_name = (lp.get("client") or "").strip()
        if not client_name:
            continue
        # See if a Customer already exists with that name (case-insensitive)
        cust = await db.customers.find_one(
            {"name": {"$regex": f"^{re.escape(client_name)}$", "$options": "i"}}, {"_id": 0}
        )
        if not cust:
            # Generate a safe alphanumeric customer_id from the name (fallback if collision)
            base = re.sub(r"[^A-Za-z0-9]+", "", client_name).upper()[:10] or "CUST"
            candidate = f"CUST-{base}"
            suffix = 1
            while await db.customers.find_one({"customer_id": candidate}, {"_id": 0}):
                candidate = f"CUST-{base}-{suffix:02d}"; suffix += 1
            cust = {
                "customer_id": candidate,
                "name": client_name, "gstin": "", "pan": "",
                "billing_address": "", "shipping_address": "",
                "contact_person": "", "phone": "", "email": "",
                "remarks": "Auto-created from legacy project.client string",
                "active": True, "created_at": now_utc(),
            }
            await db.customers.insert_one(cust)
        # Link project -> customer
        await db.projects.update_one(
            {"project_id": lp["project_id"]},
            {"$set": {"customer_id": cust["customer_id"]}},
        )

    return {"ok": True, "message": "Seed complete", "logins": [u["email"] for u in demo_users]}

# ---------------------- Startup ----------------------
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("user_id", unique=True)
    await db.user_sessions.create_index("session_token", unique=True)
    await db.user_sessions.create_index("expires_at", expireAfterSeconds=0)
    # SEC-003: Short-lived, single-use download tokens (5-minute TTL).
    await db.download_tokens.create_index("token", unique=True)
    await db.download_tokens.create_index("expires_at", expireAfterSeconds=0)
    await db.mrfs.create_index("mrf_id", unique=True)
    await db.mrfs.create_index("mrf_number", unique=True)
    await db.pos.create_index("po_id", unique=True)
    await db.pos.create_index("po_number", unique=True)
    await db.customers.create_index("customer_id", unique=True)
    await db.customer_pos.create_index("cpo_id", unique=True)
    # ---- Phase 4: Material / Variant masters ----
    await db.materials.create_index("material_uid", unique=True)
    await db.materials.create_index([("description_norm", 1)], unique=True, sparse=True)
    await db.variants.create_index("variant_uid", unique=True)
    await db.variants.create_index([("material_uid", 1), ("make_norm", 1), ("model_norm", 1)], unique=True)

@app.on_event("shutdown")
async def shutdown():
    client.close()

@api.get("/")
async def root():
    return {"app": "Vasu Infosec MRF & PO", "ok": True}

# ---------------------- Register extracted routers (Phase 9C) ----------------------
# Router-module imports MUST happen BEFORE app.include_router(api) so their
# @api.post/@api.get decorators are picked up when FastAPI snapshots the router.
from routers import llm as _llm_router  # noqa: E402,F401
from routers import notifications as _notifications_router  # noqa: E402,F401
from routers import reports as _reports_router  # noqa: E402,F401
from routers import audit as _audit_router  # noqa: E402,F401
from routers import settings as _settings_router  # noqa: E402,F401
from routers import masters as _masters_router  # noqa: E402,F401
from routers import customers as _customers_router  # noqa: E402,F401
from routers import cs as _cs_router  # noqa: E402,F401
from routers import dc as _dc_router  # noqa: E402,F401
from routers import grn as _grn_router  # noqa: E402,F401
from routers import mrf as _mrf_router  # noqa: E402,F401
from routers import po as _po_router  # noqa: E402,F401
from routers import billing as _billing_router  # noqa: E402,F401
from routers import invoice as _invoice_router  # noqa: E402,F401
from routers import imports as _imports_router  # noqa: E402,F401
from routers import vendors as _vendors_router  # noqa: E402,F401
from routers import projects as _projects_router  # noqa: E402,F401
from routers import auth as _auth_router  # noqa: E402,F401
from routers import po_exports as _po_exports_router  # noqa: E402,F401
from routers import ai_purchase as _ai_purchase_router  # noqa: E402,F401
from routers import access_security_v2 as _access_security_v2_router  # noqa: E402,F401

app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
