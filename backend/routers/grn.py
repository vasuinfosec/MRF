"""GRN (Goods Received Note) router — extracted from server.py in Phase 9C round 4.

Includes:
- List/GET/approve GRN lifecycle
- GRN PDF + Excel single-row export
- Bulk /export/grn Excel workbook
- Line-item Excel builder helper (`_grn_excel_workbook`) — used by both single &
  bulk exports.
- GRN export validator (`_validate_grn_for_export`).
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional, Dict, Any, List
from fastapi import Header, HTTPException, Body
from fastapi.responses import StreamingResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate

from server import (
    api, db, get_current_user, audit, now_utc,
    _check_po_access, _resolve_material_info, _pdf_story,
    _log_export, _vasu_footer, _excel_response, _validation_error,
    _safe_cell,
    VASU_PRIMARY, GRN_APPROVER_ROLES,
    bearer_from_url_token,
)


# ---------------------- GRN list / lookup / approve ----------------------
@api.get("/po/{po_id}/grns")
async def list_grns_for_po(po_id: str,
                            authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    po = await db.pos.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(404, "PO not found")
    await _check_po_access(u, po)
    return await db.grns.find({"po_id": po_id}, {"_id": 0}).sort("date", -1).to_list(200)


@api.get("/grns")
async def list_all_grns(authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("purchase", "gm", "director", "admin", "pm"):
        raise HTTPException(403, "Only purchase/pm/gm/director/admin")
    return await db.grns.find({}, {"_id": 0}).sort("date", -1).to_list(500)


@api.post("/grn/{grn_id}/approve")
async def approve_grn(grn_id: str, body: Dict[str, Any] = Body(default={}),
                       authorization: Optional[str] = Header(None)):
    """Admin/Accounts (or Director) approves or rejects a GRN."""
    u = await get_current_user(authorization)
    if u.role not in GRN_APPROVER_ROLES:
        raise HTTPException(403, "Only admin/accounts or director can approve GRN")
    grn = await db.grns.find_one({"grn_id": grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(404, "GRN not found")
    action = (body.get("action") or "approve").lower()
    if action not in ("approve", "reject"):
        raise HTTPException(400, "action must be 'approve' or 'reject'")
    current = grn.get("status") or "pending_approval"
    if current not in ("pending_approval", "rejected"):
        raise HTTPException(400, f"Cannot {action}: GRN is '{current}'")
    new_status = "approved" if action == "approve" else "rejected"
    hist = {
        "user_id": u.user_id, "user_name": u.name, "user_role": u.role,
        "action": action, "comment": body.get("comment") or "",
        "timestamp": now_utc().isoformat(),
    }
    await db.grns.update_one(
        {"grn_id": grn_id},
        {"$set": {"status": new_status},
         "$push": {"approval_history": hist}},
    )
    await audit("grn", grn_id, f"grn_{action}", u,
                 {"grn_number": grn.get("grn_number"),
                  "record_number": grn.get("grn_number"),
                  "old_value": {"status": current},
                  "new_value": {"status": new_status},
                  "reason": body.get("comment") or ""})
    return await db.grns.find_one({"grn_id": grn_id}, {"_id": 0})


@api.get("/grn/{grn_id}")
async def get_grn(grn_id: str, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    grn = await db.grns.find_one({"grn_id": grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(404, "GRN not found")
    po = await db.pos.find_one({"po_id": grn.get("po_id")}, {"_id": 0})
    if po:
        await _check_po_access(u, po)
    elif u.role not in ("purchase", "director", "admin", "store", "gm"):
        raise HTTPException(403, "Not allowed")
    return grn


# ---------------------- GRN PDF ----------------------
@api.get("/grn/{grn_id}/pdf")
async def grn_pdf(grn_id: str, token: Optional[str] = None,
                   force: Optional[int] = 0,
                   authorization: Optional[str] = Header(None)):
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    grn = await db.grns.find_one({"grn_id": grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(404, "GRN not found")
    po = await db.pos.find_one({"po_id": grn.get("po_id")}, {"_id": 0})
    if po:
        await _check_po_access(u, po)
    elif u.role not in ("purchase", "director", "admin", "store"):
        raise HTTPException(403, "Not allowed")
    vendor = await db.vendors.find_one({"vendor_id": grn.get("vendor_id")},
                                          {"_id": 0}) or {}
    project = await db.projects.find_one({"project_id": grn.get("project_id")},
                                            {"_id": 0}) or {}
    grn_mrf_nums: List[str] = []
    for mid in grn.get("mrf_refs") or []:
        m = await db.mrfs.find_one({"mrf_id": mid}, {"_id": 0, "mrf_number": 1})
        if m:
            grn_mrf_nums.append(m["mrf_number"])

    styles = getSampleStyleSheet()
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, leading=10)
    smallb = ParagraphStyle('sb', parent=styles['Normal'], fontSize=8, leading=10,
                              fontName="Helvetica-Bold")
    cid = (po or {}).get("customer_id") or project.get("customer_id")
    cust = await db.customers.find_one({"customer_id": cid}, {"_id": 0}) if cid else None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm,
                              leftMargin=15*mm, rightMargin=15*mm)
    story = _pdf_story("GOODS RECEIVED NOTE", grn["grn_number"],
                         grn.get("date"), project, cust, vendor,
                         extra_ref=f"PO {grn.get('po_number','')}  ·  MRF {', '.join(grn_mrf_nums)}")

    story.append(Paragraph(f"<b>Received by:</b> {grn.get('received_by_name','')}", small))
    if grn.get("vehicle_no"):
        story.append(Paragraph(f"<b>Vehicle:</b> {grn['vehicle_no']}  ·  "
                                 f"<b>Driver:</b> {grn.get('driver_name','')}", small))
    if grn.get("vendor_dc_ref"):
        story.append(Paragraph(f"<b>Vendor DC/Invoice ref:</b> {grn['vendor_dc_ref']}", small))
    story.append(Spacer(1, 6))

    header = ["#", "MAT / VAR", "Description", "Make / Model", "HSN", "Unit",
              "Qty Received", "Remarks"]
    data = [[Paragraph(h, smallb) for h in header]]
    for idx, it in enumerate(grn.get("items", []), 1):
        info = await _resolve_material_info(it)
        mv = ((info.get("material_uid") or "-") + "\n" + (info.get("variant_uid") or "-"))
        mm2 = ((info.get("make") or "-") + "\n" + (info.get("model") or "-"))
        data.append([
            Paragraph(str(idx), small),
            Paragraph(mv, small),
            Paragraph((it.get("description") or "")[:80], small),
            Paragraph(mm2, small),
            Paragraph(info.get("hsn_code") or "—", small),
            Paragraph(it.get("unit", ""), small),
            Paragraph(f"{float(it.get('qty') or 0):g}", small),
            Paragraph((it.get("remarks") or "")[:40], small),
        ])
    tbl = Table(data, colWidths=[16, 56, 150, 68, 44, 30, 52, 88], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), VASU_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor('#B7BEC9')),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#F7F8FC')]),
    ]))
    story.append(tbl)
    if grn.get("remarks"):
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<b>Remarks:</b> {grn['remarks']}", small))
    story += [
        Spacer(1, 30),
        Paragraph("_______________________________&nbsp;&nbsp;&nbsp;&nbsp;_______________________________", small),
        Paragraph("Received By&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Vendor Signature", small),
    ]
    doc.build(story, onFirstPage=_vasu_footer, onLaterPages=_vasu_footer)
    buf.seek(0)
    await _log_export(u, "grn", grn_id, "pdf", grn["grn_number"])
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition":
                 f'attachment; filename="{grn["grn_number"].replace("/", "_")}.pdf"'},
    )


# ---------------------- GRN Excel workbook helper ----------------------
async def _grn_excel_workbook(grns: List[dict]) -> "openpyxl.Workbook":
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GRN Line Items"
    headers = [
        "GRN#", "Date", "PO#", "MRF Refs", "Customer ID", "Customer Name",
        "Vendor Name", "Project Code", "Site",
        "Vehicle", "Driver", "Vendor DC/Inv Ref",
        "Line#", "MAT UID", "VAR UID", "Description", "Make", "Model", "HSN", "Unit",
        "Qty Received", "Remarks", "Received By",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    proj_map = {p["project_id"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(500)}
    vend_map = {v["vendor_id"]: v for v in await db.vendors.find({}, {"_id": 0}).to_list(2000)}
    cust_map = {c["customer_id"]: c for c in await db.customers.find({}, {"_id": 0}).to_list(500)}
    for g in grns:
        vend = vend_map.get(g.get("vendor_id")) or {}
        proj = proj_map.get(g.get("project_id")) or {}
        cust = cust_map.get(g.get("customer_id")) or {}
        mrfnums = ""
        if g.get("mrf_refs"):
            docs = await db.mrfs.find({"mrf_id": {"$in": g["mrf_refs"]}},
                                        {"_id": 0, "mrf_number": 1}).to_list(1000)
            mrfnums = ", ".join(sorted([m["mrf_number"] for m in docs]))
        gdate = g["date"].strftime("%Y-%m-%d") if isinstance(g.get("date"), datetime) else str(g.get("date", ""))[:10]
        for idx, it in enumerate(g.get("items") or [], 1):
            info = await _resolve_material_info(it)
            ws.append([
                _safe_cell(g.get("grn_number", "")), _safe_cell(gdate),
                _safe_cell(g.get("po_number", "")),
                _safe_cell(mrfnums),
                _safe_cell(cust.get("customer_id") or g.get("customer_id") or ""),
                _safe_cell(cust.get("name") or ""),
                _safe_cell(vend.get("name") or ""),
                _safe_cell(proj.get("code") or ""),
                _safe_cell(g.get("delivery_site") or proj.get("site") or ""),
                _safe_cell(g.get("vehicle_no") or ""),
                _safe_cell(g.get("driver_name") or ""),
                _safe_cell(g.get("vendor_dc_ref") or ""),
                idx,
                _safe_cell(info.get("material_uid") or ""),
                _safe_cell(info.get("variant_uid") or ""),
                _safe_cell(it.get("description") or ""),
                _safe_cell(info.get("make") or ""),
                _safe_cell(info.get("model") or ""),
                _safe_cell(info.get("hsn_code") or ""),
                _safe_cell(it.get("unit", "")),
                float(it.get("qty") or 0),
                _safe_cell(it.get("remarks") or ""),
                _safe_cell(g.get("received_by_name") or ""),
            ])
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(len(h) + 2, 28))
    return wb


def _validate_grn_for_export(grn: dict, project: Optional[dict],
                              customer: Optional[dict], fmt: str) -> tuple:
    missing: List[str] = []
    warnings: List[str] = []
    if not grn.get("grn_number"):
        missing.append("GRN number")
    if not project:
        warnings.append("Project")
    if not (customer and customer.get("customer_id")):
        warnings.append("Customer ID (permanent)")
    items = grn.get("items") or []
    if not items:
        missing.append("At least one received line")
    for i, it in enumerate(items, 1):
        if not (it.get("description") or "").strip():
            missing.append(f"Item {i} description")
        if float(it.get("qty") or 0) <= 0:
            missing.append(f"Item {i} qty > 0")
    return missing, warnings


@api.get("/grn/{grn_id}/export")
async def grn_export_dispatch(grn_id: str,
                                format: str = "pdf",
                                force: Optional[int] = 0,
                                token: Optional[str] = None,
                                authorization: Optional[str] = Header(None)):
    fmt = (format or "pdf").lower()
    if fmt not in ("pdf", "excel"):
        raise HTTPException(400, "format must be 'pdf' or 'excel'")
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    grn = await db.grns.find_one({"grn_id": grn_id}, {"_id": 0})
    if not grn:
        raise HTTPException(404, "GRN not found")
    po = await db.pos.find_one({"po_id": grn.get("po_id")}, {"_id": 0})
    if po:
        await _check_po_access(u, po)
    elif u.role not in ("purchase", "director", "admin", "store"):
        raise HTTPException(403, "Not allowed")

    if fmt == "pdf":
        return await grn_pdf(grn_id, token=token, force=force, authorization=authorization)

    project = await db.projects.find_one({"project_id": grn.get("project_id")}, {"_id": 0}) or {}
    cust = None
    cid = (po or {}).get("customer_id") or project.get("customer_id") or grn.get("customer_id")
    if cid:
        cust = await db.customers.find_one({"customer_id": cid}, {"_id": 0})
    missing, warnings = _validate_grn_for_export(grn, project, cust, fmt)
    if missing and not force:
        raise _validation_error(missing, warnings, fmt)
    wb = await _grn_excel_workbook([grn])
    await _log_export(u, "grn", grn_id, "excel", grn["grn_number"],
                        {"warnings": warnings, "forced": bool(force and missing)})
    return _excel_response(wb, f"{grn['grn_number'].replace('/', '_')}_details.xlsx")


@api.get("/export/grn")
async def export_grn(token: Optional[str] = None,
                      authorization: Optional[str] = Header(None)):
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    if u.role not in ("purchase", "director", "admin", "store", "pm", "gm"):
        raise HTTPException(403, "Not allowed")
    grns = await db.grns.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("date", -1).to_list(2000)
    wb = await _grn_excel_workbook(grns)
    await _log_export(u, "grn", "bulk", "excel", f"count={len(grns)}")
    return _excel_response(wb, "grn_export.xlsx")
