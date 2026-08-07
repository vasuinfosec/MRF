"""MRF (Material Requisition Form) router — extracted from server.py in Phase 9C round 4.

Handles the entire MRF lifecycle: draft → pm_review → approved → sent_to_purchase
→ partially/fully_ordered → received. Plus item-level edits with change_log
audit, plus PDF & Excel exports.

Access helpers `_check_mrf_access` and `_check_po_access` remain in `server.py`
because multiple routers (audit, grn, dc, invoice, po) import them.
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime
from typing import Optional, Dict, Any
from fastapi import Header, HTTPException
from fastapi.responses import StreamingResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from server import (
    api, db, get_current_user, audit, master_audit, notify, now_utc, gid,
    _check_mrf_access, _resolve_material_info,
    _log_export, _excel_response, _validation_error, _safe_cell,
    _mrf_pdf_bytes, _validate_mrf_for_export, _ALLOWED_MRF_FORMATS,
    canonical_mrf_status,
    next_mrf_number,
    MRFCreate, MRF, MRFItem, ApprovalAction,
    MRF_CREATORS,
    bearer_from_url_token,
    _project_material_balance,
)


async def _check_boq_balance(project_id: str, items: list,
                              exclude_mrf_id: Optional[str] = None) -> None:
    """Raise HTTP 400 with a boq_balance_exceeded payload if any item's
    qty_requested exceeds the project's assigned BOQ balance for that material.
    Items with no project_materials assignment are skipped (assignment is optional)."""
    for it in items:
        material_uid = it.get("material_id") if isinstance(it, dict) else getattr(it, "material_id", None)
        qty_requested = it.get("qty_requested") if isinstance(it, dict) else getattr(it, "qty_requested", None)
        if not material_uid or qty_requested is None:
            continue
        pm = await _project_material_balance(project_id, material_uid, exclude_mrf_id)
        if pm is None:
            continue
        if float(qty_requested) > pm["balance"]:
            raise HTTPException(400, {
                "error": "boq_balance_exceeded",
                "material_uid": material_uid,
                "requested": float(qty_requested),
                "balance": pm["balance"],
                "message": (f"Requested qty ({qty_requested}) exceeds the approved BOQ "
                            f"balance ({pm['balance']}) for this material. Contact Admin "
                            f"to increase the allocation."),
            })


# ---------------------- MRF create ----------------------
@api.post("/mrf")
async def create_mrf(body: MRFCreate, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in MRF_CREATORS:
        raise HTTPException(403, "Not allowed to create MRF")
    if u.role in ("site_engineer", "pm"):
        p = await db.projects.find_one({"project_id": body.project_id}, {"_id": 0})
        if not p:
            raise HTTPException(404, "Project not found")
        if u.role == "site_engineer":
            in_project = u.user_id in (p.get("site_engineers") or [])
            in_any_site = any(u.user_id in (s.get("site_engineers") or [])
                                for s in (p.get("sites") or []) if s.get("active", True))
            if not (in_project or in_any_site):
                raise HTTPException(403, "You are not assigned to this project")
        else:  # pm
            if u.user_id not in (p.get("project_managers") or []):
                raise HTTPException(403, "You are not the PM for this project")
    await _check_boq_balance(body.project_id, [i.model_dump() for i in body.items])
    items = []
    for i in body.items:
        d = i.model_dump()
        if d.get("qty_approved") is None:
            d["qty_approved"] = i.qty_requested
        items.append(MRFItem(**d))
    proj = await db.projects.find_one({"project_id": body.project_id}, {"_id": 0})
    cust_id = (proj or {}).get("customer_id")
    cust_name = ""
    if cust_id:
        c = await db.customers.find_one({"customer_id": cust_id},
                                          {"_id": 0, "name": 1})
        cust_name = (c or {}).get("name") or ""
    if not cust_name:
        cust_name = (proj or {}).get("client") or ""
    mrf = MRF(
        mrf_number=await next_mrf_number(),
        project_id=body.project_id,
        site=body.site,
        required_by=body.required_by,
        requesting_person=body.requesting_person,
        system_category=body.system_category,
        customer_id=cust_id,
        customer_name=cust_name,
        items=items,
        attachments=body.attachments,
        remarks=body.remarks or "",
        status="draft",
        created_by=u.user_id,
    )
    await db.mrfs.insert_one(mrf.model_dump())
    await audit("mrf", mrf.mrf_id, "create", u,
                 {"mrf_number": mrf.mrf_number,
                  "old_value": None,
                  "new_value": {"status": "draft",
                                 "project_id": mrf.project_id,
                                 "customer_id": cust_id,
                                 "customer_name": cust_name,
                                 "item_count": len(items),
                                 "site": mrf.site}})
    return mrf.model_dump()


# ---------------------- MRF list & lookup ----------------------
@api.get("/mrf")
async def list_mrfs(status: Optional[str] = None,
                     project_id: Optional[str] = None,
                     system_category: Optional[str] = None,
                     authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    q: Dict[str, Any] = {"deleted": False}
    if status: q["status"] = status
    if project_id: q["project_id"] = project_id
    if system_category: q["system_category"] = system_category
    if u.role == "site_engineer":
        q["created_by"] = u.user_id
    elif u.role == "pm":
        prjs = await db.projects.find(
            {"project_managers": u.user_id}, {"_id": 0, "project_id": 1}
        ).to_list(500)
        q["project_id"] = {"$in": [p["project_id"] for p in prjs]}
    elif u.role == "store":
        prjs = await db.projects.find(
            {"$or": [
                {"site_engineers": u.user_id},
                {"sites.site_engineers": u.user_id},
            ]}, {"_id": 0, "project_id": 1}
        ).to_list(500)
        q["project_id"] = {"$in": [p["project_id"] for p in prjs]}
    return await db.mrfs.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.get("/mrf/{mrf_id}")
async def get_mrf(mrf_id: str, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    d = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "MRF not found")
    await _check_mrf_access(u, d)
    return d


# ---------------------- MRF workflow transitions ----------------------
@api.post("/mrf/{mrf_id}/submit")
async def submit_mrf(mrf_id: str, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    d = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "MRF not found")
    if u.role not in ("admin", "director") and d.get("created_by") != u.user_id:
        raise HTTPException(403, "Only the MRF creator (or admin) can submit")
    if d["status"] not in ["draft", "returned"]:
        raise HTTPException(400, "Cannot submit from current status")
    await _check_boq_balance(d["project_id"], d.get("items", []), exclude_mrf_id=mrf_id)
    await db.mrfs.update_one(
        {"mrf_id": mrf_id},
        {"$set": {"status": "pm_review", "updated_at": now_utc()}}
    )
    await audit("mrf", mrf_id, "submit", u,
                 {"old_value": {"status": d["status"]},
                  "new_value": {"status": "pm_review"},
                  "mrf_number": d["mrf_number"]})
    proj = await db.projects.find_one({"project_id": d["project_id"]}, {"_id": 0}) or {}
    pm_ids = list(proj.get("project_managers") or [])
    if not pm_ids:
        pms = await db.users.find({"role": "pm"}, {"_id": 0, "user_id": 1}).to_list(50)
        pm_ids = [p["user_id"] for p in pms]
    await notify(pm_ids, "MRF pending review",
                   f"{d['mrf_number']} needs your review", f"/mrf/{mrf_id}")
    return {"ok": True}


@api.post("/mrf/{mrf_id}/approve")
async def approve_mrf(mrf_id: str, body: ApprovalAction,
                        authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("pm", "gm", "director", "admin"):
        raise HTTPException(403, "Only PM/GM/Director/admin can authorise")
    d = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "MRF not found")
    if u.role == "pm":
        proj = await db.projects.find_one({"project_id": d["project_id"]}, {"_id": 0})
        if not proj or u.user_id not in (proj.get("project_managers") or []):
            raise HTTPException(403, "You are not the PM for this project")

    if body.action == "return":
        await db.mrfs.update_one(
            {"mrf_id": mrf_id},
            {"$set": {"status": "returned",
                        "pm_comments": body.comment or "",
                        "updated_at": now_utc()}}
        )
        await audit("mrf", mrf_id, "return", u,
                     {"reason": body.comment or "",
                      "comment": body.comment,
                      "old_value": {"status": d["status"]},
                      "new_value": {"status": "returned"},
                      "mrf_number": d["mrf_number"]})
        await notify([d["created_by"]], "MRF returned",
                       f"{d['mrf_number']} returned for correction", f"/mrf/{mrf_id}")
        return {"ok": True, "status": "returned"}

    items = d["items"]
    if body.item_actions:
        for ia in body.item_actions:
            for it in items:
                if it["item_line_id"] == ia.get("item_line_id"):
                    if ia.get("action") == "reject":
                        it["status"] = "rejected"
                        it["rejection_reason"] = ia.get("reason", "")
                        it["qty_approved"] = 0
                    else:
                        it["status"] = "approved"
                        if ia.get("qty_approved") is not None:
                            it["qty_approved"] = float(ia["qty_approved"])
    else:
        for it in items:
            it["status"] = "approved"

    all_rejected = all(it["status"] == "rejected" for it in items)
    if body.action == "reject" or all_rejected:
        new_status = "rejected"
    else:
        new_status = "approved"

    await db.mrfs.update_one(
        {"mrf_id": mrf_id},
        {"$set": {"items": items, "status": new_status,
                    "pm_comments": body.comment or "", "updated_at": now_utc()}}
    )
    await audit("mrf", mrf_id, body.action, u,
                 {"reason": body.comment or "", "comment": body.comment,
                  "items": body.item_actions,
                  "old_value": {"status": d.get("status")},
                  "new_value": {"status": new_status},
                  "mrf_number": d["mrf_number"]})
    tgt = [d["created_by"]]
    if new_status == "approved":
        purchases = await db.users.find({"role": "purchase"},
                                          {"_id": 0, "user_id": 1}).to_list(50)
        tgt += [p["user_id"] for p in purchases]
    await notify(tgt, f"MRF {new_status}",
                   f"{d['mrf_number']} {new_status}", f"/mrf/{mrf_id}")
    return {"ok": True, "status": new_status}


@api.post("/mrf/{mrf_id}/send-to-purchase")
async def send_to_purchase(mrf_id: str,
                             authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("pm", "gm", "director", "admin"):
        raise HTTPException(403, "Not allowed")
    d = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not d or d["status"] != "approved":
        raise HTTPException(400, "MRF must be approved first")
    await db.mrfs.update_one({"mrf_id": mrf_id},
                              {"$set": {"status": "sent_to_purchase",
                                          "updated_at": now_utc()}})
    await audit("mrf", mrf_id, "send_to_purchase", u,
                 {"old_value": {"status": d.get("status")},
                  "new_value": {"status": "sent_to_purchase"},
                  "mrf_number": d.get("mrf_number")})
    return {"ok": True}


# ---------------------- MRF edit / delete ----------------------
@api.put("/mrf/{mrf_id}")
async def update_mrf_draft(mrf_id: str, body: dict,
                             authorization: Optional[str] = Header(None)):
    """Update a Draft or Returned MRF. Site engineers can only edit their own drafts."""
    u = await get_current_user(authorization)
    mrf = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not mrf:
        raise HTTPException(404, "MRF not found")
    editable_states = {"draft", "returned"}
    if canonical_mrf_status(mrf.get("status")) not in editable_states \
            and mrf.get("status") not in editable_states:
        raise HTTPException(
            400,
            f"MRF is '{mrf.get('status')}' — cannot edit after submission (only 'returned' MRFs can be re-edited)",
        )
    if u.role not in ("admin", "director") and mrf.get("created_by") != u.user_id:
        raise HTTPException(403, "Only the MRF creator (or admin/director) can edit this draft")

    updates: Dict[str, Any] = {}
    for k in ("site", "required_by", "requesting_person", "system_category",
              "remarks", "attachments"):
        if k in body:
            updates[k] = body[k]
    if "items" in body and isinstance(body["items"], list):
        new_items = []
        old_items = {it.get("item_line_id"): it for it in mrf.get("items", [])}
        for i in body["items"]:
            line_id = i.get("item_line_id") or f"mli_{uuid.uuid4().hex[:12]}"
            base = old_items.get(line_id) or {}
            merged = {**base, **i, "item_line_id": line_id}
            merged.setdefault("status", "pending")
            merged.setdefault("billing_status", "not_billed")
            merged.setdefault("qty_received", 0)
            merged.setdefault("qty_issued", 0)
            merged.setdefault("qty_billed", 0)
            merged.setdefault("qty_ordered", 0)
            if merged.get("qty_approved") is None:
                merged["qty_approved"] = merged.get("qty_requested")
            new_items.append(merged)
        target_project_id = body.get("project_id") or mrf.get("project_id")
        await _check_boq_balance(target_project_id, new_items, exclude_mrf_id=mrf_id)
        updates["items"] = new_items

    if body.get("project_id") and body["project_id"] != mrf.get("project_id"):
        proj = await db.projects.find_one({"project_id": body["project_id"]},
                                             {"_id": 0}) or {}
        cid = proj.get("customer_id")
        cust_name = ""
        if cid:
            c = await db.customers.find_one({"customer_id": cid}, {"_id": 0}) or {}
            cust_name = c.get("name") or ""
        updates["project_id"] = body["project_id"]
        updates["customer_id"] = cid
        updates["customer_name"] = cust_name or proj.get("client") or ""

    if not updates:
        raise HTTPException(400, "Nothing to update")

    updates["updated_at"] = now_utc()
    await db.mrfs.update_one({"mrf_id": mrf_id}, {"$set": updates})
    old_snap = {k: mrf.get(k) for k in updates.keys() if k != "updated_at"}
    new_snap = {k: v for k, v in updates.items() if k != "updated_at"}
    await audit("mrf", mrf_id, "edit_draft", u,
                 {"old_value": old_snap, "new_value": new_snap,
                  "mrf_number": mrf.get("mrf_number"),
                  "reason": body.get("reason") or ""})
    return await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})


@api.delete("/mrf/{mrf_id}")
async def delete_mrf(mrf_id: str, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role != "admin":
        raise HTTPException(403, "Only admin can delete")
    m = await db.mrfs.find_one({"mrf_id": mrf_id},
                                 {"_id": 0, "mrf_number": 1, "status": 1})
    await db.mrfs.update_one({"mrf_id": mrf_id}, {"$set": {"deleted": True}})
    await audit("mrf", mrf_id, "soft_delete", u,
                 {"old_value": {"deleted": False, "status": (m or {}).get("status")},
                  "new_value": {"deleted": True},
                  "mrf_number": (m or {}).get("mrf_number")})
    return {"ok": True}


@api.put("/mrf/{mrf_id}/items/{line_id}")
async def update_mrf_line(mrf_id: str, line_id: str, body: dict,
                            authorization: Optional[str] = Header(None)):
    """Edit an MRF line item WITH mandatory reason. Records field-level change_log."""
    u = await get_current_user(authorization)
    reason = (body.pop("reason", "") or "").strip()
    if not reason:
        raise HTTPException(400, "A reason is required for line-item edits")

    mrf = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not mrf:
        raise HTTPException(404, "MRF not found")
    await _check_mrf_access(u, mrf)
    if u.role not in ("pm", "gm", "director", "admin") and mrf.get("created_by") != u.user_id:
        raise HTTPException(403, "Not allowed to edit this MRF")
    locked = {"received", "fully_received", "closed", "cancelled", "rejected"}
    if canonical_mrf_status(mrf.get("status")) in locked and u.role not in ("admin", "director"):
        raise HTTPException(400, f"MRF is in status '{mrf.get('status')}' — line edits are locked")

    items = mrf.get("items", [])
    target = next((it for it in items if it.get("item_line_id") == line_id), None)
    if not target:
        raise HTTPException(404, "Line item not found")

    changed: Dict[str, Any] = {}
    editable_fields = ["description", "specification", "part_number", "make", "model",
                        "unit", "qty_requested", "qty_approved", "priority", "remarks",
                        "billing_status", "status"]
    for k in editable_fields:
        if k in body:
            new_v = body[k]
            if k in ("qty_requested", "qty_approved"):
                try:
                    new_v = float(new_v) if new_v is not None else new_v
                except Exception:
                    raise HTTPException(400, f"{k} must be numeric")
            if k == "priority" and new_v not in (None, "", "low", "normal", "high", "urgent"):
                raise HTTPException(400, "priority must be one of low, normal, high, urgent")
            old_v = target.get(k)
            if new_v != old_v:
                changed[k] = {"old": old_v, "new": new_v}
                target[k] = new_v
    if not changed:
        raise HTTPException(400, "No changes provided")

    log_entry = {
        "log_id": gid("chl"),
        "user_id": u.user_id, "user_name": u.name, "user_role": u.role,
        "timestamp": now_utc().isoformat(),
        "reason": reason, "changes": changed,
    }
    target.setdefault("change_log", []).append(log_entry)
    await db.mrfs.update_one({"mrf_id": mrf_id},
                               {"$set": {"items": items, "updated_at": now_utc()}})
    await master_audit("mrf.item", f"{mrf['mrf_number']}#{line_id}", "update", u,
                        reason,
                        {k: v["old"] for k, v in changed.items()},
                        {k: v["new"] for k, v in changed.items()})
    return {"ok": True, "changes": changed, "log": log_entry}


# ---------------------- MRF exports ----------------------
@api.get("/mrf/{mrf_id}/pdf")
async def mrf_pdf_route(mrf_id: str, token: Optional[str] = None,
                          force: Optional[int] = 0,
                          authorization: Optional[str] = Header(None)):
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    mrf = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not mrf:
        raise HTTPException(404, "MRF not found")
    await _check_mrf_access(u, mrf)
    project = await db.projects.find_one({"project_id": mrf["project_id"]},
                                            {"_id": 0}) or {}
    cid = mrf.get("customer_id") or project.get("customer_id")
    cust = await db.customers.find_one({"customer_id": cid}, {"_id": 0}) if cid else None
    missing, warnings = _validate_mrf_for_export(mrf, project, cust, "pdf")
    if missing and not force:
        raise _validation_error(missing, warnings, "pdf")
    pdf_bytes = await _mrf_pdf_bytes(mrf, project, cust)
    await _log_export(u, "mrf", mrf_id, "pdf", mrf.get("mrf_number", ""),
                        {"warnings": warnings, "forced": bool(force and missing)})
    return StreamingResponse(
        io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition":
                 f'attachment; filename="{mrf["mrf_number"].replace("/", "_")}.pdf"'},
    )


@api.get("/mrf/{mrf_id}/export")
async def mrf_export_dispatch(mrf_id: str,
                                format: str = "pdf",
                                force: Optional[int] = 0,
                                token: Optional[str] = None,
                                authorization: Optional[str] = Header(None)):
    fmt = (format or "pdf").lower()
    if fmt not in _ALLOWED_MRF_FORMATS:
        raise HTTPException(400, f"format must be one of {_ALLOWED_MRF_FORMATS}")
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    mrf = await db.mrfs.find_one({"mrf_id": mrf_id}, {"_id": 0})
    if not mrf:
        raise HTTPException(404, "MRF not found")
    await _check_mrf_access(u, mrf)

    if fmt == "pdf":
        return await mrf_pdf_route(mrf_id, token=token, force=force,
                                     authorization=authorization)

    project = await db.projects.find_one({"project_id": mrf["project_id"]},
                                            {"_id": 0}) or {}
    cid = mrf.get("customer_id") or project.get("customer_id")
    cust = await db.customers.find_one({"customer_id": cid}, {"_id": 0}) if cid else None
    missing, warnings = _validate_mrf_for_export(mrf, project, cust, "excel")
    if missing and not force:
        raise _validation_error(missing, warnings, "excel")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MRF Line Items"
    headers = [
        "MRF#", "Date", "Status", "Customer ID", "Customer Name",
        "Project", "Site", "Requester", "System",
        "Line#", "MAT UID", "VAR UID", "Description", "Make", "Model",
        "Unit", "Qty Requested", "Qty Approved", "Qty Ordered", "Qty Received",
        "Priority", "Item Status", "Billing Status", "Purpose", "Remarks",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for idx, it in enumerate(mrf.get("items") or [], 1):
        info = await _resolve_material_info(it)
        ws.append([
            _safe_cell(mrf.get("mrf_number", "")),
            mrf["date"].strftime("%Y-%m-%d") if isinstance(mrf.get("date"), datetime) else str(mrf.get("date", "")),
            _safe_cell(mrf.get("status", "")),
            _safe_cell(mrf.get("customer_id") or ""),
            _safe_cell(mrf.get("customer_name") or ""),
            _safe_cell(project.get("code") or mrf.get("project_id", "")),
            _safe_cell(mrf.get("site", "")),
            _safe_cell(mrf.get("requesting_person", "")),
            _safe_cell(mrf.get("system_category", "")),
            idx,
            _safe_cell(info.get("material_uid") or ""),
            _safe_cell(info.get("variant_uid") or ""),
            _safe_cell(it.get("description", "")),
            _safe_cell(info.get("make") or ""),
            _safe_cell(info.get("model") or ""),
            _safe_cell(it.get("unit", "")),
            float(it.get("qty_requested") or 0),
            float(it.get("qty_approved") or 0),
            float(it.get("qty_ordered") or 0),
            float(it.get("qty_received") or 0),
            _safe_cell((it.get("priority") or "normal")),
            _safe_cell(it.get("status", "")),
            _safe_cell(it.get("billing_status", "")),
            _safe_cell(it.get("purpose", "")),
            _safe_cell(it.get("remarks", "")),
        ])
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(len(h) + 2, 28))
    await _log_export(u, "mrf", mrf_id, "excel", mrf.get("mrf_number", ""),
                        {"warnings": warnings, "forced": bool(force and missing)})
    return _excel_response(wb, f"{mrf['mrf_number'].replace('/', '_')}_details.xlsx")
