"""Masters router — extracted from server.py in Phase 9C round 3.

Groups all material / variant / master-data endpoints:
  - /materials + counts + PM/GM approvals + bulk-action + bulk-import + template
  - /variants  (list, create)
  - /master-audit  (admin/director-only view)
  - /systems  (constant list)
  - /masters (generic category-keyed master items: units, brands, systems, etc.)

Kept out on purpose: /vendors GET/POST (fragmented across server.py; will be
extracted in the vendors router pass).
"""
from __future__ import annotations

import re
from typing import Optional, Dict, Any, List
from fastapi import Header, HTTPException

from server import (
    api, db, get_current_user, audit, master_audit,
    _norm, _word_count, _next_seq, _mat_uid, _ensure_variant,
    _strip_oids, _excel_response, now_utc,
    MAT_WORD_LIMIT, SYSTEMS, MaterialIn, MasterItem,
    openpyxl,
)
from openpyxl.styles import Font, PatternFill


# ---------------------- Materials ----------------------
@api.get("/materials")
async def list_materials(status: Optional[str] = None,
                         q: Optional[str] = None,
                         authorization: Optional[str] = Header(None)):
    """List materials. Default returns ONLY approved. Pass ?status=all to see everything."""
    await get_current_user(authorization)
    filt: Dict[str, Any] = {}
    if status and status != "all":
        filt["status"] = status
    elif not status:
        filt["status"] = "approved"
    if q:
        filt["$or"] = [
            {"description": {"$regex": re.escape(q), "$options": "i"}},
            {"item_code": {"$regex": re.escape(q), "$options": "i"}},
            {"material_uid": {"$regex": re.escape(q), "$options": "i"}},
        ]
    docs = await db.materials.find(filt, {"_id": 0}).sort([("material_uid", 1)]).to_list(2000)
    return docs


@api.get("/materials/counts")
async def material_counts(authorization: Optional[str] = Header(None)):
    await get_current_user(authorization)
    out: Dict[str, int] = {}
    async for r in db.materials.aggregate([{"$group": {"_id": "$status", "n": {"$sum": 1}}}]):
        out[r["_id"]] = r["n"]
    return out


@api.post("/materials")
async def create_material(body: MaterialIn, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("purchase", "admin", "director"):
        raise HTTPException(403, "Only Purchase/Admin/Director can add materials")
    desc = (body.description or "").strip()
    if not desc:
        raise HTTPException(400, "Description is required")
    if _word_count(desc) > MAT_WORD_LIMIT:
        raise HTTPException(400, f"Description exceeds {MAT_WORD_LIMIT} words")
    norm = _norm(desc)
    dup = await db.materials.find_one({"description_norm": norm},
                                       {"_id": 0, "material_uid": 1, "status": 1})
    if dup:
        raise HTTPException(400, f"Material already exists as {dup['material_uid']} (status: {dup['status']})")
    n = await _next_seq("materials")
    doc = {
        "material_uid": _mat_uid(n),
        "description": desc,
        "description_norm": norm,
        "category": (body.category or "").strip(),
        "unit": (body.unit or "").strip(),
        "gst_rate": float(body.gst_rate) if body.gst_rate is not None else None,
        "item_code": (body.item_code or "").strip(),
        "remarks": (body.remarks or "").strip(),
        # Admin/Director bypass PM review (still audited)
        "status": "approved" if u.role in ("admin", "director") else "pending_pm_review",
        "created_by": u.user_id, "created_by_name": u.name,
        "created_at": now_utc(), "updated_at": now_utc(),
        "approved_by": u.user_id if u.role in ("admin", "director") else None,
        "approved_at": now_utc() if u.role in ("admin", "director") else None,
    }
    await db.materials.insert_one(doc)
    doc.pop("_id", None)
    await master_audit("material", doc["material_uid"], "create", u,
                        "New material added", None, doc)
    return doc


@api.put("/materials/{material_uid}")
async def update_material(material_uid: str, body: dict,
                           authorization: Optional[str] = Header(None)):
    """Edit a material. If approved, changes go to pending_gm_approval (GM approves).
    If it's still pending_pm_review, edits are inline (Purchase can amend before PM sees it)."""
    u = await get_current_user(authorization)
    if u.role not in ("purchase", "admin", "director"):
        raise HTTPException(403, "Only Purchase/Admin/Director can edit materials")
    mat = await db.materials.find_one({"material_uid": material_uid}, {"_id": 0})
    if not mat:
        raise HTTPException(404, "Material not found")
    reason = (body.pop("reason", "") or "").strip()
    if not reason:
        raise HTTPException(400, "Reason is required for material changes")
    updates: Dict[str, Any] = {}
    for k in ("description", "category", "unit", "item_code", "remarks"):
        if k in body:
            updates[k] = (str(body[k] or "")).strip()
    if "gst_rate" in body:
        try:
            updates["gst_rate"] = float(body["gst_rate"]) if body["gst_rate"] is not None else None
        except Exception:
            raise HTTPException(400, "gst_rate must be numeric")
    if not updates:
        raise HTTPException(400, "Nothing to update")
    if "description" in updates:
        if _word_count(updates["description"]) > MAT_WORD_LIMIT:
            raise HTTPException(400, f"Description exceeds {MAT_WORD_LIMIT} words")
        updates["description_norm"] = _norm(updates["description"])
        dup = await db.materials.find_one({
            "description_norm": updates["description_norm"],
            "material_uid": {"$ne": material_uid},
        })
        if dup:
            raise HTTPException(400, f"Another material already has this description: {dup['material_uid']}")
    was_approved = mat.get("status") == "approved"
    if was_approved and u.role not in ("admin", "director"):
        updates["status"] = "pending_gm_approval"
        updates["pending_change"] = {"changed_by": u.user_id,
                                     "changed_at": now_utc().isoformat(),
                                     "reason": reason}
    updates["updated_at"] = now_utc()
    await db.materials.update_one({"material_uid": material_uid}, {"$set": updates})
    old_snap = {k: mat.get(k) for k in updates if k in mat}
    await master_audit("material", material_uid,
                        "edit_requires_gm" if was_approved else "update",
                        u, reason, old_snap, updates)
    return await db.materials.find_one({"material_uid": material_uid}, {"_id": 0})


@api.post("/materials/{material_uid}/action")
async def material_action(material_uid: str, body: dict,
                           authorization: Optional[str] = Header(None)):
    """PM approve / reject a pending material. Also handles 'needs_correction' flag toggle."""
    u = await get_current_user(authorization)
    if u.role not in ("pm", "admin", "director"):
        raise HTTPException(403, "Only PM/Admin/Director can approve materials")
    action = (body.get("action") or "").lower()
    reason = (body.get("reason") or "").strip()
    if action not in ("approve", "reject", "flag"):
        raise HTTPException(400, "action must be approve/reject/flag")
    mat = await db.materials.find_one({"material_uid": material_uid}, {"_id": 0})
    if not mat:
        raise HTTPException(404, "Material not found")
    if mat.get("status") not in ("pending_pm_review", "needs_correction"):
        raise HTTPException(400, f"Cannot {action}: material is '{mat.get('status')}'")
    new_status = {"approve": "approved", "reject": "rejected",
                  "flag": "needs_correction"}[action]
    upd: Dict[str, Any] = {"status": new_status, "updated_at": now_utc()}
    if action == "approve":
        upd["approved_by"] = u.user_id
        upd["approved_at"] = now_utc()
    await db.materials.update_one({"material_uid": material_uid}, {"$set": upd})
    await master_audit("material", material_uid, f"pm_{action}", u,
                        reason or f"PM {action}", {"status": mat.get("status")}, upd)
    return await db.materials.find_one({"material_uid": material_uid}, {"_id": 0})


@api.post("/materials/{material_uid}/gm-approve")
async def gm_approve_material(material_uid: str, body: dict,
                               authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("gm", "director", "admin"):
        raise HTTPException(403, "Only GM/Director/Admin can approve material changes")
    mat = await db.materials.find_one({"material_uid": material_uid}, {"_id": 0})
    if not mat:
        raise HTTPException(404, "Material not found")
    if mat.get("status") != "pending_gm_approval":
        raise HTTPException(400, f"Not awaiting GM approval (status={mat.get('status')})")
    action = (body.get("action") or "approve").lower()
    reason = (body.get("reason") or "").strip()
    new_status = "approved" if action == "approve" else "rejected"
    await db.materials.update_one({"material_uid": material_uid},
                                    {"$set": {"status": new_status,
                                              "gm_action_by": u.user_id,
                                              "gm_action_at": now_utc(),
                                              "pending_change": None,
                                              "updated_at": now_utc()}})
    await master_audit("material", material_uid, f"gm_{action}", u, reason,
                        {"status": "pending_gm_approval"}, {"status": new_status})
    return await db.materials.find_one({"material_uid": material_uid}, {"_id": 0})


@api.post("/materials/bulk-action")
async def bulk_material_action(body: dict, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("pm", "admin", "director"):
        raise HTTPException(403, "Only PM/Admin/Director")
    action = (body.get("action") or "").lower()
    if action not in ("approve", "reject"):
        raise HTTPException(400, "action must be approve/reject")
    uids = body.get("material_uids") or []
    reason = body.get("reason") or f"Bulk {action}"
    ok, skipped = 0, 0
    for uid in uids:
        mat = await db.materials.find_one({"material_uid": uid}, {"_id": 0})
        if not mat or mat.get("status") not in ("pending_pm_review", "needs_correction"):
            skipped += 1
            continue
        new_status = "approved" if action == "approve" else "rejected"
        upd = {"status": new_status, "updated_at": now_utc()}
        if action == "approve":
            upd["approved_by"] = u.user_id
            upd["approved_at"] = now_utc()
        await db.materials.update_one({"material_uid": uid}, {"$set": upd})
        await master_audit("material", uid, f"pm_{action}_bulk", u, reason,
                            {"status": mat.get("status")}, upd)
        ok += 1
    return {"ok": ok, "skipped": skipped}


# ---------------------- Bulk import (Excel/CSV) ----------------------
@api.get("/import/materials/template")
async def material_import_template(token: Optional[str] = None,
                                    authorization: Optional[str] = Header(None)):
    auth = authorization or (f"Bearer {token}" if token else None)
    u = await get_current_user(auth)
    if u.role not in ("purchase", "admin", "director"):
        raise HTTPException(403, "Only Purchase/Admin/Director")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materials"
    headers = ["Material Description", "Category", "Make", "Model", "Unit",
               "GST Rate %", "Item Code", "Remarks"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
    ws.append(["Smoke Detector Photoelectric 2-wire", "Fire Alarm", "Legrand",
                "XC-500", "Nos", 18, "SD-2W", "Sample row — delete before upload"])
    ws.append(["Fire Alarm Control Panel 4-Zone", "Fire Alarm", "Honeywell",
                "NFS2-3030", "Nos", 18, "FACP-4Z", ""])
    widths = [50, 20, 18, 18, 10, 12, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    return _excel_response(wb, "material_import_template.xlsx")


@api.post("/import/materials")
async def import_materials(payload: dict, authorization: Optional[str] = Header(None)):
    """Body: {'rows': [{Material Description, Category, Make, Model, Unit, GST Rate %, Item Code, Remarks}]}."""
    u = await get_current_user(authorization)
    if u.role not in ("purchase", "admin", "director"):
        raise HTTPException(403, "Only Purchase can bulk-import materials")
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(400, "No rows to import")
    created, duplicates, errors, needs_correction = [], [], [], []
    for idx, r in enumerate(rows, 2):
        desc = str(r.get("Material Description") or r.get("description") or "").strip()
        if not desc:
            errors.append({"row": idx, "reason": "empty description"})
            continue
        wc = _word_count(desc)
        if wc > MAT_WORD_LIMIT:
            errors.append({"row": idx, "description": desc[:60] + "…",
                            "reason": f"description has {wc} words (max {MAT_WORD_LIMIT})"})
            needs_correction.append({"row": idx, "description": desc, "words": wc})
            continue
        norm = _norm(desc)
        existing = await db.materials.find_one({"description_norm": norm},
                                                 {"_id": 0, "material_uid": 1})
        if existing:
            duplicates.append({"row": idx, "material_uid": existing["material_uid"],
                                "description": desc})
            continue
        try:
            gst = float(r.get("GST Rate %") or r.get("gst_rate") or 0)
        except Exception:
            gst = None
        n = await _next_seq("materials")
        doc = {
            "material_uid": _mat_uid(n),
            "description": desc, "description_norm": norm,
            "category": str(r.get("Category") or r.get("category") or "").strip(),
            "make": str(r.get("Make") or r.get("make") or "").strip(),
            "model": str(r.get("Model") or r.get("model") or "").strip(),
            "unit": str(r.get("Unit") or r.get("unit") or "").strip(),
            "gst_rate": gst,
            "item_code": str(r.get("Item Code") or r.get("item_code") or "").strip(),
            "remarks": str(r.get("Remarks") or r.get("remarks") or "").strip(),
            "status": "pending_pm_review",
            "created_by": u.user_id, "created_by_name": u.name,
            "created_at": now_utc(), "updated_at": now_utc(),
            "source": "bulk_import",
        }
        await db.materials.insert_one(doc)
        doc.pop("_id", None)
        if doc["make"]:
            await _ensure_variant(doc["material_uid"], doc["make"],
                                   doc["model"], doc["material_uid"])
        created.append({"material_uid": doc["material_uid"], "description": desc})
    await master_audit("material.bulk_import",
                        f"batch_{now_utc().timestamp()}", "bulk_import", u,
                        f"Purchase bulk import: {len(created)} created, {len(duplicates)} dup, {len(errors)} err",
                        None, {"created": len(created),
                                "duplicates": len(duplicates),
                                "errors": len(errors)})
    return {
        "created": created, "duplicates": duplicates, "errors": errors,
        "needs_correction": needs_correction,
        "summary": {
            "created": len(created), "duplicates": len(duplicates),
            "errors": len(errors), "flagged_word_limit": len(needs_correction),
        },
    }


# ---------------------- Variants ----------------------
@api.get("/variants")
async def list_variants(material_uid: Optional[str] = None,
                         authorization: Optional[str] = Header(None)):
    await get_current_user(authorization)
    q: Dict[str, Any] = {}
    if material_uid:
        q["material_uid"] = material_uid
    return await db.variants.find(q, {"_id": 0}).sort("variant_uid", 1).to_list(2000)


@api.post("/variants")
async def create_variant(body: dict, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("purchase", "site_engineer", "pm", "admin", "director"):
        raise HTTPException(403, "Not allowed")
    material_uid = (body.get("material_uid") or "").strip()
    if not material_uid:
        raise HTTPException(400, "material_uid required")
    mat = await db.materials.find_one({"material_uid": material_uid},
                                        {"_id": 0, "status": 1})
    if not mat:
        raise HTTPException(404, "Material not found")
    if mat.get("status") != "approved" and u.role not in ("admin", "director"):
        raise HTTPException(400, "Cannot add variants to un-approved materials")
    return await _ensure_variant(
        material_uid,
        body.get("make") or "",
        body.get("model") or "",
        material_uid,
    )


# ---------------------- Master-data audit log (view) ----------------------
@api.get("/master-audit")
async def master_audit_list(entity: Optional[str] = None,
                              entity_id: Optional[str] = None,
                              authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("admin", "director"):
        raise HTTPException(403, "Only admin/director")
    q: Dict[str, Any] = {}
    if entity:
        q["entity"] = entity
    if entity_id:
        q["entity_id"] = entity_id
    docs = await db.master_audit_logs.find(q, {"_id": 0}).sort("timestamp", -1).limit(500).to_list(500)
    return [_strip_oids(d) for d in docs]


# ---------------------- Systems (constant list) ----------------------
@api.get("/systems")
async def list_systems(authorization: Optional[str] = Header(None)):
    await get_current_user(authorization)
    return SYSTEMS


# ---------------------- Generic Masters (units/brands/systems/etc.) ----------------------
@api.get("/masters")
async def list_masters(authorization: Optional[str] = Header(None)):
    await get_current_user(authorization)
    items = await db.masters.find({"active": True}, {"_id": 0}).to_list(2000)
    grouped: Dict[str, List] = {"unit": [], "brand": [], "system": [], "material": [],
                                "model": [], "gst": [], "department": []}
    for it in items:
        grouped.setdefault(it["category"], []).append(it)
    grouped["system"] = grouped.get("system") or [{"name": s} for s in SYSTEMS]
    return grouped


@api.post("/masters")
async def add_master(m: MasterItem, authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("admin", "director"):
        raise HTTPException(403, "Only admin/director")
    doc = m.model_dump()
    doc["name"] = doc["name"].strip()
    if not doc["name"]:
        raise HTTPException(400, "Name required")
    if doc["category"] == "gst":
        try:
            float(doc.get("value") or "0")
        except Exception:
            raise HTTPException(400, "GST value must be a numeric percentage (e.g. 18)")
    dup = await db.masters.find_one({
        "category": doc["category"],
        "name": {"$regex": f"^{re.escape(doc['name'])}$", "$options": "i"},
        "active": True,
    })
    if dup:
        raise HTTPException(400, f"'{doc['name']}' already exists in {doc['category']}")
    await db.masters.insert_one(doc)
    doc.pop("_id", None)
    await master_audit(f"master.{doc['category']}", doc["item_id"], "create", u,
                        "Initial creation", None, doc)
    return doc


@api.put("/masters/{item_id}")
async def update_master(item_id: str, body: dict,
                         authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("admin", "director"):
        raise HTTPException(403, "Only admin/director")
    reason = (body.pop("reason", "") or "").strip()
    if not reason:
        raise HTTPException(400, "A reason is required for master-data changes")
    existing = await db.masters.find_one({"item_id": item_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Not found")
    updates: Dict[str, Any] = {}
    for k in ("name", "parent_id", "value"):
        if k in body:
            updates[k] = str(body[k] or "").strip() or None
    if "active" in body:
        updates["active"] = bool(body["active"])
    if not updates:
        raise HTTPException(400, "Nothing to update")
    await db.masters.update_one({"item_id": item_id}, {"$set": updates})
    old_snap = {k: existing.get(k) for k in updates}
    await master_audit(f"master.{existing['category']}", item_id, "update", u,
                        reason, old_snap, updates)
    return await db.masters.find_one({"item_id": item_id}, {"_id": 0})


@api.delete("/masters/{item_id}")
async def deactivate_master(item_id: str, reason: str = "Deactivated",
                              authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("admin", "director"):
        raise HTTPException(403, "Only admin/director")
    existing = await db.masters.find_one({"item_id": item_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Not found")
    await db.masters.update_one({"item_id": item_id}, {"$set": {"active": False}})
    await master_audit(f"master.{existing['category']}", item_id, "deactivate", u,
                        reason, {"active": True}, {"active": False})
    return {"ok": True}
