"""PO & bulk exports router — extracted from server.py in Phase 9C round 5 (Batch C).

Contains the five heaviest export endpoints in the codebase:
  - GET /api/po/{po_id}/pdf         — Vasu-branded PO PDF with CGST/SGST/IGST split
  - GET /api/export/mrf             — bulk MRF Excel workbook
  - GET /api/export/po              — bulk PO Excel workbook (line-item)
  - GET /api/export/tally           — bulk Tally voucher (kind=purchase | invoice)
  - GET /api/po/{po_id}/export      — unified single-record dispatcher (pdf|excel|tally)

All heavy layout / workbook helpers stay in server.py and are imported here:
  _po_excel_workbook, _po_tally_workbook, _pdf_story, _vasu_footer,
  _validate_po_for_export, _split_gst_amt, _is_intra_state,
  _log_export, _validation_error, _excel_response, _resolve_material_info,
  _safe_cell, VASU_PRIMARY, VASU_STATE, VASU_STATE_CODE.

Router endpoint bodies are copied byte-identical from the pre-refactor server.py
so behaviour, RBAC and audit rows remain unchanged.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any

from fastapi import Header, HTTPException
from fastapi.responses import StreamingResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)

from server import (
    api, db, get_current_user, now_utc,
    _check_po_access, _resolve_material_info,
    _pdf_story, _vasu_footer, _log_export, _validation_error, _excel_response,
    _validate_po_for_export, _split_gst_amt, _is_intra_state,
    _po_excel_workbook, _po_tally_workbook, _mrf_pdf_bytes,
    _validate_mrf_for_export, _safe_cell,
    VASU_PRIMARY, VASU_STATE, VASU_STATE_CODE, VASU_GSTIN,
    _ALLOWED_PO_FORMATS, _ALLOWED_MRF_FORMATS,
    bearer_from_url_token,
)


@api.get("/po/{po_id}/pdf")
async def po_pdf(po_id: str, token: Optional[str] = None,
                 force: Optional[int] = 0,
                 authorization: Optional[str] = Header(None)):
    # Allow token via query for browser download
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    po = await db.pos.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(404, "PO not found")
    await _check_po_access(u, po)
    vendor = await db.vendors.find_one({"vendor_id": po["vendor_id"]}, {"_id": 0}) or {}
    project = await db.projects.find_one({"project_id": po["project_id"]}, {"_id": 0}) or {}
    po_mrf_nums: List[str] = []
    for mid in po.get("mrf_refs") or []:
        m = await db.mrfs.find_one({"mrf_id": mid}, {"_id": 0, "mrf_number": 1})
        if m: po_mrf_nums.append(m["mrf_number"])

    # Customer snapshot (fall back to project's customer)
    cid = po.get("customer_id") or project.get("customer_id")
    cust = await db.customers.find_one({"customer_id": cid}, {"_id": 0}) if cid else None

    # Validate mandatory data
    missing, warnings = _validate_po_for_export(po, vendor, cust, project, "pdf")
    if missing and not force:
        raise _validation_error(missing, warnings, "pdf")

    # Enrich each PO item with MAT/VAR/HSN
    enriched: List[dict] = []
    for it in po.get("items", []):
        info = await _resolve_material_info(it)
        row = dict(it)
        row.update(info)
        enriched.append(row)

    # Resolve customer PO reference (latest active CPO on the customer)
    cpo_ref = ""
    if cust and cust.get("customer_id"):
        cpos = await db.customers.find_one({"customer_id": cust["customer_id"]},
                                            {"_id": 0, "pos": 1})
        if cpos and isinstance(cpos.get("pos"), list) and cpos["pos"]:
            active = [p for p in cpos["pos"] if p.get("active", True)]
            active.sort(key=lambda p: p.get("po_date") or "", reverse=True)
            if active:
                cpo_ref = f"{active[0].get('po_number','')} dt {active[0].get('po_date','')}".strip()

    # Determine intra-state for CGST/SGST vs IGST
    intra = _is_intra_state((vendor or {}).get("address") or VASU_STATE)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    extra_ref_parts = []
    if po_mrf_nums:
        extra_ref_parts.append("MRF: " + ", ".join(po_mrf_nums))
    if cpo_ref:
        extra_ref_parts.append("Customer PO: " + cpo_ref)
    if po.get("vendor_quotation"):
        extra_ref_parts.append("Vendor Quote: " + str(po["vendor_quotation"]))
    story = _pdf_story("PURCHASE ORDER", po["po_number"], po["date"], project, cust, vendor,
                       extra_ref="  ·  ".join(extra_ref_parts))

    styles = getSampleStyleSheet()
    small = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, leading=10)
    smallb = ParagraphStyle('sb', parent=styles['Normal'], fontSize=8, leading=10,
                            fontName="Helvetica-Bold")

    # Items table with MAT/VAR/HSN/Make/Model + GST breakdown
    header = ["#", "MAT / VAR", "Description", "Make / Model", "HSN", "Unit",
              "Qty", "Rate", "Disc", "Taxable", "GST %", "Total"]
    data = [[Paragraph(h, smallb) for h in header]]
    subtotal_taxable = 0.0
    total_cgst = total_sgst = total_igst = 0.0
    for idx, it in enumerate(enriched, 1):
        qty = float(it.get("qty") or 0)
        rate = float(it.get("rate") or 0)
        disc = float(it.get("discount") or 0)
        gst_pct = float(it.get("gst") or it.get("gst_pct") or 0)
        taxable = qty * rate - disc
        cgst, sgst, igst = _split_gst_amt(taxable, gst_pct, intra)
        line_total = round(taxable + cgst + sgst + igst, 2)
        subtotal_taxable += taxable
        total_cgst += cgst
        total_sgst += sgst
        total_igst += igst
        mat_var = ((it.get("material_uid") or "-") + "\n" +
                   (it.get("variant_uid") or "-"))
        make_model = ((it.get("make") or "-") + "\n" +
                      (it.get("model") or "-"))
        data.append([
            Paragraph(str(idx), small),
            Paragraph(mat_var, small),
            Paragraph((it.get("description") or "")[:80], small),
            Paragraph(make_model, small),
            Paragraph(it.get("hsn_code") or "—", small),
            Paragraph(it.get("unit") or "", small),
            Paragraph(f"{qty:g}", small),
            Paragraph(f"{rate:,.2f}", small),
            Paragraph(f"{disc:,.2f}" if disc else "—", small),
            Paragraph(f"{taxable:,.2f}", small),
            Paragraph(f"{gst_pct:g}%", small),
            Paragraph(f"{line_total:,.2f}", small),
        ])
    tbl = Table(data, colWidths=[16, 52, 130, 56, 40, 24, 26, 40, 32, 46, 26, 44], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), VASU_PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor('#B7BEC9')),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F8FC')]),
        ("ALIGN", (5, 0), (-1, -1), "RIGHT"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 8))

    # Tax summary block
    freight = float(po.get('freight') or 0)
    other = float(po.get('other_charges') or 0)
    grand = round(subtotal_taxable + total_cgst + total_sgst + total_igst + freight + other, 2)
    sum_rows = [
        [Paragraph("<b>Taxable Value</b>", small), Paragraph(f"₹ {subtotal_taxable:,.2f}", small)],
    ]
    if intra:
        sum_rows.append([Paragraph("CGST", small), Paragraph(f"₹ {total_cgst:,.2f}", small)])
        sum_rows.append([Paragraph("SGST", small), Paragraph(f"₹ {total_sgst:,.2f}", small)])
    else:
        sum_rows.append([Paragraph("IGST", small), Paragraph(f"₹ {total_igst:,.2f}", small)])
    sum_rows.append([Paragraph("Freight", small), Paragraph(f"₹ {freight:,.2f}", small)])
    sum_rows.append([Paragraph("Other Charges", small), Paragraph(f"₹ {other:,.2f}", small)])
    sum_rows.append([Paragraph("<b>Grand Total</b>", smallb),
                     Paragraph(f"<b>₹ {grand:,.2f}</b>", smallb)])
    sum_tbl = Table(sum_rows, colWidths=[380, 130])
    sum_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.4, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor('#D5D8DE')),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor('#EEF2FF')),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 10))

    # Terms & Authorisation block
    story.append(Paragraph(f"<b>Delivery Schedule:</b> {po.get('delivery_schedule','—') or '—'}", small))
    story.append(Paragraph(f"<b>Payment Terms:</b> {po.get('payment_terms','—') or '—'}", small))
    story.append(Paragraph(f"<b>Warranty:</b> {po.get('warranty_terms','—') or '—'}", small))
    if po.get("vendor_quotation"):
        story.append(Paragraph(f"<b>Vendor Quotation Ref:</b> {po['vendor_quotation']}", small))
    story.append(Spacer(1, 6))

    # Authorised signatories — from approval history (GM/Director) + explicit signatory
    signers: List[str] = []
    for h in (po.get("approval_history") or []):
        if (h.get("action") or "").lower() == "approve":
            role = (h.get("user_role") or "").upper()
            nm = h.get("user_name") or ""
            ts = (h.get("timestamp") or "")[:10]
            signers.append(f"{nm} ({role}) · {ts}")
    if po.get("authorised_signatory"):
        signers.append(str(po["authorised_signatory"]))
    if not signers:
        signers.append("For Vasu Infosec Pvt Ltd — Authorised Signatory")
    story.append(Paragraph("<b>Authorisation</b>", smallb))
    for s in signers:
        story.append(Paragraph(f"• {s}", small))

    # Warnings block (soft — printed at bottom)
    if warnings:
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            f"<font color='#B45309'><b>Advisory:</b> Optional fields missing — {', '.join(warnings[:8])}"
            + ("…" if len(warnings) > 8 else "") + "</font>", small))

    doc.build(story, onFirstPage=_vasu_footer, onLaterPages=_vasu_footer)
    buf.seek(0)
    await _log_export(u, "po", po_id, "pdf", po["po_number"],
                      {"warnings": warnings, "forced": bool(force and missing)})
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{po["po_number"].replace("/","_")}.pdf"'})

# ---------------------- Excel ----------------------


@api.get("/export/mrf")
async def export_mrf(token: Optional[str] = None,
                     authorization: Optional[str] = Header(None)):
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    if u.role not in ("purchase", "director", "admin", "pm", "gm"):
        raise HTTPException(403, "Not allowed")
    mrfs = await db.mrfs.find({"deleted": False}, {"_id": 0}).to_list(2000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MRFs"
    headers = [
        "MRF#", "Date", "Status", "Customer ID", "Customer Name",
        "Project", "Site", "Requester", "System",
        "Line#", "MAT UID", "VAR UID", "Description", "Make", "Model",
        "Unit", "Qty Requested", "Qty Approved", "Qty Ordered", "Qty Received",
        "Priority", "Item Status", "Billing Status", "Remarks",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    proj_map = {p["project_id"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(500)}
    for m in mrfs:
        proj = proj_map.get(m.get("project_id")) or {}
        for idx, it in enumerate(m.get("items") or [], 1):
            info = await _resolve_material_info(it)
            ws.append([
                _safe_cell(m.get("mrf_number", "")),
                m["date"].strftime("%Y-%m-%d") if isinstance(m.get("date"), datetime) else str(m.get("date", "")),
                _safe_cell(m.get("status", "")),
                _safe_cell(m.get("customer_id") or ""),
                _safe_cell(m.get("customer_name") or ""),
                _safe_cell(proj.get("code") or m.get("project_id", "")),
                _safe_cell(m.get("site", "")),
                _safe_cell(m.get("requesting_person", "")),
                _safe_cell(m.get("system_category", "")),
                idx,
                _safe_cell(info.get("material_uid") or ""),
                _safe_cell(info.get("variant_uid") or ""),
                _safe_cell(it.get("description", "")),
                _safe_cell(info.get("make") or ""),
                _safe_cell(info.get("model") or ""),
                _safe_cell(it.get("unit", "")),
                float(it.get("qty_requested") or 0),
                float(it.get("qty_approved") or 0),
                float(it.get("qty_ordered") or 0),
                float(it.get("qty_received") or 0),
                _safe_cell((it.get("priority") or "normal")),
                _safe_cell(it.get("status", "")),
                _safe_cell(it.get("billing_status", "")),
                _safe_cell(it.get("remarks", "")),
            ])
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(len(h) + 2, 28))
    await _log_export(u, "mrf", "bulk", "excel", f"count={len(mrfs)}")
    return _excel_response(wb, "mrf_export.xlsx")


@api.get("/export/po")
async def export_po(token: Optional[str] = None,
                    authorization: Optional[str] = Header(None)):
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    if u.role not in ("purchase", "director", "admin", "gm", "pm"):
        raise HTTPException(403, "Not allowed")
    pos = await db.pos.find({"deleted": False}, {"_id": 0}).sort("date", -1).to_list(2000)
    wb = await _po_excel_workbook(pos, single=False)
    await _log_export(u, "po", "bulk", "excel", f"count={len(pos)}")
    return _excel_response(wb, "po_export.xlsx")


@api.get("/export/tally")
async def export_tally(kind: str = "purchase", token: Optional[str] = None,
                       authorization: Optional[str] = Header(None)):
    """Bulk Tally purchase-voucher rows. kind=purchase | invoice."""
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    if u.role not in ("purchase", "director", "admin"):
        raise HTTPException(403, "Only purchase/director/admin")
    kind = (kind or "purchase").lower()
    if kind not in ("purchase", "invoice"):
        raise HTTPException(400, "kind must be 'purchase' or 'invoice'")

    if kind == "purchase":
        pos = await db.pos.find({"deleted": False}, {"_id": 0}).sort("date", 1).to_list(2000)
        wb = await _po_tally_workbook(pos)
        fname = "tally_purchase_voucher.xlsx"
        await _log_export(u, "tally", "bulk_purchase", "tally", f"count={len(pos)}")
        return _excel_response(wb, fname)

    # invoice-based purchase voucher (vendor invoices)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Voucher"
    headers = [
        "Voucher Date", "Voucher Type", "Voucher No", "Reference No", "Reference Date",
        "Company GSTIN", "Company State",
        "Party Ledger", "Party GSTIN", "Party State",
        "Customer ID", "Customer Name",
        "Item Name", "HSN/SAC", "Unit", "Quantity", "Rate", "Discount",
        "Taxable Value", "GST Rate %", "CGST", "SGST", "IGST",
        "Line Total", "Narration",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    vendors = {v["vendor_id"]: v for v in await db.vendors.find({}, {"_id": 0}).to_list(1000)}
    invs = await db.invoices.find({}, {"_id": 0}).sort("created_at", 1).to_list(2000)
    for inv in invs:
        vend = vendors.get(inv.get("vendor_id")) or {}
        vdate = str(inv.get("invoice_date", ""))[:10]
        intra = _is_intra_state(vend.get("address") or VASU_STATE)
        for it in inv.get("items", []):
            info = await _resolve_material_info(it)
            qty = float(it.get("qty") or 0)
            rate = float(it.get("rate") or 0)
            disc = float(it.get("discount") or 0)
            gst_pct = float(it.get("gst") or 0) or info.get("gst_pct") or 0
            taxable = round(qty * rate - disc, 2)
            cgst, sgst, igst = _split_gst_amt(taxable, gst_pct, intra)
            line_total = round(taxable + cgst + sgst + igst, 2)
            ws.append([
                _safe_cell(vdate), _safe_cell("Purchase"), _safe_cell(inv["invoice_number"]),
                _safe_cell(inv.get("vendor_invoice_number", "")), _safe_cell(vdate),
                _safe_cell(VASU_GSTIN), _safe_cell(VASU_STATE),
                _safe_cell(vend.get("name", "")),
                _safe_cell(vend.get("gstin", "")),
                _safe_cell((vend.get("address", "") or "")[:40]),
                _safe_cell(inv.get("customer_id", "")),
                _safe_cell(inv.get("customer_name", "")),
                _safe_cell(it.get("description", "")),
                _safe_cell(info.get("hsn_code") or ""),
                _safe_cell(it.get("unit", "")),
                qty, rate, disc,
                taxable, gst_pct, cgst, sgst, igst,
                line_total,
                _safe_cell(f"Vendor Inv {inv.get('vendor_invoice_number','')} · PO {inv.get('po_number','')}"),
            ])
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, min(len(h) + 2, 28))
    await _log_export(u, "tally", "bulk_invoice", "tally", f"count={len(invs)}")
    return _excel_response(wb, "tally_invoice_voucher.xlsx")


@api.get("/po/{po_id}/export")
async def po_export_dispatch(po_id: str,
                             format: str = "pdf",
                             force: Optional[int] = 0,
                             token: Optional[str] = None,
                             authorization: Optional[str] = Header(None)):
    """Unified download endpoint for a single PO. format = pdf | excel | tally."""
    fmt = (format or "pdf").lower()
    if fmt not in _ALLOWED_PO_FORMATS:
        raise HTTPException(400, f"format must be one of {_ALLOWED_PO_FORMATS}")
    auth = authorization or bearer_from_url_token(token)
    u = await get_current_user(auth)
    po = await db.pos.find_one({"po_id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(404, "PO not found")
    await _check_po_access(u, po)

    if fmt == "pdf":
        # Delegate to enriched po_pdf which validates internally
        return await po_pdf(po_id, token=token, force=force, authorization=authorization)

    # Excel / Tally single-record path — validate first
    vendor = await db.vendors.find_one({"vendor_id": po["vendor_id"]}, {"_id": 0}) or {}
    project = await db.projects.find_one({"project_id": po["project_id"]}, {"_id": 0}) or {}
    cid = po.get("customer_id") or project.get("customer_id")
    cust = await db.customers.find_one({"customer_id": cid}, {"_id": 0}) if cid else None
    missing, warnings = _validate_po_for_export(po, vendor, cust, project, fmt)
    if missing and not force:
        raise _validation_error(missing, warnings, fmt)

    if fmt == "excel":
        wb = await _po_excel_workbook([po], single=True)
        await _log_export(u, "po", po_id, "excel", po["po_number"],
                          {"warnings": warnings, "forced": bool(force and missing)})
        return _excel_response(wb, f"{po['po_number'].replace('/', '_')}_details.xlsx")

    # tally
    wb = await _po_tally_workbook([po])
    await _log_export(u, "po", po_id, "tally", po["po_number"],
                      {"warnings": warnings, "forced": bool(force and missing)})
    return _excel_response(wb, f"{po['po_number'].replace('/', '_')}_tally.xlsx")


