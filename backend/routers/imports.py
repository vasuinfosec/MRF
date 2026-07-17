"""Imports router — extracted from server.py in Phase 9C round 5 (Batch C).

Two flows:
  1. Vendor bulk import (Excel .xlsx / .xlsm) — admin or purchase only.
  2. MRF bulk import — groups rows by (project_code, site, required_by,
     requester, system_category) and creates one MRF per group.
Plus two template endpoints returning ready-to-fill Excel skeletons.
(Material bulk import lives in `routers/masters.py`.)
"""
from __future__ import annotations

import io
from typing import Optional, Dict, Any, List
from fastapi import Header, HTTPException, File, UploadFile

import openpyxl
from openpyxl.styles import Font, PatternFill

from server import (
    api, db, get_current_user, audit, gid,
    _excel_response,
    next_mrf_number,
    MRF, MRFItem,
    MAX_UPLOAD_BYTES,
)


@api.get("/import/vendors/template")
async def vendor_template(token: Optional[str] = None,
                            authorization: Optional[str] = Header(None)):
    auth = authorization or (f"Bearer {token}" if token else None)
    await get_current_user(auth)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendors"
    cols = ["name", "address", "gstin", "contact", "email"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
    ws.append(["Sample Vendor Pvt Ltd", "Bangalore", "29AAACS0000A1Z5",
                "9800000000", "sales@sample.com"])
    return _excel_response(wb, "vendor_template.xlsx")


@api.get("/import/mrf/template")
async def mrf_template(token: Optional[str] = None,
                        authorization: Optional[str] = Header(None)):
    auth = authorization or (f"Bearer {token}" if token else None)
    await get_current_user(auth)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MRF"
    cols = ["project_code", "site", "required_by", "requesting_person",
            "system_category", "description", "specification", "part_number",
            "unit", "qty_requested", "purpose", "drawing_ref", "billable",
            "boq_ref", "remarks"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="002FA7")
    ws.append(["VIS-101", "Bangalore Whitefield", "2026-05-15", "Ravi Kumar",
                "Fire Alarm", "Smoke Detector Photoelectric", "Honeywell 5251E",
                "5251E", "Nos", 25, "Level 1 Common Area", "FA-DWG-01", "Y",
                "BOQ-05", "Urgent"])
    return _excel_response(wb, "mrf_template.xlsx")


@api.post("/import/vendors")
async def import_vendors(file: UploadFile = File(...),
                          authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("admin", "purchase"):
        raise HTTPException(403, "Only admin/purchase")
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"File too large (max {MAX_UPLOAD_BYTES // (1024*1024)}MB)")
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
        raise HTTPException(400, "Only .xlsx / .xlsm allowed")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid xlsx: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Empty sheet")
    header = [str(c or "").strip().lower() for c in rows[0]]
    inserted, skipped, errors = 0, 0, []
    for idx, row in enumerate(rows[1:], start=2):
        rec = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
        name = (rec.get("name") or "").strip() if rec.get("name") else ""
        if not name:
            skipped += 1; continue
        existing = await db.vendors.find_one({"name": name}, {"_id": 0})
        if existing:
            skipped += 1; continue
        try:
            v = {
                "vendor_id": gid("vnd"), "name": name,
                "address": str(rec.get("address") or ""),
                "gstin": str(rec.get("gstin") or ""),
                "contact": str(rec.get("contact") or ""),
                "email": str(rec.get("email") or ""),
                "active": True,
            }
            await db.vendors.insert_one(v)
            inserted += 1
            await audit("vendor", v["vendor_id"], "import", u, {"name": name})
        except Exception as e:
            errors.append(f"row {idx}: {e}")
    return {"inserted": inserted, "skipped": skipped, "errors": errors}


@api.post("/import/mrf")
async def import_mrf(file: UploadFile = File(...),
                      authorization: Optional[str] = Header(None)):
    u = await get_current_user(authorization)
    if u.role not in ("admin", "site_engineer", "pm", "director"):
        raise HTTPException(403,
            "Only admin/site engineer/project manager can import MRFs")
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"File too large (max {MAX_UPLOAD_BYTES // (1024*1024)}MB)")
    fname = (file.filename or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xlsm")):
        raise HTTPException(400, "Only .xlsx / .xlsm allowed")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid xlsx: {e}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Empty sheet")
    header = [str(c or "").strip().lower() for c in rows[0]]

    projects = {p["code"]: p for p in await db.projects.find({}, {"_id": 0}).to_list(500)}
    groups: Dict[tuple, List[Dict[str, Any]]] = {}
    errors: List[str] = []
    for idx, row in enumerate(rows[1:], start=2):
        rec = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
        code = str(rec.get("project_code") or "").strip()
        if not code or code not in projects:
            errors.append(f"row {idx}: unknown project_code '{code}'")
            continue
        if not rec.get("description"):
            errors.append(f"row {idx}: description required")
            continue
        key = (code, str(rec.get("site") or ""),
                str(rec.get("required_by") or ""),
                str(rec.get("requesting_person") or ""),
                str(rec.get("system_category") or "Other"))
        item = {
            "description": str(rec.get("description")),
            "specification": str(rec.get("specification") or ""),
            "part_number": str(rec.get("part_number") or ""),
            "unit": str(rec.get("unit") or "Nos"),
            "qty_requested": float(rec.get("qty_requested") or 0),
            "purpose": str(rec.get("purpose") or ""),
            "drawing_ref": str(rec.get("drawing_ref") or ""),
            "billable": str(rec.get("billable") or "Y").strip().upper() != "N",
            "boq_ref": str(rec.get("boq_ref") or ""),
            "remarks": str(rec.get("remarks") or ""),
        }
        groups.setdefault(key, []).append(item)

    created: List[str] = []
    for key, items in groups.items():
        code, site, required_by, requester, syscat = key
        proj = projects[code]
        mrf_items = []
        for it in items:
            d = dict(it)
            d["qty_approved"] = d["qty_requested"]
            mrf_items.append(MRFItem(**d))
        mrf = MRF(
            mrf_number=await next_mrf_number(),
            project_id=proj["project_id"],
            site=site or proj.get("site", ""),
            required_by=required_by,
            requesting_person=requester,
            system_category=syscat,
            items=mrf_items,
            attachments=[],
            remarks="Imported from Excel",
            status="draft",
            created_by=u.user_id,
        )
        await db.mrfs.insert_one(mrf.model_dump())
        await audit("mrf", mrf.mrf_id, "import", u,
                     {"mrf_number": mrf.mrf_number, "items": len(mrf_items)})
        created.append(mrf.mrf_number)

    return {"created": created, "errors": errors, "count": len(created)}
